import json
import logging
import os
import re
import shutil
import subprocess
import tempfile
import time
import uuid
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Callable, Dict, List, Tuple

import openpyxl
import xlrd

from app.storage.database import Database
from app.core.rule_snapshot_service import RuleSnapshotService
from app.core.failure_reason_service import FailureReasonService
from app.core.rule_selection_service import RuleSelectionService


class SmartPurchaseService:
    
    FIELD_ALIASES = {
        "item_code": ["商品编码", "药品编码", "编码"],
        "source_name": ["商品名称", "品名", "药品名称", "通用名"],
        "source_spec": ["规格规格", "规格 规格", "规格", "商品规格"],
        "source_maker": ["生产厂家", "厂家", "生产企业"],
        "source_approval": ["批准文号", "国药准字"],
        "purchase_quantity": ["采购数量", "商品数量", "数量", "计划采购数量", "实际采购数量"],
        "expected_price": ["期望价格", "期望价", "最高价", "最高允许价"],
        "smart_name": ["药师帮商品名称", "药师帮品名"],
        "smart_spec": ["药师帮规格"],
        "smart_approval": ["药师帮批准文号"],
        "smart_supplier": ["商家", "供应商", "药师帮供应商"],
        "smart_supplier_full": ["企业名称", "供应商全称"],
        "min_purchase_quantity": ["起购数量", "起订量"],
        "available_stock": ["可用库存", "库存"],
        "actual_purchase_quantity": ["实际采购数量"],
        "smart_price": ["药师帮采购价", "采购价", "价格", "单价"],
        "selected": ["是否选中", "选中"],
        "activity_type": ["活动类型", "活动"],
        "ysb_code": ["药师帮编码", "wholesaleId", "商品ID"],
        "barcode": ["条形码", "条码"],
    }
    
    REQUIRED_FIELDS = ["source_name", "purchase_quantity"]
    
    def __init__(self, db: Database):
        self.db = db
        self._rule_snapshot_service = RuleSnapshotService(db)
        self._failure_reason_service = FailureReasonService(db)
        self._rule_selection_service = RuleSelectionService(db)
        self._ensure_tables()
    
    def _ensure_tables(self):
        conn = self.db.get_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS smart_purchase_batches (
                batch_id TEXT PRIMARY KEY,
                batch_name TEXT,
                source_file TEXT,
                sheet_name TEXT,
                supplier_scope TEXT,
                allow_keep_cart INTEGER DEFAULT 1,
                total_count INTEGER DEFAULT 0,
                valid_count INTEGER DEFAULT 0,
                invalid_count INTEGER DEFAULT 0,
                status TEXT DEFAULT 'imported',
                created_by TEXT,
                imported_at TEXT,
                created_at TEXT,
                rule_set_code TEXT,
                rule_set_version TEXT,
                rule_select_mode TEXT DEFAULT 'default',
                rule_select_reason TEXT,
                rule_snapshot_id TEXT,
                rule_selected_by TEXT,
                rule_selected_at TEXT
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS smart_purchase_items (
                item_id TEXT PRIMARY KEY,
                batch_id TEXT NOT NULL,
                row_number INTEGER NOT NULL,
                business_key TEXT,
                item_code TEXT,
                source_name TEXT,
                source_spec TEXT,
                source_maker TEXT,
                source_approval TEXT,
                purchase_quantity TEXT,
                expected_price TEXT,
                smart_name TEXT,
                smart_spec TEXT,
                smart_approval TEXT,
                smart_supplier TEXT,
                smart_supplier_full TEXT,
                min_purchase_quantity TEXT,
                available_stock TEXT,
                actual_purchase_quantity TEXT,
                smart_price TEXT,
                selected TEXT,
                activity_type TEXT,
                ysb_code TEXT,
                actual_ysb_code TEXT,
                barcode TEXT,
                raw_data TEXT,
                normalized_data TEXT,
                import_status TEXT DEFAULT 'valid',
                validation_message TEXT,
                purchase_status TEXT DEFAULT 'pending',
                purchase_supplier TEXT,
                purchase_product TEXT,
                purchase_spec TEXT,
                purchase_maker TEXT,
                purchase_valid_date TEXT,
                purchase_quantity_result TEXT,
                purchase_price TEXT,
                max_allowed_price TEXT,
                purchase_reason TEXT,
                candidate_count INTEGER DEFAULT 0,
                executed_at TEXT,
                failure_stage TEXT,
                failure_code TEXT,
                created_at TEXT,
                updated_at TEXT
            )
        ''')
        
        cursor.execute('''
            CREATE UNIQUE INDEX IF NOT EXISTS idx_smart_purchase_batch_row
            ON smart_purchase_items(batch_id, row_number)
        ''')
        
        cursor.execute('''
            CREATE INDEX IF NOT EXISTS idx_smart_purchase_batch
            ON smart_purchase_items(batch_id)
        ''')
        
        self._ensure_item_columns(cursor)
        
        conn.commit()
    
    def _ensure_item_columns(self, cursor):
        cursor.execute("PRAGMA table_info(smart_purchase_items)")
        columns = {row[1] for row in cursor.fetchall()}
        if "actual_ysb_code" not in columns:
            cursor.execute("ALTER TABLE smart_purchase_items ADD COLUMN actual_ysb_code TEXT")
        if "purchase_valid_date" not in columns:
            cursor.execute("ALTER TABLE smart_purchase_items ADD COLUMN purchase_valid_date TEXT")
        
        # 四期整改：购物车反写状态独立设计
        cart_write_columns = {
            "cart_write_status": "TEXT DEFAULT 'NOT_STARTED'",  # 购物车反写状态
            "cart_write_time": "TEXT",  # 购物车反写时间
            "cart_write_message": "TEXT",  # 购物车反写信息
            "cart_item_id": "TEXT",  # 购物车商品标识
            "cart_write_attempts": "INTEGER DEFAULT 0",  # 购物车反写次数
        }
        for col_name, col_def in cart_write_columns.items():
            if col_name not in columns:
                cursor.execute(f"ALTER TABLE smart_purchase_items ADD COLUMN {col_name} {col_def}")
    
    def _is_xls_file(self, file_path: str) -> bool:
        return Path(file_path).suffix.lower() == ".xls"
    
    def get_sheets(self, file_path: str) -> Tuple[List[str], str]:
        try:
            if self._is_xls_file(file_path):
                workbook = xlrd.open_workbook(file_path)
                sheets = workbook.sheet_names()
                return sheets, ""
            else:
                workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                sheets = workbook.sheetnames
                workbook.close()
                return sheets, ""
        except Exception as e:
            logging.error(f"读取智能采购Excel工作表失败: {e}")
            return [], str(e)
    
    def read_preview(
        self,
        file_path: str,
        sheet_name: str = "",
        max_rows: int = 100
    ) -> Tuple[List[str], List[List[str]], str]:
        try:
            if self._is_xls_file(file_path):
                workbook = xlrd.open_workbook(file_path)
                if not workbook.sheet_names():
                    return [], [], "Excel文件中没有工作表"
                
                target_sheet_name = sheet_name if sheet_name in workbook.sheet_names() else workbook.sheet_names()[0]
                sheet = workbook.sheet_by_name(target_sheet_name)
                
                headers = []
                for col in range(sheet.ncols):
                    value = sheet.cell_value(0, col)
                    headers.append(self._cell_to_text(value) or f"列{col + 1}")
                
                rows = []
                for row_idx in range(1, min(sheet.nrows, max_rows + 1)):
                    row_values = []
                    for col in range(len(headers)):
                        value = sheet.cell_value(row_idx, col)
                        row_values.append(self._cell_to_text(value))
                    if any(row_values):
                        rows.append(row_values)
                
                return headers, rows, ""
            else:
                workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                if not workbook.sheetnames:
                    workbook.close()
                    return [], [], "Excel文件中没有工作表"
                
                target_sheet = sheet_name if sheet_name in workbook.sheetnames else workbook.sheetnames[0]
                sheet = workbook[target_sheet]
                headers = self._read_headers(sheet)
                
                rows = []
                for row in sheet.iter_rows(min_row=2, max_row=min(sheet.max_row, max_rows + 1), values_only=True):
                    row_values = [self._cell_to_text(value) for value in row[:len(headers)]]
                    if any(row_values):
                        rows.append(row_values)
                
                workbook.close()
                return headers, rows, ""
        except Exception as e:
            logging.error(f"读取智能采购Excel预览失败: {e}")
            return [], [], str(e)
    
    def import_excel(
        self,
        file_path: str,
        sheet_name: str,
        supplier_scope: str = "",
        allow_keep_cart: bool = True,
        imported_by: str = "system"
    ) -> Tuple[str, int, int, str]:
        try:
            batch_id = f"SP{datetime.now().strftime('%Y%m%d%H%M%S')}{uuid.uuid4().hex[:4]}"
            now = datetime.now().isoformat()
            
            conn = self.db.get_connection()
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO smart_purchase_batches
                (batch_id, batch_name, source_file, sheet_name, supplier_scope,
                 allow_keep_cart, total_count, valid_count, invalid_count, status,
                 created_by, imported_at, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                batch_id,
                Path(file_path).name,
                file_path,
                sheet_name,
                supplier_scope,
                1 if allow_keep_cart else 0,
                0, 0, 0,
                "importing",
                imported_by,
                now,
                now
            ))
            
            total_count = 0
            valid_count = 0
            invalid_count = 0
            
            if self._is_xls_file(file_path):
                workbook = xlrd.open_workbook(file_path)
                if sheet_name not in workbook.sheet_names():
                    return "", 0, 0, f"工作表不存在: {sheet_name}"
                
                sheet = workbook.sheet_by_name(sheet_name)
                headers = []
                for col in range(sheet.ncols):
                    value = sheet.cell_value(0, col)
                    headers.append(self._cell_to_text(value) or f"列{col + 1}")
                
                for row_number in range(1, sheet.nrows):
                    row_values = []
                    for col in range(len(headers)):
                        value = sheet.cell_value(row_number, col)
                        row_values.append(value)
                    
                    row_data = self._row_to_dict_xls(headers, row_values)
                    if not any(str(value).strip() for value in row_data.values()):
                        continue
                    
                    total_count += 1
                    normalized = self._normalize_row(row_data)
                    errors = self._validate_row(normalized)
                    import_status = "invalid" if errors else "valid"
                    validation_message = "；".join(errors)
                    if errors:
                        invalid_count += 1
                    else:
                        valid_count += 1
                    
                    item_id = f"{batch_id}_{row_number + 1}"
                    business_key = (
                        normalized.get("item_code")
                        or normalized.get("source_approval")
                        or normalized.get("source_name")
                        or f"ROW_{row_number + 1}"
                    )
                    
                    cursor.execute('''
                        INSERT OR REPLACE INTO smart_purchase_items
                        (item_id, batch_id, row_number, business_key, item_code,
                         source_name, source_spec, source_maker, source_approval,
                         purchase_quantity, expected_price, smart_name, smart_spec,
                         smart_approval, smart_supplier, smart_supplier_full,
                         min_purchase_quantity, available_stock, actual_purchase_quantity,
                         smart_price, selected, activity_type, ysb_code, barcode,
                         raw_data, normalized_data, import_status, validation_message,
                         purchase_status, created_at, updated_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        item_id,
                        batch_id,
                        row_number + 1,
                        business_key,
                        normalized.get("item_code", ""),
                        normalized.get("source_name", ""),
                        normalized.get("source_spec", ""),
                        normalized.get("source_maker", ""),
                        normalized.get("source_approval", ""),
                        normalized.get("purchase_quantity", ""),
                        normalized.get("expected_price", ""),
                        normalized.get("smart_name", ""),
                        normalized.get("smart_spec", ""),
                        normalized.get("smart_approval", ""),
                        normalized.get("smart_supplier", ""),
                        normalized.get("smart_supplier_full", ""),
                        normalized.get("min_purchase_quantity", ""),
                        normalized.get("available_stock", ""),
                        normalized.get("actual_purchase_quantity", ""),
                        normalized.get("smart_price", ""),
                        normalized.get("selected", ""),
                        normalized.get("activity_type", ""),
                        normalized.get("ysb_code", ""),
                        normalized.get("barcode", ""),
                        json.dumps(row_data, ensure_ascii=False),
                        json.dumps(normalized, ensure_ascii=False),
                        import_status,
                        validation_message,
                        "pending" if import_status == "valid" else "invalid",
                        now,
                        now
                    ))
            else:
                workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                if sheet_name not in workbook.sheetnames:
                    workbook.close()
                    return "", 0, 0, f"工作表不存在: {sheet_name}"
                
                sheet = workbook[sheet_name]
                headers = self._read_headers(sheet)
                
                for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    row_data = self._row_to_dict(headers, row)
                    if not any(str(value).strip() for value in row_data.values()):
                        continue
                    
                    total_count += 1
                    normalized = self._normalize_row(row_data)
                    errors = self._validate_row(normalized)
                    import_status = "invalid" if errors else "valid"
                    validation_message = "；".join(errors)
                    if errors:
                        invalid_count += 1
                    else:
                        valid_count += 1
                    
                    item_id = f"{batch_id}_{row_number}"
                    business_key = (
                        normalized.get("item_code")
                        or normalized.get("source_approval")
                        or normalized.get("source_name")
                        or f"ROW_{row_number}"
                    )
                    
                    cursor.execute('''
                        INSERT OR REPLACE INTO smart_purchase_items
                        (item_id, batch_id, row_number, business_key, item_code,
                         source_name, source_spec, source_maker, source_approval,
                         purchase_quantity, expected_price, smart_name, smart_spec,
                         smart_approval, smart_supplier, smart_supplier_full,
                         min_purchase_quantity, available_stock, actual_purchase_quantity,
                         smart_price, selected, activity_type, ysb_code, barcode,
                         raw_data, normalized_data, import_status, validation_message,
                         purchase_status, created_at, updated_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        item_id,
                        batch_id,
                        row_number,
                        business_key,
                        normalized.get("item_code", ""),
                        normalized.get("source_name", ""),
                        normalized.get("source_spec", ""),
                        normalized.get("source_maker", ""),
                        normalized.get("source_approval", ""),
                        normalized.get("purchase_quantity", ""),
                        normalized.get("expected_price", ""),
                        normalized.get("smart_name", ""),
                        normalized.get("smart_spec", ""),
                        normalized.get("smart_approval", ""),
                        normalized.get("smart_supplier", ""),
                        normalized.get("smart_supplier_full", ""),
                        normalized.get("min_purchase_quantity", ""),
                        normalized.get("available_stock", ""),
                        normalized.get("actual_purchase_quantity", ""),
                        normalized.get("smart_price", ""),
                        normalized.get("selected", ""),
                        normalized.get("activity_type", ""),
                        normalized.get("ysb_code", ""),
                        normalized.get("barcode", ""),
                        json.dumps(row_data, ensure_ascii=False),
                        json.dumps(normalized, ensure_ascii=False),
                        import_status,
                        validation_message,
                        "pending" if import_status == "valid" else "invalid",
                        now,
                        now
                    ))
                
                workbook.close()
            
            cursor.execute('''
                UPDATE smart_purchase_batches
                SET total_count = ?, valid_count = ?, invalid_count = ?, status = ?
                WHERE batch_id = ?
            ''', (total_count, valid_count, invalid_count, "imported", batch_id))
            
            conn.commit()
            return batch_id, valid_count, invalid_count, ""
        except Exception as e:
            logging.error(f"导入智能采购Excel失败: {e}")
            try:
                self.db.get_connection().rollback()
            except Exception:
                pass
            return "", 0, 0, str(e)
    
    def get_batches(self) -> List[Dict]:
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT batch_id, batch_name, source_file, sheet_name, supplier_scope,
                   allow_keep_cart, total_count, valid_count, invalid_count,
                   status, imported_at
            FROM smart_purchase_batches
            ORDER BY imported_at DESC
        ''')
        return [dict(row) for row in cursor.fetchall()]
    
    def get_batch_items(
        self,
        batch_id: str,
        status_filter: str = "all",
        keyword: str = ""
    ) -> List[Dict]:
        sql = '''
            SELECT item_id, batch_id, row_number, business_key, item_code,
                   source_name, source_spec, source_maker, source_approval,
                   purchase_quantity, expected_price, smart_supplier,
                   smart_supplier_full, smart_price, ysb_code, actual_ysb_code, import_status,
                   validation_message, purchase_status, purchase_supplier,
                   purchase_product, purchase_spec, purchase_maker,
                   purchase_quantity_result, purchase_price, purchase_reason,
                   candidate_count, executed_at, activity_type,
                   cart_write_status, cart_write_time, cart_write_message,
                   cart_item_id, cart_write_attempts
            FROM smart_purchase_items
            WHERE batch_id = ?
        '''
        params = [batch_id]
        
        if status_filter != "all":
            sql += " AND purchase_status = ?"
            params.append(status_filter)
        
        if keyword:
            sql += '''
                AND (
                    business_key LIKE ? OR item_code LIKE ? OR source_name LIKE ?
                    OR source_spec LIKE ? OR source_maker LIKE ? OR source_approval LIKE ?
                    OR smart_supplier LIKE ? OR smart_supplier_full LIKE ?
                )
            '''
            like_keyword = f"%{keyword}%"
            params.extend([like_keyword] * 8)
        
        sql += " ORDER BY row_number"
        
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute(sql, params)
        return [dict(row) for row in cursor.fetchall()]
    
    def delete_batch(self, batch_id: str) -> Tuple[bool, str]:
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            cursor.execute("DELETE FROM smart_purchase_items WHERE batch_id = ?", (batch_id,))
            cursor.execute("DELETE FROM smart_purchase_batches WHERE batch_id = ?", (batch_id,))
            conn.commit()
            return True, ""
        except Exception as e:
            logging.error(f"删除智能采购批次失败: {e}")
            return False, str(e)
    
    def _clear_purchase_backfill_history(
        self,
        batch_id: str,
        remove_cart_extra: bool = True,
        preserve_purchase_result: bool = False,
    ) -> int:
        conn = self.db.get_connection()
        cursor = conn.cursor()
        removed = 0
        if remove_cart_extra:
            cursor.execute('''
                DELETE FROM smart_purchase_items
                WHERE batch_id = ?
                  AND (business_key LIKE 'CART_EXTRA_%' OR activity_type = 'cart_extra')
            ''', (batch_id,))
            removed = cursor.rowcount if cursor.rowcount is not None else 0
        if preserve_purchase_result:
            cursor.execute('''
                UPDATE smart_purchase_items
                SET cart_write_status = 'NOT_STARTED',
                    cart_write_time = '',
                    cart_write_message = '',
                    cart_item_id = '',
                    cart_write_attempts = 0,
                    updated_at = ?
                WHERE batch_id = ?
                  AND import_status = 'valid'
            ''', (datetime.now().isoformat(), batch_id))
        else:
            cursor.execute('''
                UPDATE smart_purchase_items
                SET purchase_status = 'pending',
                    purchase_supplier = '',
                    purchase_product = '',
                    purchase_spec = '',
                    purchase_maker = '',
                    purchase_quantity_result = '',
                    actual_purchase_quantity = '',
                    purchase_price = '',
                    max_allowed_price = '',
                    purchase_reason = '',
                    actual_ysb_code = '',
                    candidate_count = 0,
                    executed_at = '',
                    cart_write_status = 'NOT_STARTED',
                    cart_write_time = '',
                    cart_write_message = '',
                    cart_item_id = '',
                    cart_write_attempts = 0,
                    updated_at = ?
                WHERE batch_id = ?
                  AND import_status = 'valid'
            ''', (datetime.now().isoformat(), batch_id))
        conn.commit()
        if remove_cart_extra:
            self._refresh_batch_counts(batch_id)
        return removed

    def execute_batch_purchase(
        self,
        batch_id: str,
        retry_failed: bool = False,
        max_rows: int = 0,
        use_cart_adapter: bool = False
    ) -> Tuple[Dict, List[str], str]:
        try:
            batch = self.get_batch(batch_id)
            if not batch:
                return {}, [], "采购批次不存在"
            
            if not retry_failed:
                self._clear_purchase_backfill_history(batch_id, remove_cart_extra=True)
            items = self._get_executable_items(batch_id, retry_failed=retry_failed, max_rows=max_rows)
            summary = {
                "total": len(items),
                "success": 0,
                "failed": 0,
                "skipped": 0,
            }
            logs = []
            
            if not items:
                return summary, ["没有可执行的采购明细"], ""
            
            if use_cart_adapter and not any(str(item.get("ysb_code") or "").strip() for item in items):
                return summary, [
                    "当前批次没有任何药师帮编码，未执行逐个采购。",
                    "请先完成药师帮商品匹配，或导入包含“药师帮编码/wholesaleId/商品ID”的采购清单后再执行。",
                ], "当前批次没有药师帮编码，无法加入购物车"
            
            supplier_scope = self._parse_supplier_scope(batch.get("supplier_scope", ""))
            self._use_cart_adapter = use_cart_adapter
            
            for item in items:
                result = self._execute_single_item(item, supplier_scope)
                self._save_purchase_result(item["item_id"], result, batch_id=batch_id)
                
                status = result["purchase_status"]
                if status == "success":
                    summary["success"] += 1
                elif status == "skipped":
                    summary["skipped"] += 1
                else:
                    summary["failed"] += 1
                
                logs.append(
                    f"第{item['row_number']}行 {item.get('source_name') or item.get('business_key')}: "
                    f"{self._status_text(status)} - {result['purchase_reason']}"
                )
            
            return summary, logs, ""
        except Exception as e:
            logging.error(f"执行智能采购批次失败: {e}")
            return {}, [], str(e)
    
    def execute_batch_purchase_real(
        self,
        batch_id: str,
        retry_failed: bool = False,
        max_rows: int = 0,
        use_cart_adapter: bool = True,
        progress_callback: Callable[[str], None] = None,
        web_error_callback: Callable[[str], bool] = None,
        pause_callback: Callable[[], None] = None
    ) -> Tuple[Dict, List[str], str]:
        try:
            batch = self.get_batch(batch_id)
            if not batch:
                # 增强批次不存在提示（方案要求：提示批次ID和批次状态统计）
                error_msg = f"采购批次不存在（批次ID: {batch_id}）。请检查批次是否已删除或批次ID是否正确。建议：请确认批次ID后再执行。"
                return {}, [], error_msg

            if use_cart_adapter:
                adapter_syntax_error = self._check_cart_adapter_syntax()
                if adapter_syntax_error:
                    return {}, [], f"购物车执行脚本语法检查失败，已停止本批次：{adapter_syntax_error}"
            
            if not retry_failed:
                self._clear_purchase_backfill_history(batch_id, remove_cart_extra=True)
            items = self._get_executable_items(batch_id, retry_failed=retry_failed, max_rows=max_rows)
            status_counts = self._get_batch_purchase_status_counts(batch_id)
            summary = {"total": len(items), "success": 0, "failed": 0, "skipped": 0}
            logs = []
            if not items:
                # 增强无可执行明细提示（方案要求：提示总行数、成功行数、失败行数、跳过行数）
                all_count = status_counts.get('all', 0)
                success_count = status_counts.get('success', 0)
                failed_count = status_counts.get('failed', 0)
                pending_count = status_counts.get('pending', 0)
                other_count = status_counts.get('other', 0)
                
                if retry_failed:
                    # 重试失败时，提示失败明细情况
                    error_msg = (
                        f"当前批次没有可重试的失败明细。\n"
                        f"批次状态统计：总行数 {all_count}，已成功 {success_count} 行，失败 {failed_count} 行，待执行 {pending_count} 行，其他 {other_count} 行。\n"
                        f"建议：如果失败明细已全部重试成功，请选择'开始逐个采购'而非'重试失败明细'。"
                    )
                else:
                    # 开始逐个采购时，提示批次状态
                    error_msg = (
                        f"当前批次没有可执行的采购明细。\n"
                        f"批次状态统计：总行数 {all_count}，已成功 {success_count} 行，失败 {failed_count} 行，待执行 {pending_count} 行，其他 {other_count} 行。\n"
                        f"建议：如果所有明细都已执行完成，请查看执行结果或导出采购结果。"
                    )
                
                return summary, [error_msg], ""
            
            trace_path = self._create_purchase_trace_path(batch_id)
            self._append_purchase_trace(
                trace_path,
                f"START batch={batch_id} total={len(items)} retry_failed={retry_failed} "
                f"max_rows={max_rows} use_cart_adapter={use_cart_adapter}",
                progress_callback
            )
            self._append_purchase_trace(
                trace_path,
                "BATCH_STATUS "
                f"all={status_counts.get('all', 0)} "
                f"pending={status_counts.get('pending', 0)} "
                f"failed={status_counts.get('failed', 0)} "
                f"success={status_counts.get('success', 0)} "
                f"other={status_counts.get('other', 0)} "
                f"executable={len(items)}",
                progress_callback
            )
            logs.append(f"本次逐个采购日志：{trace_path}")
            if progress_callback:
                progress_callback(f"本次逐个采购日志：{trace_path}")
            
            supplier_scope = self._parse_supplier_scope(batch.get("supplier_scope", ""))

            # 三期整改：通过 RuleSelectionService 解析规则集
            # 优先使用批次已保存的规则集，否则使用默认规则
            saved_rule_set_code = self._rule_selection_service.get_batch_rule_set_code(batch_id)
            rule_selection = self._rule_selection_service.resolve_rule_set(
                batch_id, manual_rule_set_code=saved_rule_set_code
            )
            resolved_rule_set_code = rule_selection["rule_set_code"]
            rule_select_mode = rule_selection["rule_select_mode"]

            # 保存规则选择结果到批次
            self._rule_selection_service.save_rule_selection_to_batch(batch_id, rule_selection)

            # 三期整改：将实际规则集传入快照生成
            rule_snapshot_id, snapshot_error = self._rule_snapshot_service.generate_rule_snapshot(
                batch_id, resolved_rule_set_code
            )
            if rule_snapshot_id:
                # 将快照ID回写批次
                conn = self.db.get_connection()
                cursor = conn.cursor()
                cursor.execute(
                    "UPDATE smart_purchase_batches SET rule_snapshot_id = ? WHERE batch_id = ?",
                    (rule_snapshot_id, batch_id)
                )
                conn.commit()
                self._append_purchase_trace(
                    trace_path,
                    f"RULE_SNAPSHOT_GENERATED snapshot_id={rule_snapshot_id} "
                    f"rule_set={resolved_rule_set_code} mode={rule_select_mode}",
                    progress_callback
                )
            else:
                self._append_purchase_trace(
                    trace_path,
                    f"RULE_SNAPSHOT_FAILED rule_set={resolved_rule_set_code} error={snapshot_error}",
                    progress_callback
                )
            
            # 获取规则快照配置（供Node使用）
            rule_config_for_node = {}
            if rule_snapshot_id:
                rule_config_for_node = self._rule_snapshot_service.get_rule_config_for_node(rule_snapshot_id)
            
            ready_items = []
            ready_results = {}
            
            for item_index, item in enumerate(items, start=1):
                if pause_callback and pause_callback():
                    skipped_count = len(items) - item_index + 1
                    summary["skipped"] += skipped_count
                    message = f"用户请求停止，剩余 {skipped_count} 条未执行"
                    self._append_purchase_trace(trace_path, f"STOP_REQUESTED stage=precheck remaining={skipped_count}", progress_callback)
                    logs.append(message)
                    return summary, logs, ""
                result = self._prepare_cart_purchase_item(item, supplier_scope)
                self._append_purchase_trace(
                    trace_path,
                    f"PRECHECK row={item.get('row_number')} item_id={item.get('item_id')} "
                    f"ysb_code={item.get('ysb_code')} name={item.get('source_name')} "
                    f"status={result.get('purchase_status')} ready={result.get('ready_for_cart')} "
                    f"reason={result.get('purchase_reason', '')}",
                    progress_callback
                )
                if result.get("ready_for_cart"):
                    adapter_item = self._build_cart_adapter_item(
                        item, result, supplier_scope,
                        rule_config=rule_config_for_node,
                        rule_snapshot_id=rule_snapshot_id
                    )
                    ready_items.append(adapter_item)
                    ready_results[item["item_id"]] = result
                    continue
                
                self._save_purchase_result(item["item_id"], result, batch_id=batch_id,
                                           rule_snapshot_id=rule_snapshot_id, rule_set_code=resolved_rule_set_code)
                status = result.get("purchase_status", "failed")
                summary[status if status in summary else "failed"] += 1
                logs.append(
                    f"第{item['row_number']}行 {item.get('source_name') or item.get('business_key')}: "
                    f"{self._status_text(status)} - {result.get('purchase_reason', '')}"
                )
            
            if ready_items:
                self._append_purchase_trace(
                    trace_path,
                    f"CART_ADAPTER_START ready_count={len(ready_items)} mode=one_by_one_writeback",
                    progress_callback
                )
                original_ready_items = list(ready_items)
                successful_cart_results = []
                for index, adapter_item in enumerate(original_ready_items, start=1):
                    if pause_callback and pause_callback():
                        skipped_count = len(original_ready_items) - index + 1
                        summary["skipped"] += skipped_count
                        message = f"用户请求停止，剩余 {skipped_count} 条未执行"
                        self._append_purchase_trace(trace_path, f"STOP_REQUESTED stage=cart remaining={skipped_count}", progress_callback)
                        logs.append(message)
                        return summary, logs, ""
                    self._append_purchase_trace(
                        trace_path,
                        f"CART_ADAPTER_ITEM_START index={index}/{len(original_ready_items)} "
                        f"row={adapter_item.get('rowNumber', '')} item_id={adapter_item.get('itemId', '')}",
                        progress_callback
                    )
                    if use_cart_adapter:
                        adapter_results = self._run_cart_adapter_batch(
                            batch_id,
                            [adapter_item],
                            trace_path,
                            progress_callback,
                            stop_callback=pause_callback,
                        )
                        adapter_results, batch_aborted, abort_reason = self._handle_web_error_retry(
                            batch_id,
                            [adapter_item],
                            adapter_results,
                            trace_path,
                            progress_callback,
                            web_error_callback,
                            stop_callback=pause_callback,
                        )
                    else:
                        adapter_results = self._adapter_failure_results([adapter_item], "未启用药师帮购物车真实加购，未执行加购")
                        batch_aborted = False
                        abort_reason = ""
                        self._append_purchase_trace(trace_path, "CART_ADAPTER_SKIPPED disabled", progress_callback)
                    if not adapter_results:
                        adapter_results = self._adapter_failure_results([adapter_item], "购物车执行器未返回本行结果")

                    for adapter_result in adapter_results:
                        item_id = adapter_result.get("itemId", "")
                        base_result = ready_results.get(item_id)
                        if not base_result:
                            self._append_purchase_trace(
                                trace_path,
                                f"CART_RESULT_IGNORED item_id={item_id} result={json.dumps(adapter_result, ensure_ascii=False)}",
                                progress_callback
                            )
                            continue
                        result = self._merge_cart_adapter_result(base_result, adapter_result)
                        self._save_purchase_result(item_id, result, batch_id=batch_id,
                                                   rule_snapshot_id=rule_snapshot_id, rule_set_code=resolved_rule_set_code)
                        status = result.get("purchase_status", "failed")
                        if use_cart_adapter and status in ("success", "skipped"):
                            successful_cart_results.append((item_id, adapter_item, adapter_result, result))
                        summary[status if status in summary else "failed"] += 1
                        display_reason = self._result_display_reason(result)
                        self._append_purchase_trace(
                            trace_path,
                            f"WRITE_RESULT row={adapter_result.get('rowNumber', '')} item_id={item_id} "
                            f"wholesale_id={adapter_result.get('wholesaleId', '')} status={status} "
                            f"actual_ysb_code={result.get('actual_ysb_code', '')} "
                            f"verified_amount={adapter_result.get('verifiedAmount', '')} "
                            f"match_source={adapter_result.get('matchSource', '')} "
                            f"match_score={adapter_result.get('matchScore', '')} "
                            f"match_reason={adapter_result.get('matchReason', '')} "
                            f"reason={display_reason}",
                            progress_callback
                        )
                        logs.append(
                            f"第{adapter_result.get('rowNumber', '')}行 {adapter_result.get('name', '')}: "
                            f"{self._status_text(status)} - {display_reason}"
                        )
                        if progress_callback:
                            progress_callback(
                                f"进度 {index}/{len(original_ready_items)} 行{adapter_result.get('rowNumber', '')}: "
                                f"{self._status_text(status)} - {display_reason}"
                            )
                    if batch_aborted:
                        remaining_count = len(original_ready_items) - index
                        summary["skipped"] += remaining_count
                        message = (
                            f"本批次已停止：{abort_reason or '药师帮网页异常'}；"
                            f"剩余 {remaining_count} 条保持未执行"
                        )
                        logs.append(message)
                        self._append_purchase_trace(
                            trace_path,
                            f"BATCH_ABORTED row={adapter_result.get('rowNumber', '')} "
                            f"remaining={remaining_count} reason={abort_reason}",
                            progress_callback,
                        )
                        return summary, logs, ""

                if use_cart_adapter and successful_cart_results:
                    self._reconcile_successful_cart_results(
                        batch_id,
                        successful_cart_results,
                        trace_path,
                        progress_callback,
                    )
            
            # 增强执行完成弹窗增加失败原因 Top 3（方案要求：统计失败原因分类，弹窗展示失败原因 Top 3 分类摘要）
            if summary.get("failed", 0) > 0:
                failure_reason_summary = self._get_failure_reason_summary(batch_id)
                if failure_reason_summary:
                    top3_message = "失败原因 Top 3 分类摘要：\n"
                    for index, (reason_type, count) in enumerate(failure_reason_summary[:3], start=1):
                        top3_message += f"{index}. {reason_type}：{count} 次\n"
                    logs.append(top3_message)
                    self._append_purchase_trace(trace_path, f"FAILURE_TOP3 {top3_message.strip()}", progress_callback)
                    if progress_callback:
                        progress_callback(top3_message.strip())
            
            self._append_purchase_trace(trace_path, f"FINISH summary={json.dumps(summary, ensure_ascii=False)}", progress_callback)
            return summary, logs, ""
        except Exception as e:
            logging.error(f"执行智能采购真实加购失败: {e}")
            return {}, [], str(e)
    
    def execute_cart_backfill(
        self,
        batch_id: str,
        progress_callback: Callable[[str], None] = None
    ) -> Tuple[Dict, List[str], str]:
        try:
            batch = self.get_batch(batch_id)
            if not batch:
                return {}, [], "采购批次不存在"
            
            trace_path = self._create_purchase_trace_path(batch_id)
            removed_extra = self._clear_purchase_backfill_history(
                batch_id,
                remove_cart_extra=True,
                preserve_purchase_result=True,
            )
            self._append_purchase_trace(trace_path, f"CART_BACKFILL_CLEAR removed_extra={removed_extra}", progress_callback)
            
            items = self._get_executable_items(batch_id, retry_failed=False, max_rows=0)
            summary = {"total": len(items), "updated": 0, "unmatched": 0, "extra": 0}
            logs = [f"已清除历史反写数据，删除旧购物车额外登记 {removed_extra} 条"]
            if not items:
                return summary, ["当前批次没有可反写的明细"], ""

            self._append_purchase_trace(trace_path, f"CART_BACKFILL_ONLY_START batch={batch_id} total={len(items)}", progress_callback)
            
            # 三期整改：获取批次的规则集和快照ID
            resolved_rule_set_code = self._rule_selection_service.get_batch_rule_set_code(batch_id) or "default_v1"
            conn = self.db.get_connection()
            cursor = conn.cursor()
            cursor.execute(
                "SELECT rule_snapshot_id FROM smart_purchase_batches WHERE batch_id = ?",
                (batch_id,)
            )
            batch_row = cursor.fetchone()
            rule_snapshot_id = batch_row["rule_snapshot_id"] if batch_row and batch_row["rule_snapshot_id"] else ""
            
            # 生成统一的 snapshot_batch_id（方案要求：确保购物车快照和反写匹配使用相同的 snapshot_batch_id）
            snapshot_batch_id = f"snapshot_{batch_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}"
            
            cart_items = self._run_cart_snapshot(batch_id, trace_path, progress_callback, snapshot_batch_id)
            if not cart_items:
                for item in items:
                    self._save_cart_writeback_state(
                        item["item_id"],
                        "NOT_FOUND",
                        "未读取到购物车商品",
                    )
                return summary, ["未读取到购物车商品，请确认药师帮购物车页面已登录并打开"], ""
            self._append_purchase_trace(trace_path, f"CART_BACKFILL_ONLY_SNAPSHOT count={len(cart_items)} snapshot_batch_id={snapshot_batch_id}", progress_callback)

            matched_cart_keys = set()
            for item in items:
                adapter_item = self._build_cart_adapter_item(item, {"max_allowed_price": ""}, [])
                # 保留之前的采购原因，避免清除未成功下单的原因
                previous_reason = item.get("purchase_reason", "")
                result = {
                    "purchase_status": "success",
                    "purchase_supplier": "",
                    "purchase_product": "",
                    "purchase_spec": "",
                    "purchase_maker": "",
                    "purchase_valid_date": "",
                    "purchase_quantity_result": "",
                    "purchase_price": "",
                    "max_allowed_price": "",
                    "purchase_reason": previous_reason,
                    "actual_ysb_code": item.get("ysb_code") or "",
                    "candidate_count": 1,
                    "executed_at": datetime.now().isoformat(),
                    "cart_write_status": "NOT_STARTED",
                    "cart_write_time": "",
                    "cart_write_message": "",
                    "cart_item_id": "",
                    "cart_write_attempts": 0,
                }
                adapter_result = {
                    "candidateSupplier": item.get("smart_supplier") or item.get("smart_supplier_full") or "",
                }
                backfilled = self._apply_cart_backfill_to_result(
                    result,
                    adapter_item,
                    cart_items,
                    adapter_result,
                    excluded_cart_keys=matched_cart_keys,
                    batch_id=batch_id,
                    purchase_detail_id=item.get("item_id") or "",
                    snapshot_batch_id=snapshot_batch_id  # 使用统一的 snapshot_batch_id
                )
                if backfilled.get("matched"):
                    cart_key = backfilled.get("cartKey", "")
                    if cart_key and cart_key in matched_cart_keys:
                        summary["unmatched"] += 1
                        self._save_cart_writeback_state(
                            item["item_id"],
                            "SKIPPED",
                            "购物车商品已匹配其他采购明细，未重复反写",
                            backfilled.get("wholesaleId", ""),
                        )
                        logs.append(f"第{item.get('row_number')}行 {item.get('source_name')}: 购物车商品已匹配其他行，未重复反写")
                        self._append_purchase_trace(
                            trace_path,
                            f"CART_BACKFILL_DUPLICATE row={item.get('row_number')} cart_key={cart_key} "
                            f"wholesale_id={backfilled.get('wholesaleId', '')}",
                            progress_callback
                        )
                        continue
                    self._save_purchase_result(
                        item["item_id"],
                        result,
                        batch_id=batch_id,
                        rule_snapshot_id=rule_snapshot_id,
                        rule_set_code=resolved_rule_set_code,
                        preserve_purchase_reason=True,
                    )
                    if cart_key:
                        matched_cart_keys.add(cart_key)
                    summary["updated"] += 1
                    logs.append(
                        f"第{item.get('row_number')}行 {item.get('source_name')}: 已反写 - "
                        f"{result.get('purchase_supplier', '')} {result.get('purchase_price', '')}"
                    )
                else:
                    summary["unmatched"] += 1
                    self._save_cart_writeback_state(
                        item["item_id"],
                        result.get("cart_write_status", "NOT_FOUND"),
                        result.get("cart_write_message", "购物车未匹配到"),
                        result.get("cart_item_id", ""),
                    )
                    logs.append(f"第{item.get('row_number')}行 {item.get('source_name')}: 购物车未匹配到")
                self._append_purchase_trace(
                    trace_path,
                    f"CART_BACKFILL_ONLY row={item.get('row_number')} matched={backfilled.get('matched', False)} "
                    f"score={backfilled.get('score', '')} supplier={result.get('purchase_supplier', '')} "
                    f"price={result.get('purchase_price', '')} reason={backfilled.get('reason', '')}",
                    progress_callback
                )

            for cart_item in cart_items:
                cart_key = self._cart_item_key(cart_item)
                if cart_key in matched_cart_keys:
                    continue
                row_number, created = self._upsert_cart_extra_item(batch_id, cart_item)
                summary["extra"] += 1
                action = "新增" if created else "更新"
                logs.append(
                    f"第{row_number}行 {cart_item.get('name')}: 购物车额外商品已{action}登记 - "
                    f"{cart_item.get('supplier', '')} {cart_item.get('price', '')}"
                )
                self._append_purchase_trace(
                    trace_path,
                    f"CART_BACKFILL_EXTRA row={row_number} created={created} wholesale_id={cart_item.get('wholesaleId', '')} "
                    f"name={cart_item.get('name', '')} supplier={cart_item.get('supplier', '')} price={cart_item.get('price', '')}",
                    progress_callback
                )
            if summary["extra"]:
                self._refresh_batch_counts(batch_id)

            self._append_purchase_trace(trace_path, f"CART_BACKFILL_ONLY_FINISH summary={json.dumps(summary, ensure_ascii=False)}", progress_callback)
            return summary, logs, ""
        except Exception as e:
            logging.error(f"购物车反写失败: {e}")
            return {}, [], str(e)

    def _build_cart_adapter_item(self, item: Dict, base_result: Dict, supplier_scope: List[str] = None,
                                     rule_config: Dict = None, rule_snapshot_id: str = "") -> Dict:
        quantity = item.get("actual_purchase_quantity") or item.get("purchase_quantity") or "0"
        source_name = str(item.get("source_name") or "").strip()
        source_spec = str(item.get("source_spec") or "").strip()
        smart_name = str(item.get("smart_name") or "").strip()
        if smart_name in ("未匹配", "未找到", "无", "None", "null", "-"):
            smart_name = ""
        smart_spec = str(item.get("smart_spec") or "").strip()
        if smart_spec in ("未匹配", "未找到", "无", "None", "null", "-"):
            smart_spec = ""
        return {
            "itemId": item.get("item_id", ""),
            "rowNumber": item.get("row_number", ""),
            "wholesaleId": str(item.get("ysb_code") or "").strip(),
            "amount": str(quantity).strip(),
            "name": source_name or smart_name or "",
            "spec": source_spec or smart_spec or "",
            "smartName": smart_name,
            "smartSpec": smart_spec,
            "manufacturer": item.get("source_maker") or "",
            "approval": item.get("source_approval") or "",
            "supplier": item.get("smart_supplier") or "",
            "supplierFull": item.get("smart_supplier_full") or "",
            "price": item.get("smart_price") or "",
            "minAmount": item.get("min_purchase_quantity") or "",
            "stock": item.get("available_stock") or "",
            "expectedPrice": item.get("expected_price") or "",
            "maxAllowedPrice": base_result.get("max_allowed_price", ""),
            "supplierScope": supplier_scope or [],
            "ruleConfig": rule_config or {},
            "ruleSnapshotId": rule_snapshot_id or "",
        }
    
    def _create_purchase_trace_path(self, batch_id: str) -> str:
        log_dir = Path(__file__).resolve().parents[2] / "logs" / "smart_purchase"
        log_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return str(log_dir / f"{batch_id}_{timestamp}.log")
    
    def _append_purchase_trace(self, trace_path: str, message: str, progress_callback: Callable[[str], None] = None):
        if not trace_path:
            return
        try:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            with open(trace_path, "a", encoding="utf-8") as log_file:
                log_file.write(f"[{timestamp}] {message}\n")
            if progress_callback:
                progress_callback(message)
        except Exception as e:
            logging.warning(f"写入智能采购日志失败: {e}")
    
    def _save_single_candidate_score(
        self,
        cursor,
        batch_id: str,
        item: Dict,
        candidate_data: Dict,
        adapter_item: Dict,
        purchase_status: str,
        candidate_rank: int,
        is_selected: bool,
        rule_snapshot_id: str = ""
    ) -> None:
        """保存单个候选的评分记录"""
        # 获取候选商品信息
        candidate_name = candidate_data.get("name") or ""
        candidate_spec = candidate_data.get("spec") or ""
        candidate_maker = candidate_data.get("manufacturer") or ""
        candidate_supplier = candidate_data.get("supplier") or ""
        candidate_supplier_full = candidate_data.get("supplierFull") or ""
        candidate_min_amount = candidate_data.get("minAmount") or ""
        candidate_stock = candidate_data.get("stock") or ""
        candidate_price = self._to_decimal(candidate_data.get("price") or 0)
        
        # 获取评分详情
        match_detail = candidate_data.get("detail") or {}
        match_score = candidate_data.get("score") or 0
        
        # 计算折算比较价
        compare_price = candidate_price * Decimal("0.97") if candidate_price else None
        
        # 获取最高允许价
        max_allowed_price = self._to_decimal(adapter_item.get("max_allowed_price") or 0)
        
        # 获取评分详情中的各项分数
        name_score = match_detail.get("nameScore") or 0
        spec_score = match_detail.get("specScore") or 0
        maker_score = match_detail.get("makerScore") or 0
        total_score = match_score
        
        # 获取各项是否通过的标志（优先读取完整评估字段，字段缺失时保持为空字符串）
        identity_pass = 1 if match_detail.get("identityOk") is True else (0 if match_detail.get("identityOk") is False else "")
        spec_conflict = 1 if match_detail.get("specConflict") is True else (0 if match_detail.get("specConflict") is False else "")
        spec_pass = 1 if candidate_data.get("specOk") is True else (0 if candidate_data.get("specOk") is False else "")
        maker_pass = 1 if candidate_data.get("manufacturerOk") is True else (0 if candidate_data.get("manufacturerOk") is False else "")
        supplier_pass = 1 if candidate_data.get("supplierOk") is True else (0 if candidate_data.get("supplierOk") is False else "")
        price_pass = 1 if candidate_data.get("priceOk") is True else (0 if candidate_data.get("priceOk") is False else "")
        qty_pass = 1 if candidate_data.get("qtyOk") is True else (0 if candidate_data.get("qtyOk") is False else "")
        stock_pass = 1 if candidate_data.get("stockOk") is True else (0 if candidate_data.get("stockOk") is False else "")
        
        # 判断最终是否合格（基于purchase_status和是否选中）
        final_pass = 1 if (purchase_status == "success" and is_selected) else 0
        
        # 判断是否最终选中
        selected = 1 if is_selected else 0
        
        # 获取搜索关键词
        search_keyword = adapter_item.get("name") or ""
        
        # 获取候选起购数量和库存
        min_purchase_quantity = str(candidate_min_amount) if candidate_min_amount else ""
        candidate_stock_str = str(candidate_stock) if candidate_stock else ""
        
        # 获取候选供应商全称
        candidate_supplier_full_str = str(candidate_supplier_full) if candidate_supplier_full else ""
        
        # 获取规则集代码（三期整改：从批次获取实际规则集）
        rule_set_code = self._rule_selection_service.get_batch_rule_set_code(batch_id) or "default_v1"
        
        # 获取拒绝原因
        reject_reason = candidate_data.get("reason") or ""
        
        # 获取候选原始数据
        raw_data = json.dumps(candidate_data, ensure_ascii=False, default=str)
        
        # 插入purchase_candidate_scores表
        cursor.execute('''
            INSERT INTO purchase_candidate_scores (
                purchase_batch_id,
                purchase_detail_id,
                purchase_status,
                rule_set_code,
                search_keyword,
                candidate_rank,
                candidate_name,
                candidate_spec,
                candidate_maker,
                candidate_supplier,
                candidate_supplier_full,
                candidate_price,
                compare_price,
                max_allowed_price,
                min_purchase_quantity,
                candidate_stock,
                name_score,
                spec_score,
                maker_score,
                total_score,
                identity_pass,
                spec_conflict,
                spec_pass,
                maker_pass,
                supplier_pass,
                price_pass,
                qty_pass,
                stock_pass,
                final_pass,
                selected,
                reject_reason,
                raw_data,
                rule_snapshot_id,
                created_at,
                updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            batch_id,
            item.get("itemId") or "",
            purchase_status,
            rule_set_code,
            search_keyword,
            candidate_rank,
            candidate_name,
            candidate_spec,
            candidate_maker,
            candidate_supplier,
            candidate_supplier_full_str,
            str(candidate_price) if candidate_price else "",
            str(compare_price) if compare_price else "",
            str(max_allowed_price) if max_allowed_price else "",
            min_purchase_quantity,
            candidate_stock_str,
            name_score,
            spec_score,
            maker_score,
            total_score,
            identity_pass,
            spec_conflict,
            spec_pass,
            maker_pass,
            supplier_pass,
            price_pass,
            qty_pass,
            stock_pass,
            final_pass,
            selected,
            reject_reason,
            raw_data,
            rule_snapshot_id,
            datetime.now().isoformat(),
            datetime.now().isoformat()
        ))
        
        # 同时插入smart_purchase_candidates表（方案要求）
        cursor.execute('''
            INSERT INTO smart_purchase_candidates (
                purchase_batch_id,
                purchase_detail_id,
                rule_set_code,
                search_keyword,
                candidate_rank,
                candidate_name,
                candidate_spec,
                candidate_maker,
                candidate_supplier,
                candidate_supplier_full,
                candidate_price,
                compare_price,
                max_allowed_price,
                min_purchase_quantity,
                candidate_stock,
                name_score,
                spec_score,
                maker_score,
                total_score,
                identity_pass,
                spec_conflict,
                spec_pass,
                maker_pass,
                supplier_pass,
                price_pass,
                qty_pass,
                stock_pass,
                final_pass,
                selected,
                reject_reason,
                raw_data,
                rule_snapshot_id,
                created_at,
                updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            batch_id,
            item.get("itemId") or "",
            rule_set_code,
            search_keyword,
            candidate_rank,
            candidate_name,
            candidate_spec,
            candidate_maker,
            candidate_supplier,
            candidate_supplier_full_str,
            str(candidate_price) if candidate_price else "",
            str(compare_price) if compare_price else "",
            str(max_allowed_price) if max_allowed_price else "",
            min_purchase_quantity,
            candidate_stock_str,
            name_score,
            spec_score,
            maker_score,
            total_score,
            identity_pass,
            spec_conflict,
            spec_pass,
            maker_pass,
            supplier_pass,
            price_pass,
            qty_pass,
            stock_pass,
            final_pass,
            selected,
            reject_reason,
            raw_data,
            rule_snapshot_id,
            datetime.now().isoformat(),
            datetime.now().isoformat()
        ))
    
    def _save_candidate_scores(
        self,
        batch_id: str,
        item: Dict,
        result: Dict,
        adapter_item: Dict,
        rule_snapshot_id: str = ""
    ) -> None:
        """保存候选评分数据到数据库（一期落库）"""
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            # 获取候选列表（前10个候选）
            candidates = result.get("candidates") or []
            
            # 获取采购状态
            purchase_status = result.get("status") or ""
            
            # 如果candidates字段存在，则批量保存前10个候选的评分记录
            if candidates and len(candidates) > 0:
                for i, candidate in enumerate(candidates):
                    # 根据Node返回的isSelected标记判断是否选中（不再默认i==0为选中）
                    is_selected = candidate.get("isSelected") is True
                    
                    # 调用_save_single_candidate_score函数保存单个候选的评分记录
                    self._save_single_candidate_score(
                        cursor,
                        batch_id,
                        item,
                        candidate,
                        adapter_item,
                        purchase_status,
                        i + 1,  # candidate_rank从1开始
                        is_selected,
                        rule_snapshot_id=rule_snapshot_id
                    )
                
                conn.commit()
                logging.info(f"✓ 已保存前{len(candidates)}个候选评分数据: batch_id={batch_id}, item_id={item.get('itemId')}")
            else:
                # 如果没有candidates字段，则保存单个候选的评分记录（兼容旧逻辑）
                # 获取评分详情
                match_detail = result.get("matchDetail") or {}
                match_score = result.get("matchScore") or 0
                
                # 计算折算比较价
                candidate_price = self._to_decimal(result.get("matchedPrice") or 0)
                compare_price = candidate_price * Decimal("0.97") if candidate_price else None
                
                # 获取最高允许价
                max_allowed_price = self._to_decimal(adapter_item.get("max_allowed_price") or 0)
                
                # 获取候选商品信息
                candidate_name = result.get("matchedName") or ""
                candidate_spec = result.get("matchedSpec") or ""
                candidate_maker = result.get("matchedManufacturer") or ""
                candidate_supplier = result.get("matchedSupplier") or ""
                candidate_supplier_full = result.get("matchedSupplierFull") or ""
                candidate_min_amount = result.get("matchedMinAmount") or ""
                candidate_stock = result.get("matchedStock") or ""
                candidate_rank = result.get("candidateRank") or 0
                
                # 获取评分详情中的各项分数
                name_score = match_detail.get("nameScore") or 0
                spec_score = match_detail.get("specScore") or 0
                maker_score = match_detail.get("makerScore") or 0
                total_score = match_score
                
                # 获取各项是否通过的标志（优先读取完整评估字段，字段缺失时保持为空字符串）
                # 如果match_detail中包含字段，则使用该字段的值；如果不包含，则保持为空字符串
                identity_pass = 1 if match_detail.get("identityOk") is True else (0 if match_detail.get("identityOk") is False else "")
                spec_conflict = 1 if match_detail.get("specConflict") is True else (0 if match_detail.get("specConflict") is False else "")
                spec_pass = 1 if match_detail.get("specOk") is True else (0 if match_detail.get("specOk") is False else "")
                maker_pass = 1 if match_detail.get("manufacturerOk") is True else (0 if match_detail.get("manufacturerOk") is False else "")
                supplier_pass = 1 if match_detail.get("supplierOk") is True else (0 if match_detail.get("supplierOk") is False else "")
                price_pass = 1 if match_detail.get("priceOk") is True else (0 if match_detail.get("priceOk") is False else "")
                qty_pass = 1 if match_detail.get("qtyOk") is True else (0 if match_detail.get("qtyOk") is False else "")
                stock_pass = 1 if match_detail.get("stockOk") is True else (0 if match_detail.get("stockOk") is False else "")
                
                # 判断最终是否合格（基于purchase_status）
                # 如果purchase_status=success，则final_pass=1；否则final_pass=0
                final_pass = 1 if purchase_status == "success" else 0
                
                # 判断是否最终选中（基于final_pass）
                # 如果final_pass=1，则selected=1；如果final_pass=0，则selected=0
                selected = final_pass
                
                # 获取搜索关键词
                search_keyword = adapter_item.get("name") or ""
                
                # 获取候选起购数量和库存
                min_purchase_quantity = str(candidate_min_amount) if candidate_min_amount else ""
                candidate_stock_str = str(candidate_stock) if candidate_stock else ""
                
                # 获取候选供应商全称
                candidate_supplier_full_str = str(candidate_supplier_full) if candidate_supplier_full else ""
                
                # 获取候选原始顺序
                candidate_rank_int = int(candidate_rank) if candidate_rank else 0
                
                # 获取规则集代码（三期整改：从批次获取实际规则集）
                rule_set_code = self._rule_selection_service.get_batch_rule_set_code(batch_id) or "default_v1"
                
                # 获取拒绝原因
                reject_reason = match_detail.get("rejectReason") or ""
                
                # 获取候选原始数据
                raw_data = json.dumps(result, ensure_ascii=False)
                
                # 插入数据库
                cursor.execute('''
                    INSERT INTO purchase_candidate_scores (
                        purchase_batch_id,
                        purchase_detail_id,
                        purchase_status,
                        rule_set_code,
                        search_keyword,
                        candidate_rank,
                        candidate_name,
                        candidate_spec,
                        candidate_maker,
                        candidate_supplier,
                        candidate_supplier_full,
                        candidate_price,
                        compare_price,
                        max_allowed_price,
                        min_purchase_quantity,
                        candidate_stock,
                        name_score,
                        spec_score,
                        maker_score,
                        total_score,
                        identity_pass,
                        spec_conflict,
                        spec_pass,
                        maker_pass,
                        supplier_pass,
                        price_pass,
                        qty_pass,
                        stock_pass,
                        final_pass,
                        selected,
                        reject_reason,
                        raw_data,
                        rule_snapshot_id,
                        created_at,
                        updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    batch_id,
                    item.get("itemId") or "",
                    purchase_status,
                    rule_set_code,
                    search_keyword,
                    candidate_rank_int,
                    candidate_name,
                    candidate_spec,
                    candidate_maker,
                    candidate_supplier,
                    candidate_supplier_full,
                    str(candidate_price) if candidate_price else "",
                    str(compare_price) if compare_price else "",
                    str(max_allowed_price) if max_allowed_price else "",
                    min_purchase_quantity,
                    candidate_stock_str,
                    name_score,
                    spec_score,
                    maker_score,
                    total_score,
                    identity_pass,
                    spec_conflict,
                    spec_pass,
                    maker_pass,
                    supplier_pass,
                    price_pass,
                    qty_pass,
                    stock_pass,
                    final_pass,
                    selected,
                    reject_reason,
                    raw_data,
                    rule_snapshot_id,
                    datetime.now().isoformat(),
                    datetime.now().isoformat()
                ))
                
                # P2-1整改：同步写入smart_purchase_candidates表，保持双表一致
                # 双表写入放在同一事务中，任一表失败时整体回滚
                cursor.execute('''
                    INSERT INTO smart_purchase_candidates (
                        purchase_batch_id,
                        purchase_detail_id,
                        rule_set_code,
                        search_keyword,
                        candidate_rank,
                        candidate_name,
                        candidate_spec,
                        candidate_maker,
                        candidate_supplier,
                        candidate_supplier_full,
                        candidate_price,
                        compare_price,
                        max_allowed_price,
                        min_purchase_quantity,
                        candidate_stock,
                        name_score,
                        spec_score,
                        maker_score,
                        total_score,
                        identity_pass,
                        spec_conflict,
                        spec_pass,
                        maker_pass,
                        supplier_pass,
                        price_pass,
                        qty_pass,
                        stock_pass,
                        final_pass,
                        selected,
                        reject_reason,
                        raw_data,
                        rule_snapshot_id,
                        created_at,
                        updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    batch_id,
                    item.get("itemId") or "",
                    rule_set_code,
                    search_keyword,
                    candidate_rank_int,
                    candidate_name,
                    candidate_spec,
                    candidate_maker,
                    candidate_supplier,
                    candidate_supplier_full_str,
                    str(candidate_price) if candidate_price else "",
                    str(compare_price) if compare_price else "",
                    str(max_allowed_price) if max_allowed_price else "",
                    min_purchase_quantity,
                    candidate_stock_str,
                    name_score,
                    spec_score,
                    maker_score,
                    total_score,
                    identity_pass,
                    spec_conflict,
                    spec_pass,
                    maker_pass,
                    supplier_pass,
                    price_pass,
                    qty_pass,
                    stock_pass,
                    final_pass,
                    selected,
                    reject_reason,
                    raw_data,
                    rule_snapshot_id,
                    datetime.now().isoformat(),
                    datetime.now().isoformat()
                ))
                
                # 双表统一提交，确保事务一致性
                conn.commit()
                logging.info(f"✓ 已保存候选评分数据（双表一致）: batch_id={batch_id}, item_id={item.get('itemId')}, score={total_score}")
            
        except Exception as e:
            logging.warning(f"保存候选评分数据失败: {e}")
            try:
                conn.rollback()
            except:
                pass
    
    def _check_cart_adapter_syntax(self) -> str:
        """Return a concise error when the Node adapter cannot be parsed."""
        node_bin = os.environ.get("NODE_EXE") or shutil.which("node")
        if not node_bin:
            return "未找到 Node.js，无法启动购物车执行器"
        script_path = Path(__file__).resolve().parents[1] / "automation" / "ysbang_cart_add_onebyone.mjs"
        if not script_path.exists():
            return f"购物车执行脚本不存在: {script_path}"
        try:
            completed = subprocess.run(
                [node_bin, "--check", str(script_path)],
                cwd=str(Path(__file__).resolve().parents[2]),
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                encoding="utf-8",
                errors="replace",
                timeout=30,
            )
        except Exception as error:
            return f"购物车执行脚本语法检查异常: {error}"
        if completed.returncode == 0:
            return ""
        detail = (completed.stderr or completed.stdout or "Node syntax check failed").strip()
        return detail[:2000]

    def _run_cart_adapter_batch(
        self,
        batch_id: str,
        adapter_items: List[Dict],
        trace_path: str = "",
        progress_callback: Callable[[str], None] = None,
        stop_callback: Callable[[], bool] = None
    ) -> List[Dict]:
        node_bin = os.environ.get("NODE_EXE") or shutil.which("node")
        if not node_bin:
            self._append_purchase_trace(trace_path, "CART_ADAPTER_ERROR Node.js not found", progress_callback)
            return self._adapter_failure_results(adapter_items, "未找到 Node.js，无法连接药师帮购物车执行器")
        
        script_path = Path(__file__).resolve().parents[1] / "automation" / "ysbang_cart_add_onebyone.mjs"
        if not script_path.exists():
            self._append_purchase_trace(trace_path, f"CART_ADAPTER_ERROR script not found: {script_path}", progress_callback)
            return self._adapter_failure_results(adapter_items, f"购物车执行脚本不存在: {script_path}")
        
        temp_dir = Path(tempfile.gettempdir()) / "rpa_smart_purchase"
        temp_dir.mkdir(parents=True, exist_ok=True)
        input_path = temp_dir / f"{batch_id}_cart_input.json"
        output_path = temp_dir / f"{batch_id}_cart_result.json"
        input_path.write_text(
            json.dumps({
                "items": adapter_items,
                "logPath": trace_path,
                "ruleConfig": adapter_items[0].get("ruleConfig", {}) if adapter_items else {},
                "ruleSnapshotId": adapter_items[0].get("ruleSnapshotId", "") if adapter_items else "",
            }, ensure_ascii=False, indent=2),
            encoding="utf-8"
        )
        if output_path.exists():
            output_path.unlink()
        self._append_purchase_trace(
            trace_path,
            f"CART_ADAPTER_CALL node={node_bin} script={script_path} input={input_path} output={output_path}",
            progress_callback
        )
        
        completed_stdout = ""
        completed_stderr = ""
        try:
            trace_position = Path(trace_path).stat().st_size if trace_path and Path(trace_path).exists() else 0
            process = subprocess.Popen(
                [node_bin, str(script_path), str(input_path), str(output_path)],
                cwd=str(Path(__file__).resolve().parents[2]),
                env={**os.environ, "YSBANG_CDP_URL": os.environ.get("YSBANG_CDP_URL", "http://127.0.0.1:9222")},
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                encoding="utf-8",
                errors="replace"
            )
            deadline = time.monotonic() + max(180, len(adapter_items) * 75)
            while process.poll() is None:
                trace_position = self._emit_new_trace_lines(trace_path, trace_position, progress_callback)
                if stop_callback and stop_callback():
                    process.kill()
                    completed_stdout, completed_stderr = process.communicate(timeout=5)
                    trace_position = self._emit_new_trace_lines(trace_path, trace_position, progress_callback)
                    self._append_purchase_trace(
                        trace_path,
                        f"CART_ADAPTER_STOPPED stdout={completed_stdout.strip()} stderr={completed_stderr.strip()}",
                        progress_callback
                    )
                    for path in (input_path, output_path):
                        try:
                            if path.exists():
                                path.unlink()
                        except Exception as cleanup_error:
                            self._append_purchase_trace(
                                trace_path,
                                f"CART_ADAPTER_STOP_CLEANUP_ERROR path={path} error={cleanup_error}",
                                progress_callback
                            )
                    return self._adapter_failure_results(adapter_items, "用户请求停止，已结束当前 Node 加购进程")
                if time.monotonic() > deadline:
                    process.kill()
                    completed_stdout, completed_stderr = process.communicate(timeout=5)
                    self._append_purchase_trace(trace_path, "CART_ADAPTER_TIMEOUT", progress_callback)
                    return self._adapter_failure_results(adapter_items, "药师帮购物车执行超时，已停止本批次")
                time.sleep(0.5)
            completed_stdout, completed_stderr = process.communicate()
            trace_position = self._emit_new_trace_lines(trace_path, trace_position, progress_callback)
            self._append_purchase_trace(
                trace_path,
                f"CART_ADAPTER_EXIT code={process.returncode} stdout={completed_stdout.strip()} stderr={completed_stderr.strip()}",
                progress_callback
            )
        except Exception as e:
            self._append_purchase_trace(trace_path, f"CART_ADAPTER_EXCEPTION {e}", progress_callback)
            return self._adapter_failure_results(adapter_items, f"调用药师帮购物车执行器失败: {e}")
        
        if output_path.exists():
            try:
                payload = json.loads(output_path.read_text(encoding="utf-8"))
                results = payload.get("results", [])
                if results:
                    self._append_purchase_trace(trace_path, f"CART_ADAPTER_RESULTS count={len(results)}", progress_callback)
                    
                    # 保存候选评分数据到数据库（一期落库）
                    for result in results:
                        # 找到对应的adapter_item
                        adapter_item = None
                        for item in adapter_items:
                            if item.get("itemId") == result.get("itemId"):
                                adapter_item = item
                                break
                        
                        if adapter_item:
                            # 构造item字典（用于保存）
                            item_dict = {
                                "itemId": result.get("itemId"),
                                "rowNumber": result.get("rowNumber"),
                                "name": adapter_item.get("name"),
                                "spec": adapter_item.get("spec"),
                                "manufacturer": adapter_item.get("manufacturer"),
                                "amount": adapter_item.get("amount"),
                            }
                            
                            # 保存候选评分数据
                            # 从adapter_item获取ruleSnapshotId
                            rs_id = adapter_item.get("ruleSnapshotId", "")
                            self._save_candidate_scores(batch_id, item_dict, result, adapter_item, rule_snapshot_id=rs_id)
                    
                    return self._complete_missing_adapter_results(
                        adapter_items,
                        results,
                        "购物车执行器提前结束，本行未执行；请检查网页状态后重试失败项",
                    )
            except Exception as e:
                self._append_purchase_trace(trace_path, f"CART_ADAPTER_READ_RESULT_ERROR {e}", progress_callback)
                return self._adapter_failure_results(adapter_items, f"读取购物车执行结果失败: {e}")
        
        message = completed_stderr.strip() or completed_stdout.strip() or "药师帮购物车执行器未返回结果"
        return self._adapter_failure_results(adapter_items, message)

    def _run_cart_snapshot(
        self,
        batch_id: str,
        trace_path: str = "",
        progress_callback: Callable[[str], None] = None,
        snapshot_batch_id: str = ""
    ) -> List[Dict]:
        node_bin = os.environ.get("NODE_EXE") or shutil.which("node")
        if not node_bin:
            self._append_purchase_trace(trace_path, "CART_SNAPSHOT_ERROR Node.js not found", progress_callback)
            return []
        script_path = Path(__file__).resolve().parents[1] / "automation" / "ysbang_cart_snapshot.mjs"
        if not script_path.exists():
            self._append_purchase_trace(trace_path, f"CART_SNAPSHOT_ERROR script not found: {script_path}", progress_callback)
            return []
        temp_dir = Path(tempfile.gettempdir()) / "rpa_smart_purchase"
        temp_dir.mkdir(parents=True, exist_ok=True)
        output_path = temp_dir / f"{batch_id}_cart_snapshot.json"
        if output_path.exists():
            output_path.unlink()
        self._append_purchase_trace(
            trace_path,
            f"CART_SNAPSHOT_CALL node={node_bin} script={script_path} output={output_path}",
            progress_callback
        )
        try:
            self._append_purchase_trace(trace_path, f"CART_SNAPSHOT_START script={script_path}", progress_callback)
            completed = subprocess.run(
                [node_bin, str(script_path), str(output_path)],
                cwd=str(Path(__file__).resolve().parents[2]),
                env={**os.environ, "YSBANG_CDP_URL": os.environ.get("YSBANG_CDP_URL", "http://127.0.0.1:9222")},
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                encoding="utf-8",
                errors="replace",
                timeout=120
            )
            self._append_purchase_trace(
                trace_path,
                f"CART_SNAPSHOT_EXIT code={completed.returncode} stdout={completed.stdout.strip()} stderr={completed.stderr.strip()}",
                progress_callback
            )
            if completed.returncode != 0 or not output_path.exists():
                return []
            payload = json.loads(output_path.read_text(encoding="utf-8"))
            items = payload.get("items", [])
            cart_items = items if isinstance(items, list) else []
            
            # 保存购物车快照到数据库（方案要求）
            # 使用传入的 snapshot_batch_id，确保与反写匹配使用相同的 snapshot_batch_id
            if snapshot_batch_id:
                self._save_cart_snapshot(batch_id, cart_items, snapshot_batch_id)
            else:
                # 如果没有传入 snapshot_batch_id，则生成一个新的
                self._save_cart_snapshot(batch_id, cart_items)
            
            return cart_items
        except Exception as e:
            self._append_purchase_trace(trace_path, f"CART_SNAPSHOT_EXCEPTION {e}", progress_callback)
            return []
    
    def _save_cart_snapshot(self, batch_id: str, cart_items: List[Dict], snapshot_batch_id: str = "") -> None:
        """保存购物车快照到数据库（方案要求）"""
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            # 使用传入的 snapshot_batch_id，或生成一个新的
            if not snapshot_batch_id:
                snapshot_batch_id = f"snapshot_{batch_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}"
            
            # 确定快照状态
            snapshot_status = "empty" if len(cart_items) == 0 else "completed"
            
            # 插入购物车快照批次记录
            cursor.execute('''
                INSERT INTO smart_cart_snapshots (
                    snapshot_batch_id,
                    snapshot_type,
                    total_items,
                    matched_items,
                    unmatched_items,
                    snapshot_status,
                    snapshot_time,
                    created_by,
                    created_at,
                    updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                snapshot_batch_id,
                "cart_backfill",  # 快照类型：购物车反写
                len(cart_items),  # 总商品数
                0,  # 匹配商品数（初始为0，后续更新）
                len(cart_items),  # 未匹配商品数（初始为总商品数）
                snapshot_status,  # 快照状态
                datetime.now().isoformat(),  # 快照时间
                "system",  # 创建人
                datetime.now().isoformat(),  # 创建时间
                datetime.now().isoformat()  # 更新时间
            ))
            
            # 插入购物车快照明细记录
            for i, cart_item in enumerate(cart_items):
                cursor.execute('''
                    INSERT INTO smart_cart_snapshot_items (
                        snapshot_batch_id,
                        item_index,
                        item_name,
                        item_spec,
                        item_maker,
                        item_supplier,
                        item_price,
                        item_quantity,
                        item_stock,
                        match_status,
                        match_detail,
                        created_at,
                        updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    snapshot_batch_id,
                    i + 1,  # 商品索引从1开始
                    cart_item.get("name") or "",
                    cart_item.get("spec") or "",
                    cart_item.get("manufacturer") or "",
                    cart_item.get("supplier") or "",
                    cart_item.get("price") or "",
                    cart_item.get("quantity") or "",
                    cart_item.get("stock") or "",
                    "pending",  # 匹配状态（初始为pending）
                    json.dumps(cart_item, ensure_ascii=False, default=str),  # 匹配详情（原始购物车商品数据）
                    datetime.now().isoformat(),  # 创建时间
                    datetime.now().isoformat()  # 更新时间
                ))
            
            conn.commit()
            logging.info(f"✓ 已保存购物车快照: snapshot_batch_id={snapshot_batch_id}, total_items={len(cart_items)}, status={snapshot_status}")
            
        except Exception as e:
            logging.warning(f"保存购物车快照失败: {e}")
            try:
                conn.rollback()
            except:
                pass

    def _cart_text(self, value) -> str:
        return "" if value is None else str(value).strip()

    def _cart_norm(self, value) -> str:
        text = self._cart_text(value).lower()
        text = re.sub(r"[()（）【】\[\]\s,，。.;；:_/\\-]", "", text)
        return text.replace("×", "*").replace("x", "*")

    def _cart_code(self, value) -> str:
        text = self._cart_text(value).upper()
        text = re.sub(r"^(YSB|YB|ID)", "", text)
        digits = re.sub(r"\D", "", text)
        return digits or text.strip()

    def _cart_text_score(self, source, target) -> int:
        source_text = self._cart_norm(source)
        target_text = self._cart_norm(target)
        if not source_text or not target_text:
            return 0
        if source_text == target_text:
            return 100
        if source_text in target_text or target_text in source_text:
            return 85
        chars = set(source_text)
        hits = sum(1 for char in target_text if char in chars)
        return round(hits / max(len(source_text), len(target_text)) * 100)

    def _cart_product_core_name(self, value) -> str:
        text = self._cart_text(value).lower()
        text = re.sub(r"[\(\uFF08\[].*?[\)\uFF09\]]", " ", text)
        text = re.sub(
            r"\d+(?:\.\d+)?\s*(?:mg|g|kg|ml|l|ug|\u03bcg|iu|%|\u6beb\u514b|\u514b|\u5343\u514b|\u6beb\u5347|\u5fae\u514b)",
            " ",
            text,
            flags=re.I
        )
        text = re.sub(
            r"\d+(?:\.\d+)?\s*(?:\u4e38|\u7247|\u7c92|\u888b|\u652f|\u74f6|\u76d2|\u677f|\u8d34|\u5305|\u7ba1|\u63d0|\u677f\u88c5|\u7247\u88c5|\u7c92\u88c5)",
            " ",
            text,
            flags=re.I
        )
        text = re.sub(
            r"(?:rx|otc|\u5305\u90ae|\u8d77\u8d2d|\u9996\u63a8|\u63a8\u8350|\u7279\u4ef7|\u70ed\u5356|\u65e5\u5e38|\u5c0f\u836f\u7cbe\u9009\w*)",
            " ",
            text,
            flags=re.I
        )
        return re.sub(r"[^0-9a-z\u4e00-\u9fff]", "", self._cart_norm(text))

    def _cart_core_name_compatible(self, source, target) -> bool:
        source_core = self._cart_product_core_name(source)
        target_core = self._cart_product_core_name(target)
        if len(source_core) < 3 or len(target_core) < 3:
            return True
        if source_core in target_core or target_core in source_core:
            return True
        return self._cart_text_score(source_core, target_core) >= 70

    def _cart_brand_hint(self, value, core_name: str) -> str:
        text = self._cart_norm(value)
        if core_name:
            text = text.replace(core_name, "")
        text = re.sub(r"\d+(?:\.\d+)?(?:mg|g|kg|ml|l|ug|\u03bcg|iu|%)?", "", text, flags=re.I)
        return re.sub(r"[^0-9a-z\u4e00-\u9fff]", "", text)

    def _cart_brand_compatible(self, source, target) -> bool:
        source_core = self._cart_product_core_name(source)
        target_core = self._cart_product_core_name(target)
        source_brands = {
            self._cart_brand_hint(source, source_core),
            self._cart_brand_hint(source, target_core),
        }
        target_brands = {
            self._cart_brand_hint(target, source_core),
            self._cart_brand_hint(target, target_core),
        }
        for source_brand in source_brands:
            for target_brand in target_brands:
                if len(source_brand) < 2 or len(target_brand) < 2:
                    continue
                if (
                    source_brand in target_brand
                    or target_brand in source_brand
                    or self._cart_text_score(source_brand, target_brand) >= 60
                ):
                    return True
        return False

    def _cart_package_total_conflict(self, source, target) -> bool:
        source_strength, source_total = self._cart_spec_parts(source)
        target_strength, target_total = self._cart_spec_parts(target)
        return bool(
            source_total
            and target_total
            and source_total != target_total
            and (not source_strength or not target_strength or source_strength == target_strength)
        )

    def _cart_spec_numbers(self, value) -> List[str]:
        text = self._cart_text(value).lower().replace("×", "*").replace("x", "*")
        return [match.group(1).rstrip("0").rstrip(".") for match in re.finditer(r"(?<![a-z0-9])(\d+(?:\.\d+)?)(?!\d)", text, re.I)]

    def _cart_spec_unit_value(self, value) -> str:
        text = re.sub(r"\s+", "", self._cart_text(value).lower())
        text = text.replace("ug", "μg").replace("µg", "μg").replace("渭g", "μg")
        match = re.fullmatch(r"(\d+(?:\.\d+)?)(mg|g|ml|μg|iu|%)", text)
        if not match:
            return text
        number = Decimal(match.group(1))
        unit = match.group(2)
        if unit == "mg" and number >= 1000:
            number = number / Decimal("1000")
            unit = "g"
        elif unit == "mg" and number < 1000:
            normalized = number / Decimal("1000")
            if normalized == normalized.quantize(Decimal("0.001")):
                return f"{self._decimal_plain(normalized)}g"
        return f"{self._decimal_plain(number)}{unit}"

    def _decimal_plain(self, value: Decimal) -> str:
        text = format(value.normalize(), "f")
        return text.rstrip("0").rstrip(".") if "." in text else text

    def _cart_spec_unit_set(self, value) -> set:
        text = self._cart_text(value).lower().replace("×", "*").replace("脳", "*").replace("x", "*")
        return {
            self._cart_spec_unit_value(match.group(0))
            for match in re.finditer(r"\d+(?:\.\d+)?\s*(?:mg|g|ml|ug|μg|µg|渭g|iu|%)", text, re.I)
        }

    def _cart_has_count_unit(self, value) -> bool:
        text = self._cart_text(value).lower()
        return bool(re.search(r"\d+(?:\.\d+)?\s*(片|粒|丸|袋|支|瓶|盒|板|贴|包|枚|只|管|条|s)", text, re.I))

    def _cart_spec_parts(self, value) -> Tuple[str, int]:
        text = self._cart_text(value).lower().replace("×", "*").replace("脳", "*").replace("x", "*")
        strength = ""
        strength_match = re.search(r"(\d+(?:\.\d+)?\s*(?:mg|g|ml|ug|μg|µg|渭g|iu|%))", text, re.I)
        if strength_match:
            strength = self._cart_spec_unit_value(strength_match.group(1))
        count_unit_pattern = r"(?:片|粒|丸|袋|支|瓶|盒|板|贴|包|枚|只|管|条|s)"
        numbers = [
            int(float(match.group(1)))
            for match in re.finditer(rf"(?<![a-z0-9])(\d+(?:\.\d+)?)(?!\d)\s*{count_unit_pattern}", text, re.I)
        ]
        while len(numbers) > 1 and len(numbers) % 2 == 0:
            half = len(numbers) // 2
            if numbers[:half] != numbers[half:]:
                break
            numbers = numbers[:half]
        total = 0
        if numbers:
            total = 1
            for number in numbers:
                total *= number
        return strength, total

    def _cart_spec_score(self, source, target) -> int:
        source_strength, source_total = self._cart_spec_parts(source)
        target_strength, target_total = self._cart_spec_parts(target)
        source_units = self._cart_spec_unit_set(source)
        target_units = self._cart_spec_unit_set(target)
        if source_strength and target_strength and source_strength == target_strength:
            if source_total and target_total and source_total == target_total:
                return 100
            if source_total and target_total and source_total != target_total:
                if self._cart_has_count_unit(source) and self._cart_has_count_unit(target):
                    return 45
                shared_units = source_units & target_units
                if any(unit.endswith(("g", "ml")) and not unit.endswith("mg") for unit in shared_units):
                    return 85
                return 45
            if not source_total or not target_total:
                return 85
        source_numbers = self._cart_spec_numbers(source)
        target_numbers = self._cart_spec_numbers(target)
        if source_units and target_units and source_units & target_units:
            shared_units = source_units & target_units
            if any(unit.endswith(("g", "ml")) and not unit.endswith("mg") for unit in shared_units):
                return 85
            if source_units <= target_units or target_units <= source_units:
                return 85
        if source_numbers and target_numbers:
            source_set = set(source_numbers)
            target_set = set(target_numbers)
            shared = source_set & target_set
            if len(source_numbers) >= 2 and len(shared) >= min(2, len(source_set)):
                return 100
            if len(source_numbers) == 1 and source_numbers[0] in target_set:
                return 85
        return self._cart_text_score(source, target)

    def _cart_match_score(self, item: Dict, cart_item: Dict) -> int:
        target_name = item.get("name") or item.get("matchedName") or ""
        target_spec = item.get("spec") or item.get("matchedSpec") or ""
        target_maker = item.get("manufacturer") or item.get("matchedManufacturer") or ""
        cart_name = cart_item.get("name") or cart_item.get("drugName") or ""
        cart_spec = cart_item.get("spec") or ""
        cart_maker = cart_item.get("manufacturer") or cart_item.get("maker") or ""
        name_score = self._cart_text_score(target_name, cart_name)
        spec_score = self._cart_spec_score(target_spec, f"{cart_spec}{cart_name}") if target_spec else 60
        maker_score = self._cart_text_score(target_maker, cart_maker) if target_maker else 60
        core_ok = self._cart_core_name_compatible(target_name, cart_name)
        brand_ok = self._cart_brand_compatible(target_name, cart_name)
        core_score = self._cart_text_score(
            self._cart_product_core_name(target_name),
            self._cart_product_core_name(cart_name)
        )
        identity_ok = core_ok or (brand_ok and core_score >= 60 and (spec_score >= 80 or maker_score >= 80))
        if spec_score >= 95 and name_score >= 50 and maker_score == 0:
            maker_score = 60
        score = round(name_score * 0.62 + spec_score * 0.20 + maker_score * 0.18)
        if target_spec and self._cart_package_total_conflict(target_spec, f"{cart_spec}{cart_name}"):
            return min(score, 61)
        if target_spec and spec_score < 60 and not (core_ok and brand_ok and spec_score >= 45):
            return min(score, 61)
        if not identity_ok:
            return min(score, 61)
        if target_maker and (not cart_maker or maker_score < 70) and not (identity_ok and brand_ok and spec_score >= 45):
            return min(score, 61)
        if identity_ok and brand_ok and spec_score >= 45:
            score = max(score, 70)
        if spec_score >= 95 and maker_score >= 80 and name_score >= 30:
            return max(score, 82)
        if spec_score >= 80 and maker_score >= 55 and name_score >= 55:
            return max(score, 70)
        if spec_score >= 80 and maker_score >= 90 and name_score >= 40:
            return max(score, 70)
        if spec_score >= 95 and maker_score >= 90 and name_score >= 30:
            return max(score, 82)
        return score

    def _find_cart_backfill_item(
        self,
        adapter_item: Dict,
        result: Dict,
        cart_items: List[Dict],
        excluded_cart_keys: set = None
    ) -> Tuple[Dict, int, str]:
        excluded_cart_keys = excluded_cart_keys or set()
        wanted_id = self._cart_code(result.get("actual_ysb_code") or adapter_item.get("wholesaleId"))
        if wanted_id:
            target_spec = adapter_item.get("spec") or adapter_item.get("matchedSpec") or ""
            same_id = [
                item for item in cart_items
                if self._cart_code(item.get("wholesaleId")) == wanted_id
                and self._cart_item_key(item) not in excluded_cart_keys
                and not self._cart_package_total_conflict(
                    target_spec,
                    f"{item.get('spec') or ''}{item.get('name') or item.get('drugName') or ''}"
                )
            ]
            if same_id:
                chosen = sorted(same_id, key=lambda item: self._to_decimal(item.get("price")) or Decimal("999999"))[0]
                return chosen, 100, "wholesaleId"
        scored = []
        for cart_item in cart_items:
            if self._cart_item_key(cart_item) in excluded_cart_keys:
                continue
            score = self._cart_match_score(adapter_item, cart_item)
            if score >= 62:
                scored.append((score, cart_item))
        if not scored:
            return {}, 0, "not_found"
        scored.sort(key=lambda pair: (-pair[0], self._to_decimal(pair[1].get("price")) or Decimal("999999")))
        return scored[0][1], scored[0][0], "name_spec_manufacturer"

    def _is_internal_cart_supplier(self, supplier: str) -> bool:
        return bool(re.fullmatch(r"摇钱树\d+", self._cart_text(supplier)))

    def _cart_item_key(self, cart_item: Dict) -> str:
        return "|".join([
            self._cart_text(cart_item.get("wholesaleId")),
            self._cart_norm(cart_item.get("name")),
            self._cart_norm(cart_item.get("spec")),
            self._cart_norm(cart_item.get("manufacturer") or cart_item.get("maker")),
            self._cart_norm(cart_item.get("supplier")),
            self._cart_text(cart_item.get("price")),
        ])

    def _apply_cart_backfill_to_result(
        self,
        result: Dict,
        adapter_item: Dict,
        cart_items: List[Dict],
        adapter_result: Dict = None,
        excluded_cart_keys: set = None,
        batch_id: str = "",
        purchase_detail_id: str = "",
        snapshot_batch_id: str = ""
    ) -> Dict:
        adapter_result = adapter_result or {}
        cart_item, score, reason = self._find_cart_backfill_item(adapter_item, result, cart_items, excluded_cart_keys)
        
        # 保存购物车反写匹配过程到数据库（方案要求）
        match_status = "matched" if cart_item else "unmatched"
        match_detail = {
            "score": score,
            "reason": reason,
            "adapter_item": adapter_item,
            "cart_item": cart_item if cart_item else {},
            "excluded_cart_keys": list(excluded_cart_keys) if excluded_cart_keys else [],
        }
        self._save_cart_backfill_match(
            batch_id,
            purchase_detail_id,
            match_status,
            reason,  # match_type: wholesaleId或name_spec_manufacturer
            score,
            match_detail,
            snapshot_batch_id,  # 使用传入的 snapshot_batch_id，确保与购物车快照使用相同的 snapshot_batch_id
            rule_snapshot_id=adapter_item.get("ruleSnapshotId", "")  # 二期整改：关联规则快照
        )
        
        if not cart_item:
            result["cart_write_status"] = "NOT_FOUND"
            result["cart_write_time"] = datetime.now().isoformat()
            result["cart_write_message"] = f"购物车未匹配，匹配方式={reason}，分数={score}"
            return {"matched": False, "score": score, "reason": reason}
        result["actual_ysb_code"] = self._cart_text(cart_item.get("wholesaleId")) or result.get("actual_ysb_code", "")
        result["purchase_product"] = self._cart_text(cart_item.get("name") or cart_item.get("drugName"))
        result["purchase_spec"] = self._cart_text(cart_item.get("spec"))
        result["purchase_maker"] = self._cart_text(cart_item.get("manufacturer") or cart_item.get("maker"))
        result["purchase_valid_date"] = self._cart_text(cart_item.get("validDate"))
        cart_supplier = self._cart_text(cart_item.get("supplier"))
        candidate_supplier = self._cart_text(adapter_result.get("candidateSupplier") or adapter_item.get("supplier"))
        result["purchase_supplier"] = candidate_supplier if self._is_internal_cart_supplier(cart_supplier) and candidate_supplier else cart_supplier
        price = cart_item.get("price")
        amount = cart_item.get("amount")
        quantity = "" if amount is None else str(amount)
        result["purchase_price"] = "" if price is None else str(price)
        result["purchase_quantity_result"] = quantity
        result["actual_purchase_quantity"] = quantity
        result["purchase_quantity"] = quantity
        result["sync_purchase_quantity"] = True
        if result.get("cart_write_status") not in ("WRITTEN", "ALREADY_EXISTS"):
            result["cart_write_status"] = "ALREADY_EXISTS"
        result["cart_write_time"] = datetime.now().isoformat()
        result["cart_write_message"] = f"购物车反写成功，匹配方式={reason}，分数={score}"
        result["cart_item_id"] = result.get("actual_ysb_code", "")
        return {
            "matched": True,
            "score": score,
            "reason": reason,
            "cartKey": self._cart_item_key(cart_item),
            "wholesaleId": result.get("actual_ysb_code", ""),
            "supplier": result.get("purchase_supplier", ""),
        }
    
    def _save_cart_backfill_match(
        self,
        batch_id: str,
        purchase_detail_id: str,
        match_status: str,
        match_type: str,
        match_score: int,
        match_detail: Dict,
        snapshot_batch_id: str = "",
        rule_snapshot_id: str = ""
    ) -> None:
        """保存购物车反写匹配过程到数据库（方案要求）"""
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            # 使用传入的 snapshot_batch_id，或生成一个新的
            if not snapshot_batch_id:
                snapshot_batch_id = f"snapshot_{batch_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}"
            
            # 插入购物车反写匹配记录
            cursor.execute('''
                INSERT INTO smart_cart_backfill_matches (
                    snapshot_batch_id,
                    snapshot_item_id,
                    purchase_batch_id,
                    purchase_detail_id,
                    match_type,
                    match_score,
                    match_status,
                    match_detail,
                    rule_snapshot_id,
                    created_at,
                    updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                snapshot_batch_id,  # 快照批次ID
                "",  # 快照商品ID（暂时为空，后续可以关联）
                batch_id,  # 采购批次ID
                purchase_detail_id,  # 采购明细ID
                match_type,  # 匹配方式：wholesaleId或name_spec_manufacturer
                match_score,  # 匹配分数
                match_status,  # 匹配状态：matched或unmatched
                json.dumps(match_detail, ensure_ascii=False, default=str),  # 匹配详情（原始匹配过程数据）
                rule_snapshot_id,  # 规则快照ID（二期整改）
                datetime.now().isoformat(),  # 创建时间
                datetime.now().isoformat()  # 更新时间
            ))
            
            conn.commit()
            logging.info(f"✓ 已保存购物车反写匹配: batch_id={batch_id}, detail_id={purchase_detail_id}, status={match_status}, type={match_type}, score={match_score}")
            
        except Exception as e:
            logging.warning(f"保存购物车反写匹配失败: {e}")
            try:
                conn.rollback()
            except:
                pass

    def _reconcile_successful_cart_results(
        self,
        batch_id: str,
        successful_results: List[Tuple[str, Dict, Dict, Dict]],
        trace_path: str,
        progress_callback: Callable[[str], None] = None,
    ) -> None:
        """Run one end-of-batch snapshot without rewriting purchase result fields."""
        snapshot_batch_id = (
            f"snapshot_{batch_id}_final_{datetime.now().strftime('%Y%m%d%H%M%S')}"
        )
        cart_items = self._run_cart_snapshot(
            batch_id,
            trace_path,
            progress_callback,
            snapshot_batch_id,
        )
        self._append_purchase_trace(
            trace_path,
            f"CART_FINAL_SNAPSHOT count={len(cart_items)} "
            f"success_count={len(successful_results)} snapshot_batch_id={snapshot_batch_id}",
            progress_callback,
        )
        if not cart_items:
            self._append_purchase_trace(
                trace_path,
                "CART_FINAL_RECONCILE_SKIPPED empty_snapshot; keep_node_verified_state=true",
                progress_callback,
            )
            return

        matched_cart_keys = set()
        for item_id, adapter_item, adapter_result, result in successful_results:
            cart_item, score, reason = self._find_cart_backfill_item(
                adapter_item,
                result,
                cart_items,
                excluded_cart_keys=matched_cart_keys,
            )
            if not cart_item:
                self._save_cart_writeback_state(
                    item_id,
                    "NOT_FOUND",
                    "批次结束复核未在购物车找到该商品；请人工确认购物车",
                    result.get("cart_item_id", ""),
                    int(result.get("cart_write_attempts") or 0),
                )
                self._append_purchase_trace(
                    trace_path,
                    f"CART_FINAL_RECONCILE item_id={item_id} matched=false score={score} reason={reason}",
                    progress_callback,
                )
                continue

            cart_key = self._cart_item_key(cart_item)
            if cart_key:
                matched_cart_keys.add(cart_key)
            original_status = result.get("cart_write_status", "WRITTEN")
            final_status = (
                "ALREADY_EXISTS" if original_status == "ALREADY_EXISTS" else "WRITTEN"
            )
            cart_item_id = self._cart_text(cart_item.get("wholesaleId"))
            self._save_cart_writeback_state(
                item_id,
                final_status,
                f"批次结束购物车复核通过，匹配方式={reason}，分数={score}",
                cart_item_id,
                int(result.get("cart_write_attempts") or 0),
            )
            self._append_purchase_trace(
                trace_path,
                f"CART_FINAL_RECONCILE item_id={item_id} matched=true "
                f"status={final_status} score={score} reason={reason} cart_item_id={cart_item_id}",
                progress_callback,
            )

    def _cart_extra_business_key(self, cart_item: Dict) -> str:
        source = self._cart_item_key(cart_item) or str(uuid.uuid4())
        return f"CART_EXTRA_{uuid.uuid5(uuid.NAMESPACE_URL, source).hex[:16]}"

    def _upsert_cart_extra_item(self, batch_id: str, cart_item: Dict) -> Tuple[int, bool]:
        conn = self.db.get_connection()
        cursor = conn.cursor()
        business_key = self._cart_extra_business_key(cart_item)
        now = datetime.now().isoformat()
        raw_data = json.dumps({"source": "cart_backfill_extra", "cart_item": cart_item}, ensure_ascii=False)
        normalized_data = json.dumps({
            "source_name": cart_item.get("name", ""),
            "source_spec": cart_item.get("spec", ""),
            "source_maker": cart_item.get("manufacturer") or cart_item.get("maker") or "",
            "purchase_quantity": cart_item.get("amount", ""),
            "ysb_code": cart_item.get("wholesaleId", ""),
        }, ensure_ascii=False)
        cursor.execute('''
            SELECT item_id, row_number
            FROM smart_purchase_items
            WHERE batch_id = ? AND business_key = ?
        ''', (batch_id, business_key))
        old = cursor.fetchone()
        if old:
            row_number = int(old["row_number"])
            item_id = old["item_id"]
            created = False
        else:
            cursor.execute("SELECT COALESCE(MAX(row_number), 1) + 1 FROM smart_purchase_items WHERE batch_id = ?", (batch_id,))
            row_number = int(cursor.fetchone()[0])
            item_id = f"{batch_id}_cart_extra_{row_number}_{uuid.uuid4().hex[:8]}"
            created = True

        values = (
            batch_id,
            row_number,
            business_key,
            "",
            self._cart_text(cart_item.get("name")),
            self._cart_text(cart_item.get("spec")),
            self._cart_text(cart_item.get("manufacturer") or cart_item.get("maker")),
            "",
            self._cart_text(cart_item.get("amount")),
            "",
            self._cart_text(cart_item.get("name")),
            self._cart_text(cart_item.get("spec")),
            "",
            self._cart_text(cart_item.get("supplier")),
            "",
            "",
            "",
            self._cart_text(cart_item.get("amount")),
            self._cart_text(cart_item.get("price")),
            "1",
            "cart_extra",
            self._cart_text(cart_item.get("wholesaleId")),
            self._cart_text(cart_item.get("wholesaleId")),
            "",
            raw_data,
            normalized_data,
            "valid",
            "",
            "success",
            self._cart_text(cart_item.get("supplier")),
            self._cart_text(cart_item.get("name")),
            self._cart_text(cart_item.get("spec")),
            self._cart_text(cart_item.get("manufacturer") or cart_item.get("maker")),
            self._cart_text(cart_item.get("validDate")),
            self._cart_text(cart_item.get("amount")),
            self._cart_text(cart_item.get("price")),
            "",
            "",
            1,
            now,
            now,
            now,
            item_id,
            "ALREADY_EXISTS",
            now,
            "购物车额外商品已登记",
            self._cart_text(cart_item.get("wholesaleId")),
            0,
        )

        if created:
            cursor.execute('''
                INSERT INTO smart_purchase_items
                (batch_id, row_number, business_key, item_code,
                 source_name, source_spec, source_maker, source_approval,
                 purchase_quantity, expected_price, smart_name, smart_spec,
                 smart_approval, smart_supplier, smart_supplier_full,
                 min_purchase_quantity, available_stock, actual_purchase_quantity,
                 smart_price, selected, activity_type, ysb_code, actual_ysb_code, barcode,
                 raw_data, normalized_data, import_status, validation_message,
                 purchase_status, purchase_supplier, purchase_product, purchase_spec,
                 purchase_maker, purchase_valid_date, purchase_quantity_result, purchase_price, max_allowed_price,
                 purchase_reason, candidate_count, executed_at, created_at, updated_at, item_id,
                 cart_write_status, cart_write_time, cart_write_message, cart_item_id, cart_write_attempts)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', values)
        else:
            cursor.execute('''
                UPDATE smart_purchase_items
                SET source_name = ?,
                    source_spec = ?,
                    source_maker = ?,
                    purchase_quantity = ?,
                    actual_purchase_quantity = ?,
                    smart_name = ?,
                    smart_spec = ?,
                    smart_supplier = ?,
                    smart_price = ?,
                    selected = '1',
                    activity_type = 'cart_extra',
                    ysb_code = ?,
                    actual_ysb_code = ?,
                    raw_data = ?,
                    normalized_data = ?,
                    import_status = 'valid',
                    purchase_status = 'success',
                    purchase_supplier = ?,
                    purchase_product = ?,
                    purchase_spec = ?,
                    purchase_maker = ?,
                    purchase_valid_date = ?,
                    purchase_quantity_result = ?,
                    purchase_price = ?,
                    candidate_count = 1,
                    cart_write_status = 'ALREADY_EXISTS',
                    cart_write_time = ?,
                    cart_write_message = ?,
                    cart_item_id = ?,
                    cart_write_attempts = 0,
                    executed_at = ?,
                    updated_at = ?
                WHERE item_id = ?
            ''', (
                self._cart_text(cart_item.get("name")),
                self._cart_text(cart_item.get("spec")),
                self._cart_text(cart_item.get("manufacturer") or cart_item.get("maker")),
                self._cart_text(cart_item.get("amount")),
                self._cart_text(cart_item.get("amount")),
                self._cart_text(cart_item.get("name")),
                self._cart_text(cart_item.get("spec")),
                self._cart_text(cart_item.get("supplier")),
                self._cart_text(cart_item.get("price")),
                self._cart_text(cart_item.get("wholesaleId")),
                self._cart_text(cart_item.get("wholesaleId")),
                raw_data,
                normalized_data,
                self._cart_text(cart_item.get("supplier")),
                self._cart_text(cart_item.get("name")),
                self._cart_text(cart_item.get("spec")),
                self._cart_text(cart_item.get("manufacturer") or cart_item.get("maker")),
                self._cart_text(cart_item.get("validDate")),
                self._cart_text(cart_item.get("amount")),
                self._cart_text(cart_item.get("price")),
                now,
                "购物车额外商品已登记",
                self._cart_text(cart_item.get("wholesaleId")),
                now,
                now,
                item_id,
            ))
        conn.commit()
        return row_number, created

    def _refresh_batch_counts(self, batch_id: str):
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT COUNT(*) AS total_count,
                   SUM(CASE WHEN import_status = 'valid' THEN 1 ELSE 0 END) AS valid_count,
                   SUM(CASE WHEN import_status = 'invalid' THEN 1 ELSE 0 END) AS invalid_count
            FROM smart_purchase_items
            WHERE batch_id = ?
        ''', (batch_id,))
        row = cursor.fetchone()
        cursor.execute('''
            UPDATE smart_purchase_batches
            SET total_count = ?,
                valid_count = ?,
                invalid_count = ?
            WHERE batch_id = ?
        ''', (
            int(row["total_count"] or 0),
            int(row["valid_count"] or 0),
            int(row["invalid_count"] or 0),
            batch_id,
        ))
        conn.commit()

    def _emit_new_trace_lines(
        self,
        trace_path: str,
        start_position: int,
        progress_callback: Callable[[str], None] = None
    ) -> int:
        if not trace_path or not progress_callback:
            return start_position
        path = Path(trace_path)
        if not path.exists():
            return start_position
        try:
            with open(path, "r", encoding="utf-8", errors="ignore") as log_file:
                log_file.seek(start_position)
                lines = log_file.readlines()
                position = log_file.tell()
            for line in lines:
                message = line.strip()
                if not message:
                    continue
                if "] JS " in message:
                    message = message.split("] JS ", 1)[1]
                elif "] " in message:
                    message = message.split("] ", 1)[1]
                progress_callback(message)
            return position
        except Exception as e:
            logging.warning(f"读取智能采购实时日志失败: {e}")
            return start_position
    
    def _adapter_failure_results(self, adapter_items: List[Dict], reason: str) -> List[Dict]:
        return [
            {
                "itemId": item.get("itemId", ""),
                "rowNumber": item.get("rowNumber", ""),
                "name": item.get("name", ""),
                "wholesaleId": item.get("wholesaleId", ""),
                "status": "failed",
                "reason": reason,
            }
            for item in adapter_items
        ]

    def _complete_missing_adapter_results(
        self,
        adapter_items: List[Dict],
        adapter_results: List[Dict],
        reason: str
    ) -> List[Dict]:
        completed = list(adapter_results or [])
        seen = {str(result.get("itemId", "")) for result in completed}
        for item in adapter_items:
            item_id = str(item.get("itemId", ""))
            if item_id in seen:
                continue
            completed.append({
                "itemId": item_id,
                "rowNumber": item.get("rowNumber", ""),
                "name": item.get("name", ""),
                "wholesaleId": item.get("wholesaleId", ""),
                "status": "failed",
                "reason": reason,
                "noPurchaseInfo": True,
            })
        return completed

    def _handle_web_error_retry(
        self,
        batch_id: str,
        adapter_items: List[Dict],
        adapter_results: List[Dict],
        trace_path: str,
        progress_callback: Callable[[str], None] = None,
        web_error_callback: Callable[[str], bool] = None,
        stop_callback: Callable[[], bool] = None
    ) -> Tuple[List[Dict], bool, str]:
        combined = list(adapter_results or [])
        consecutive_web_errors = 0
        while True:
            web_result = next((result for result in combined if result.get("status") == "web_error"), None)
            if not web_result:
                return combined, False, ""
            message = web_result.get("reason", "药师帮网页异常，无法继续")
            consecutive_web_errors += 1
            self._append_purchase_trace(trace_path, f"WEB_ERROR_PAUSE reason={message}", progress_callback)
            if consecutive_web_errors >= 2:
                circuit_reason = f"连续 {consecutive_web_errors} 次药师帮连接异常，已触发熔断：{message}"
                self._append_purchase_trace(
                    trace_path,
                    f"WEB_ERROR_CIRCUIT_BREAKER count={consecutive_web_errors} reason={message}",
                    progress_callback,
                )
                return combined, True, circuit_reason
            should_continue = bool(web_error_callback and web_error_callback(message))
            if not should_continue:
                self._append_purchase_trace(trace_path, "WEB_ERROR_CANCEL user_cancelled", progress_callback)
                return combined, True, f"用户取消继续执行：{message}"
            failed_item_id = web_result.get("itemId", "")
            retry_started = False
            retry_items = []
            completed_ids = {
                result.get("itemId")
                for result in combined
                if result.get("itemId") and result.get("status") != "web_error"
            }
            for item in adapter_items:
                if item.get("itemId") == failed_item_id:
                    retry_started = True
                if retry_started and item.get("itemId") not in completed_ids:
                    retry_items.append(item)
            if not retry_items:
                self._append_purchase_trace(trace_path, "WEB_ERROR_RETRY no_remaining_items", progress_callback)
                return combined, False, ""
            if stop_callback and stop_callback():
                self._append_purchase_trace(trace_path, "WEB_ERROR_RETRY skipped_by_stop", progress_callback)
                return combined, True, "用户请求停止"
            combined = [
                result
                for result in combined
                if result.get("status") != "web_error" and result.get("itemId") not in {item.get("itemId") for item in retry_items}
            ]
            self._append_purchase_trace(trace_path, f"WEB_ERROR_RETRY remaining_count={len(retry_items)}", progress_callback)
            retry_results = self._run_cart_adapter_batch(
                batch_id,
                retry_items,
                trace_path,
                progress_callback,
                stop_callback=stop_callback,
            )
            combined.extend(retry_results)

    def _get_batch_purchase_status_counts(self, batch_id: str) -> Dict[str, int]:
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT COALESCE(purchase_status, '') AS status, COUNT(*) AS count
            FROM smart_purchase_items
            WHERE batch_id = ?
              AND import_status = 'valid'
            GROUP BY COALESCE(purchase_status, '')
        ''', (batch_id,))
        counts = {"all": 0, "pending": 0, "failed": 0, "success": 0, "other": 0}
        for row in cursor.fetchall():
            status = row["status"] or "pending"
            count = int(row["count"] or 0)
            counts["all"] += count
            if status in counts and status != "all":
                counts[status] += count
            else:
                counts["other"] += count
        return counts
    
    def _get_failure_reason_summary(self, batch_id: str) -> List[Tuple[str, int]]:
        """统计失败原因的分类和数量（方案要求：统计失败原因分类，弹窗展示失败原因 Top 3 分类摘要）"""
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT purchase_reason, cart_write_message, cart_write_status, failure_code
            FROM smart_purchase_items
            WHERE batch_id = ?
              AND purchase_status = 'failed'
        ''', (batch_id,))
        
        # 统计失败原因分类
        failure_reason_counts = {}
        for row in cursor.fetchall():
            cart_status = row["cart_write_status"] or ""
            reason = row["purchase_reason"] or ""
            if cart_status in ("FAILED_RETRYABLE", "FAILED_MANUAL"):
                reason = row["cart_write_message"] or reason
            reason_type = self._classify_failure_reason(reason, row["failure_code"] or "")
            if reason_type in failure_reason_counts:
                failure_reason_counts[reason_type] += 1
            else:
                failure_reason_counts[reason_type] = 1
        
        # 按数量排序，返回Top 3
        sorted_reasons = sorted(failure_reason_counts.items(), key=lambda x: x[1], reverse=True)
        return sorted_reasons
    
    def _classify_failure_reason(self, reason: str, failure_code: str = "") -> str:
        """分类失败原因（根据关键词进行分类）"""
        if failure_code in ("CDP_CONNECTION_FAILED", "EXECUTION_TIMEOUT"):
            return "连接或执行异常"
        if failure_code in ("LOGIN_NOT_CONFIRMED", "BROWSER_NOT_FOUND", "PAGE_CLOSED"):
            return "浏览器或登录异常"
        if not reason:
            return "未知失败原因"
        
        # 根据关键词进行分类
        if "未找到候选商品" in reason or "搜索无候选" in reason:
            return "搜索无候选"
        elif "候选匹配分数过低" in reason or "低分候选" in reason:
            return "候选匹配分数过低"
        elif "规格不一致" in reason or "规格不匹配" in reason:
            return "规格不一致"
        elif "厂家不匹配" in reason or "厂家不一致" in reason:
            return "厂家不匹配"
        elif "价格超限" in reason or "价格超过最高允许价" in reason:
            return "价格超限"
        elif "库存不足" in reason or "候选库存不足" in reason:
            return "库存不足"
        elif "起购不满足" in reason or "起购数量" in reason:
            return "起购不满足"
        elif "加购接口返回异常" in reason or "加购接口异常" in reason:
            return "加购接口异常"
        elif "加购后购物车数量不足" in reason or "购物车数量未达到要求" in reason:
            return "加购后购物车数量不足"
        elif "浏览器异常" in reason or "未找到药师帮浏览器页面" in reason:
            return "浏览器异常"
        elif "未确认登录" in reason or "登录状态" in reason:
            return "登录异常"
        elif "执行超时" in reason or "超时" in reason:
            return "执行超时"
        elif re.search(r"fetch failed|ECONNREFUSED|ECONNRESET|连接异常", reason, re.IGNORECASE):
            return "连接或执行异常"
        elif "缺少商品名称" in reason:
            return "缺少商品名称"
        elif "采购数量无效" in reason:
            return "采购数量无效"
        elif "候选供应商不在允许范围" in reason:
            return "候选供应商不在允许范围"
        elif "未配置供应商范围" in reason:
            return "未配置供应商范围"
        else:
            return "其他失败原因"
    
    def _merge_cart_adapter_result(self, base_result: Dict, adapter_result: Dict) -> Dict:
        result = dict(base_result)
        status = adapter_result.get("status", "failed")
        if status in ("success", "skipped"):
            result["purchase_status"] = status
        else:
            result["purchase_status"] = "failed"
        # 技术执行信息使用独立字段；只有业务规则拒绝会更新采购原因。
        adapter_reason = adapter_result.get("reason", "")
        failure_category = adapter_result.get("failureCategory", "")
        if not failure_category and status not in ("success", "skipped"):
            if re.search(
                r"fetch failed|ECONNREFUSED|ECONNRESET|timeout|网络|WebSocket|未确认登录|浏览器|页面被关闭",
                adapter_reason,
                re.IGNORECASE,
            ):
                failure_category = "TECHNICAL_RETRYABLE"
            else:
                failure_category = "BUSINESS_REJECTED"
        result["failure_category"] = failure_category
        # 搜索、评分、价格、厂家等业务拒绝属于采购原因；连接和页面异常仅写购物车信息。
        if (
            status not in ("success", "skipped")
            and failure_category in ("BUSINESS_REJECTED", "NO_CANDIDATE")
            and adapter_reason
        ):
            result["purchase_reason"] = adapter_reason
        cart_write_status = adapter_result.get("cartWriteStatus", "")
        if not cart_write_status:
            if status == "success":
                cart_write_status = "WRITTEN"
            elif status == "skipped":
                cart_write_status = "SKIPPED"
            elif re.search(
                r"timeout|Runtime\.evaluate|network|fetch failed|ECONNREFUSED|ECONNRESET",
                adapter_reason,
                re.IGNORECASE,
            ):
                cart_write_status = "FAILED_RETRYABLE"
            else:
                cart_write_status = "FAILED_MANUAL"
        result["cart_write_status"] = cart_write_status
        result["cart_write_time"] = adapter_result.get("cartWriteTime") or datetime.now().isoformat()
        result["cart_write_message"] = adapter_result.get("cartWriteMessage") or adapter_reason
        result["cart_item_id"] = adapter_result.get("cartItemId") or adapter_result.get("wholesaleId", "")
        result["cart_write_attempts"] = int(adapter_result.get("cartWriteAttempts") or 0)
        # 二期整改：从Node结果中提取结构化失败编码
        if adapter_result.get("failureStage"):
            result["failure_stage"] = adapter_result.get("failureStage")
        if adapter_result.get("failureCode"):
            result["failure_code"] = adapter_result.get("failureCode")
        if adapter_result.get("failureDetail"):
            result["failure_detail"] = adapter_result.get("failureDetail")
        if adapter_result.get("suggestion"):
            result["suggestion"] = adapter_result.get("suggestion")
        if adapter_result.get("ruleSnapshotId"):
            result["rule_snapshot_id"] = adapter_result.get("ruleSnapshotId")
        if adapter_result.get("noPurchaseInfo"):
            result["purchase_supplier"] = ""
            result["purchase_product"] = ""
            result["purchase_spec"] = ""
            result["purchase_maker"] = ""
            result["purchase_quantity_result"] = ""
            result["purchase_price"] = ""
            result["actual_ysb_code"] = ""
            result["executed_at"] = datetime.now().isoformat()
            result["ready_for_cart"] = False
            return result
        if adapter_result.get("matchedSupplier"):
            result["purchase_supplier"] = adapter_result.get("matchedSupplier", "")
        if adapter_result.get("matchedName"):
            result["purchase_product"] = adapter_result.get("matchedName", "")
        if adapter_result.get("matchedSpec"):
            result["purchase_spec"] = adapter_result.get("matchedSpec", "")
        if adapter_result.get("matchedManufacturer"):
            result["purchase_maker"] = adapter_result.get("matchedManufacturer", "")
        if adapter_result.get("matchedPrice"):
            result["purchase_price"] = adapter_result.get("matchedPrice", "")
        result["purchase_quantity_result"] = str(adapter_result.get("verifiedAmount") or result.get("purchase_quantity_result") or "")
        if adapter_result.get("wholesaleId"):
            result["actual_ysb_code"] = str(adapter_result.get("wholesaleId") or "")
        result["executed_at"] = datetime.now().isoformat()
        result["ready_for_cart"] = False
        return result

    def _result_display_reason(self, result: Dict) -> str:
        if result.get("failure_category", "").startswith("TECHNICAL_"):
            return result.get("cart_write_message") or result.get("purchase_reason", "")
        return result.get("purchase_reason") or result.get("cart_write_message", "")
    
    def _prepare_cart_purchase_item(self, item: Dict, supplier_scope: List[str]) -> Dict:
        item = {key: ("" if value is None else value) for key, value in item.items()}
        now = datetime.now().isoformat()
        result = {
            "purchase_status": "failed",
            "purchase_supplier": item.get("smart_supplier") or item.get("smart_supplier_full") or "",
            "purchase_product": item.get("smart_name") or item.get("source_name") or "",
            "purchase_spec": item.get("smart_spec") or item.get("source_spec") or "",
            "purchase_maker": item.get("source_maker") or "",
            "purchase_valid_date": "",
            "purchase_quantity_result": item.get("actual_purchase_quantity") or item.get("purchase_quantity") or "",
            "purchase_price": item.get("smart_price") or "",
            "max_allowed_price": "",
            "purchase_reason": "",
            "actual_ysb_code": item.get("actual_ysb_code") or "",
            "candidate_count": 1 if (item.get("ysb_code") or item.get("source_name")) else 0,
            "executed_at": now,
            "ready_for_cart": False,
        }
        
        validation_error = self._validate_purchase_item(item)
        if validation_error:
            result["purchase_reason"] = validation_error
            return result
        
        if supplier_scope and (item.get("smart_supplier") or item.get("smart_supplier_full")):
            supplier_error = self._validate_supplier_scope(item, supplier_scope)
            if supplier_error:
                result["purchase_reason"] = supplier_error
                return result
        
        price_error, max_allowed_price = self._validate_price(item)
        result["max_allowed_price"] = str(max_allowed_price) if max_allowed_price is not None else ""
        if price_error and item.get("smart_price"):
            result["purchase_reason"] = price_error
            return result
        
        quantity_error = self._validate_min_quantity_and_stock(item)
        if quantity_error:
            result["purchase_reason"] = quantity_error
            return result
        
        result["purchase_status"] = "pending"
        result["purchase_reason"] = "已通过基础校验，等待药师帮页面按名称/规格/厂家匹配并加购"
        result["ready_for_cart"] = True
        return result
    
    def get_batch(self, batch_id: str) -> Dict:
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT batch_id, batch_name, source_file, sheet_name, supplier_scope,
                   allow_keep_cart, total_count, valid_count, invalid_count,
                   status, imported_at
            FROM smart_purchase_batches
            WHERE batch_id = ?
        ''', (batch_id,))
        row = cursor.fetchone()
        return dict(row) if row else {}
    
    def _get_executable_items(self, batch_id: str, retry_failed: bool, max_rows: int) -> List[Dict]:
        sql = '''
            SELECT item_id, batch_id, row_number, business_key, item_code,
                   source_name, source_spec, source_maker, source_approval,
                   purchase_quantity, expected_price, smart_name, smart_spec,
                   smart_approval, smart_supplier, smart_supplier_full,
                   min_purchase_quantity, available_stock, actual_purchase_quantity,
                   smart_price, selected, activity_type, ysb_code, actual_ysb_code, barcode,
                   import_status, validation_message, purchase_status, purchase_reason,
                   cart_write_status, cart_write_time, cart_write_message,
                   cart_item_id, cart_write_attempts
            FROM smart_purchase_items
            WHERE batch_id = ?
              AND import_status = 'valid'
        '''
        params = [batch_id]
        
        if retry_failed:
            sql += " AND purchase_status = 'failed'"
        
        sql += " ORDER BY row_number"
        if max_rows and max_rows > 0:
            sql += " LIMIT ?"
            params.append(max_rows)
        
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute(sql, params)
        return [dict(row) for row in cursor.fetchall()]
    
    def _execute_single_item(self, item: Dict, supplier_scope: List[str]) -> Dict:
        now = datetime.now().isoformat()
        base_result = {
            "purchase_status": "failed",
            "purchase_supplier": item.get("smart_supplier") or item.get("smart_supplier_full") or "",
            "purchase_product": item.get("smart_name") or item.get("source_name") or "",
            "purchase_spec": item.get("smart_spec") or item.get("source_spec") or "",
            "purchase_maker": item.get("source_maker") or "",
            "purchase_quantity_result": item.get("actual_purchase_quantity") or item.get("purchase_quantity") or "",
            "purchase_price": item.get("smart_price") or "",
            "max_allowed_price": "",
            "purchase_reason": "",
            "actual_ysb_code": item.get("actual_ysb_code") or item.get("ysb_code") or "",
            "candidate_count": 1 if (item.get("ysb_code") or item.get("smart_name")) else 0,
            "executed_at": now,
        }
        
        validation_error = self._validate_purchase_item(item)
        if validation_error:
            base_result["purchase_reason"] = validation_error
            return base_result
        
        supplier_error = self._validate_supplier_scope(item, supplier_scope)
        if supplier_error:
            base_result["purchase_reason"] = supplier_error
            return base_result
        
        price_error, max_allowed_price = self._validate_price(item)
        base_result["max_allowed_price"] = str(max_allowed_price) if max_allowed_price is not None else ""
        if price_error:
            base_result["purchase_reason"] = price_error
            return base_result
        
        quantity_error = self._validate_min_quantity_and_stock(item)
        if quantity_error:
            base_result["purchase_reason"] = quantity_error
            return base_result
        
        if not item.get("ysb_code") and not item.get("smart_name"):
            base_result["purchase_reason"] = "未找到满足供应商、品种、规格、厂家/批准文号、价格、起购数量的候选"
            return base_result
        
        # 这里已经完成逐个采购前的匹配和规则判断。真实购物车扫描、加购、验证由后续药师帮适配器接入。
        base_result["purchase_reason"] = "药师帮购物车执行适配器未接入，未执行加购"
        return base_result
    
    def _save_cart_writeback_state(
        self,
        item_id: str,
        status: str,
        message: str = "",
        cart_item_id: str = "",
        attempts: int = 0,
    ) -> None:
        """Update cart writeback metadata without touching purchase result fields."""
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            UPDATE smart_purchase_items
            SET cart_write_status = ?,
                cart_write_time = ?,
                cart_write_message = ?,
                cart_item_id = ?,
                cart_write_attempts = ?,
                updated_at = ?
            WHERE item_id = ?
        ''', (
            status or "NOT_STARTED",
            datetime.now().isoformat(),
            message or "",
            cart_item_id or "",
            int(attempts or 0),
            datetime.now().isoformat(),
            item_id,
        ))
        conn.commit()

    def _save_purchase_result(
        self,
        item_id: str,
        result: Dict,
        batch_id: str = "",
        rule_snapshot_id: str = "",
        rule_set_code: str = "",
        preserve_purchase_reason: bool = False,
    ):
        conn = self.db.get_connection()
        cursor = conn.cursor()
        status = result.get("purchase_status", "failed")

        # 二期整改：提取结构化失败信息
        failure_stage = result.get("failure_stage") or result.get("failureStage") or ""
        failure_code = result.get("failure_code") or result.get("failureCode") or ""

        # 如果没有结构化失败编码，从原始原因分类
        if not failure_code and status not in ("success", "skipped"):
            classified = self._failure_reason_service._classify_raw_reason(
                result.get("purchase_reason", "")
            )
            failure_stage = classified["stage"]
            failure_code = classified["code"]
            # 将分类结果回写到result中
            result["failure_stage"] = failure_stage
            result["failure_code"] = failure_code

        if status in ("success", "skipped"):
            cursor.execute('''
                UPDATE smart_purchase_items
                SET purchase_status = ?,
                    purchase_supplier = ?,
                    purchase_product = ?,
                    purchase_spec = ?,
                    purchase_maker = ?,
                    purchase_valid_date = ?,
                    purchase_quantity_result = ?,
                    actual_purchase_quantity = ?,
                    purchase_quantity = CASE WHEN ? THEN ? ELSE purchase_quantity END,
                    purchase_price = ?,
                    max_allowed_price = ?,
                    purchase_reason = CASE WHEN ? = 1 THEN purchase_reason ELSE ? END,
                    actual_ysb_code = ?,
                    candidate_count = ?,
                    executed_at = ?,
                    failure_stage = ?,
                    failure_code = ?,
                    cart_write_status = ?,
                    cart_write_time = ?,
                    cart_write_message = ?,
                    cart_item_id = ?,
                    cart_write_attempts = ?,
                    updated_at = ?
                WHERE item_id = ?
            ''', (
                status,
                result.get("purchase_supplier", ""),
                result.get("purchase_product", ""),
                result.get("purchase_spec", ""),
                result.get("purchase_maker", ""),
                result.get("purchase_valid_date", ""),
                result.get("purchase_quantity_result", ""),
                result.get("actual_purchase_quantity", result.get("purchase_quantity_result", "")),
                1 if result.get("sync_purchase_quantity") else 0,
                result.get("purchase_quantity", result.get("purchase_quantity_result", "")),
                result.get("purchase_price", ""),
                result.get("max_allowed_price", ""),
                1 if preserve_purchase_reason else 0,
                result.get("purchase_reason", ""),
                result.get("actual_ysb_code", ""),
                result.get("candidate_count", 0),
                result.get("executed_at", ""),
                failure_stage,
                failure_code,
                result.get("cart_write_status", "NOT_STARTED"),
                result.get("cart_write_time", ""),
                result.get("cart_write_message", ""),
                result.get("cart_item_id", ""),
                int(result.get("cart_write_attempts") or 0),
                datetime.now().isoformat(),
                item_id
            ))
        else:
            cursor.execute('''
                UPDATE smart_purchase_items
                SET purchase_status = ?,
                    purchase_supplier = '',
                    purchase_product = '',
                    purchase_spec = '',
                    purchase_maker = '',
                    purchase_valid_date = '',
                    purchase_quantity_result = '',
                    purchase_price = '',
                    actual_ysb_code = '',
                    purchase_reason = CASE WHEN ? = 1 THEN purchase_reason ELSE ? END,
                    executed_at = ?,
                    failure_stage = ?,
                    failure_code = ?,
                    cart_write_status = ?,
                    cart_write_time = ?,
                    cart_write_message = ?,
                    cart_item_id = ?,
                    cart_write_attempts = ?,
                    updated_at = ?
                WHERE item_id = ?
            ''', (
                status,
                1 if preserve_purchase_reason else 0,
                result.get("purchase_reason", ""),
                result.get("executed_at", ""),
                failure_stage,
                failure_code,
                result.get("cart_write_status", "NOT_STARTED"),
                result.get("cart_write_time", ""),
                result.get("cart_write_message", ""),
                result.get("cart_item_id", ""),
                int(result.get("cart_write_attempts") or 0),
                datetime.now().isoformat(),
                item_id
            ))

        # 二期整改：保存结构化失败原因到 smart_purchase_failure_reasons 表
        if failure_code and batch_id:
            # 从 item_id 中提取行号（格式：batch_id_rowNumber）
            row_number = result.get("row_number") or result.get("rowNumber") or 0
            if not row_number:
                try:
                    # item_id 格式为 batch_id_rowNumber，取最后一个下划线后的数字
                    parts = item_id.rsplit("_", 1)
                    if len(parts) == 2 and parts[1].isdigit():
                        row_number = int(parts[1])
                except Exception:
                    row_number = 0

            # 优先保存 Node 返回的 failureDetail 和 suggestion
            failure_detail = result.get("failure_detail") or result.get("failureDetail") or ""
            suggestion = result.get("suggestion") or ""

            # Node 未返回时，调用失败原因服务的分类结果生成默认详情和建议
            if not failure_detail or not suggestion:
                classified = self._failure_reason_service._classify_raw_reason(
                    result.get("purchase_reason", "")
                )
                if not failure_detail:
                    failure_detail = classified.get("detail", "")
                if not suggestion:
                    suggestion = classified.get("suggestion", "")

            self._failure_reason_service.save_failure_reason(
                batch_id=batch_id,
                item_id=item_id,
                row_number=row_number,
                failure_stage=failure_stage,
                failure_code=failure_code,
                failure_message=result.get("purchase_reason", ""),
                failure_detail=failure_detail,
                suggestion=suggestion,
                raw_reason=result.get("purchase_reason", ""),
                rule_set_code=rule_set_code,
                rule_snapshot_id=rule_snapshot_id
            )

        conn.commit()
    
    def _validate_purchase_item(self, item: Dict) -> str:
        """增强行级基础校验失败原因（方案要求：提示行号、商品名、失败环节、关键对比值和建议动作）"""
        row_number = item.get("row_number", "")
        item_name = item.get("source_name") or item.get("business_key") or ""
        
        if not item.get("source_name"):
            return f"第{row_number}行 {item_name}: 缺少商品名称。建议：请检查导入文件，确保每行都有商品名称字段。"
        
        purchase_quantity = self._to_decimal(item.get("purchase_quantity"))
        if purchase_quantity <= 0:
            return f"第{row_number}行 {item_name}: 采购数量无效（当前数量: {purchase_quantity}）。建议：请修改采购数量为大于0的数值。"
        
        return ""
    
    def _validate_supplier_scope(self, item: Dict, supplier_scope: List[str]) -> str:
        """增强供应商范围校验失败原因（方案要求：提示行号、候选供应商、允许范围、建议调整）"""
        row_number = item.get("row_number", "")
        item_name = item.get("source_name") or item.get("business_key") or ""
        
        if not supplier_scope:
            return f"第{row_number}行 {item_name}: 未配置供应商范围。建议：请在批次设置中配置允许采购的供应商范围。"
        
        supplier_text = " ".join([
            item.get("smart_supplier", ""),
            item.get("smart_supplier_full", ""),
        ]).lower()
        
        if not supplier_text.strip():
            return f"第{row_number}行 {item_name}: 未找到候选供应商。建议：请先完成药师帮商品匹配，或导入包含供应商信息的采购清单。"
        
        candidate_supplier = item.get("smart_supplier") or item.get("smart_supplier_full") or ""
        allowed_suppliers = ", ".join(supplier_scope)
        
        for supplier in supplier_scope:
            if supplier and supplier.lower() in supplier_text:
                return ""
        
        return f"第{row_number}行 {item_name}: 候选供应商不在允许范围（候选供应商: {candidate_supplier}，允许范围: {allowed_suppliers}）。建议：请调整供应商范围或选择其他供应商。"
    
    def _validate_price(self, item: Dict) -> Tuple[str, Decimal | None]:
        """增强价格校验失败原因（方案要求：提示行号、商品名、期望价、最高允许价、候选价、折后比较价、建议动作）"""
        row_number = item.get("row_number", "")
        item_name = item.get("source_name") or item.get("business_key") or ""
        
        expected_price = self._to_decimal(item.get("expected_price"))
        smart_price = self._to_decimal(item.get("smart_price"))
        
        if expected_price <= 0:
            return "", None
        
        max_allowed_price = min(
            expected_price * Decimal("1.05"),
            expected_price + Decimal("1")
        )
        
        if smart_price <= 0:
            # 增强未找到候选采购价格提示（方案要求：提示行号、商品名、建议动作）
            return f"第{row_number}行 {item_name}: 未找到候选采购价格。建议：请先完成药师帮商品匹配，或选择有价格的候选供应商。", max_allowed_price
        
        compare_price = smart_price * Decimal("0.97")
        if compare_price > max_allowed_price:
            # 增强候选价格超限提示（方案要求：提示行号、商品名、期望价、最高允许价、候选价、折后比较价、建议动作）
            return f"第{row_number}行 {item_name}: 候选价格超限（期望价: {expected_price}，最高允许价: {max_allowed_price}，候选价: {smart_price}，折后比较价: {compare_price}）。建议：请调整价格上限或选择其他供应商。", max_allowed_price
        
        return "", max_allowed_price
    
    def _validate_min_quantity_and_stock(self, item: Dict) -> str:
        """增强起购和库存校验失败原因（方案要求：提示行号、商品名、采购数量、候选起购数量、候选库存、建议动作）"""
        row_number = item.get("row_number", "")
        item_name = item.get("source_name") or item.get("business_key") or ""
        
        purchase_quantity = self._to_decimal(item.get("purchase_quantity"))
        min_purchase_quantity = self._to_decimal(item.get("min_purchase_quantity"))
        available_stock = self._to_decimal(item.get("available_stock"))
        
        if min_purchase_quantity > 0 and min_purchase_quantity > purchase_quantity:
            # 增强候选起购数量大于采购数量提示（方案要求：提示行号、商品名、采购数量、起购数量、建议动作）
            return f"第{row_number}行 {item_name}: 候选起购数量大于采购数量（采购数量: {purchase_quantity}，候选起购数量: {min_purchase_quantity}）。建议：请提高采购数量或选择其他供应商。"
        
        if available_stock > 0 and available_stock < purchase_quantity:
            # 增强候选库存不足提示（方案要求：提示行号、商品名、采购数量、库存、建议动作）
            return f"第{row_number}行 {item_name}: 候选库存不足（采购数量: {purchase_quantity}，候选库存: {available_stock}）。建议：请降低采购数量或选择其他供应商。"
        
        return ""
    
    def _parse_supplier_scope(self, supplier_scope: str) -> List[str]:
        normalized = supplier_scope.replace("，", ",").replace("；", ",").replace(";", ",").replace("\n", ",")
        return [item.strip() for item in normalized.split(",") if item.strip()]
    
    def _status_text(self, status: str) -> str:
        return {
            "success": "成功",
            "failed": "失败",
            "skipped": "跳过",
            "pending": "待处理",
        }.get(status, status)
    
    def _read_headers(self, sheet) -> List[str]:
        headers = []
        for col in range(1, sheet.max_column + 1):
            value = sheet.cell(row=1, column=col).value
            header = self._cell_to_text(value) or f"列{col}"
            headers.append(header)
        return headers
    
    def _row_to_dict(self, headers: List[str], row) -> Dict[str, str]:
        row_data = {}
        for index, header in enumerate(headers):
            value = row[index] if index < len(row) else ""
            row_data[header] = self._cell_to_text(value)
        return row_data
    
    def _row_to_dict_xls(self, headers: List[str], row_values: List) -> Dict[str, str]:
        row_data = {}
        for index, header in enumerate(headers):
            value = row_values[index] if index < len(row_values) else ""
            row_data[header] = self._cell_to_text(value)
        return row_data
    
    def _normalize_row(self, row_data: Dict[str, str]) -> Dict[str, str]:
        normalized = {}
        for field, aliases in self.FIELD_ALIASES.items():
            normalized[field] = self._first_value(row_data, aliases)
        return normalized
    
    def _validate_row(self, normalized: Dict[str, str]) -> List[str]:
        errors = []
        for field in self.REQUIRED_FIELDS:
            if not normalized.get(field):
                errors.append(f"缺少必填字段: {field}")
        
        quantity = normalized.get("purchase_quantity")
        if quantity and self._to_decimal(quantity) <= 0:
            errors.append("采购数量必须大于0")
        
        expected_price = normalized.get("expected_price")
        if expected_price and self._to_decimal(expected_price) < 0:
            errors.append("期望价格不能小于0")
        
        return errors
    
    def _first_value(self, row_data: Dict[str, str], aliases: List[str]) -> str:
        normalized_keys = {str(key).strip().lower(): key for key in row_data.keys()}
        for alias in aliases:
            if alias in row_data and row_data[alias] not in (None, ""):
                return str(row_data[alias]).strip()
            key = normalized_keys.get(alias.strip().lower())
            if key and row_data.get(key) not in (None, ""):
                return str(row_data[key]).strip()
        return ""
    
    def _cell_to_text(self, value) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()
    
    def _to_decimal(self, value) -> Decimal:
        try:
            return Decimal(str(value).replace(",", "").strip())
        except Exception:
            return Decimal("0")
    
    RESULT_FIELD_ALIASES = {
        "purchase_status": ["采购状态", "状态"],
        "purchase_supplier": ["采购供应商", "实际供应商"],
        "purchase_product": ["采购品种", "采购商品", "实际商品名称"],
        "purchase_spec": ["采购规格", "实际规格"],
        "purchase_maker": ["采购厂家", "实际厂家"],
        "purchase_valid_date": ["有效期", "采购有效期"],
        "purchase_quantity_result": ["采购结果数量"],
        "purchase_price": ["采购价格", "实际采购价"],
        "actual_ysb_code": ["药师帮编码", "实际药师帮编码", "wholesaleId"],
        "max_allowed_price": ["最高允许价", "最高允许价格"],
        "purchase_reason": ["采购原因", "原因", "备注"],
        "candidate_count": ["候选数量"],
        "executed_at": ["执行时间"],
    }
    
    def export_results(self, batch_id: str, output_path: str = None) -> Tuple[str, str]:
        batch = self.get_batch(batch_id)
        if not batch:
            return "", "采购批次不存在"
        
        source_file = batch.get("source_file", "")
        sheet_name = batch.get("sheet_name", "")
        if not source_file or not sheet_name:
            return "", "批次缺少源文件或工作表信息"
        
        items = self.get_batch_items(batch_id)
        if not items:
            return "", "批次没有采购明细数据"
        
        try:
            if self._is_xls_file(source_file):
                xls_workbook = xlrd.open_workbook(source_file)
                if sheet_name not in xls_workbook.sheet_names():
                    return "", f"工作表不存在: {sheet_name}"
                
                xls_sheet = xls_workbook.sheet_by_name(sheet_name)
                headers = []
                for col in range(xls_sheet.ncols):
                    value = xls_sheet.cell_value(0, col)
                    headers.append(self._cell_to_text(value) or f"列{col + 1}")
                
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = sheet_name
                
                for col_idx, header in enumerate(headers, start=1):
                    sheet.cell(row=1, column=col_idx, value=header)
                
                for row_idx in range(1, xls_sheet.nrows):
                    for col_idx in range(len(headers)):
                        value = xls_sheet.cell_value(row_idx, col_idx)
                        sheet.cell(row=row_idx + 1, column=col_idx + 1, value=self._cell_to_text(value))
            else:
                workbook = openpyxl.load_workbook(source_file)
                if sheet_name not in workbook.sheetnames:
                    workbook.close()
                    return "", f"工作表不存在: {sheet_name}"
                
                sheet = workbook[sheet_name]
                headers = self._read_headers(sheet)
            
            result_col_map = {}
            for field, aliases in self.RESULT_FIELD_ALIASES.items():
                for alias in aliases:
                    for col_idx, header in enumerate(headers, start=1):
                        if header.strip().lower() == alias.strip().lower():
                            result_col_map[field] = col_idx
                            break
                    if field in result_col_map:
                        break
            
            new_result_cols = {}
            for field in self.RESULT_FIELD_ALIASES.keys():
                if field not in result_col_map:
                    new_col = len(headers) + len(new_result_cols) + 1
                    new_result_cols[field] = new_col
                    header_name = self.RESULT_FIELD_ALIASES[field][0]
                    sheet.cell(row=1, column=new_col, value=header_name)
            
            all_result_cols = {**result_col_map, **new_result_cols}
            
            for col_idx in all_result_cols.values():
                for row_idx in range(2, sheet.max_row + 1):
                    sheet.cell(row=row_idx, column=col_idx, value="")
            
            item_by_row = {item.get("row_number"): item for item in items}
            cart_extra_items = [
                item for item in items
                if str(item.get("activity_type") or "").strip() == "cart_extra"
                or str(item.get("business_key") or "").startswith("CART_EXTRA_")
            ]
            
            for row_idx in range(2, sheet.max_row + 1):
                item = item_by_row.get(row_idx)
                if not item:
                    continue
                
                row_values = {
                    "purchase_status": self._status_text(item.get("purchase_status", "")),
                    "purchase_supplier": item.get("purchase_supplier", ""),
                    "purchase_product": item.get("purchase_product", ""),
                    "purchase_spec": item.get("purchase_spec", ""),
                    "purchase_maker": item.get("purchase_maker", ""),
                    "purchase_valid_date": item.get("purchase_valid_date", ""),
                    "purchase_quantity_result": item.get("purchase_quantity_result", ""),
                    "purchase_price": item.get("purchase_price", ""),
                    "actual_ysb_code": item.get("actual_ysb_code", ""),
                    "max_allowed_price": item.get("max_allowed_price", ""),
                    "purchase_reason": item.get("purchase_reason", ""),
                    "candidate_count": str(item.get("candidate_count", 0)),
                    "executed_at": item.get("executed_at", ""),
                }
                
                for field, col_idx in all_result_cols.items():
                    value = row_values.get(field, "")
                    sheet.cell(row=row_idx, column=col_idx, value=value)

            source_col_map = {}
            for field, aliases in self.FIELD_ALIASES.items():
                for alias in aliases:
                    for col_idx, header in enumerate(headers, start=1):
                        if header.strip().lower() == alias.strip().lower():
                            source_col_map[field] = col_idx
                            break
                    if field in source_col_map:
                        break

            def write_purchase_result_row(row_idx: int, item: Dict):
                row_values = {
                    "purchase_status": self._status_text(item.get("purchase_status", "")),
                    "purchase_supplier": item.get("purchase_supplier", ""),
                    "purchase_product": item.get("purchase_product", ""),
                    "purchase_spec": item.get("purchase_spec", ""),
                    "purchase_maker": item.get("purchase_maker", ""),
                    "purchase_valid_date": item.get("purchase_valid_date", ""),
                    "purchase_quantity_result": item.get("purchase_quantity_result", ""),
                    "purchase_price": item.get("purchase_price", ""),
                    "actual_ysb_code": item.get("actual_ysb_code", ""),
                    "max_allowed_price": item.get("max_allowed_price", ""),
                    "purchase_reason": item.get("purchase_reason", ""),
                    "candidate_count": str(item.get("candidate_count", 0)),
                    "executed_at": item.get("executed_at", ""),
                }
                for field, col_idx in all_result_cols.items():
                    sheet.cell(row=row_idx, column=col_idx, value=row_values.get(field, ""))

            def write_source_row(row_idx: int, item: Dict):
                row_values = {
                    "item_code": item.get("item_code", ""),
                    "source_name": item.get("source_name", "") or item.get("purchase_product", ""),
                    "source_spec": item.get("source_spec", "") or item.get("purchase_spec", ""),
                    "source_maker": item.get("source_maker", "") or item.get("purchase_maker", ""),
                    "purchase_quantity": item.get("purchase_quantity", "") or item.get("purchase_quantity_result", ""),
                    "actual_purchase_quantity": item.get("purchase_quantity_result", "") or item.get("purchase_quantity", ""),
                    "smart_name": item.get("purchase_product", "") or item.get("source_name", ""),
                    "smart_spec": item.get("purchase_spec", "") or item.get("source_spec", ""),
                    "smart_supplier": item.get("purchase_supplier", "") or item.get("smart_supplier", ""),
                    "smart_price": item.get("purchase_price", "") or item.get("smart_price", ""),
                    "activity_type": "cart_extra",
                    "ysb_code": item.get("actual_ysb_code", "") or item.get("ysb_code", ""),
                }
                for field, col_idx in source_col_map.items():
                    value = row_values.get(field)
                    if value not in (None, ""):
                        sheet.cell(row=row_idx, column=col_idx, value=value)

            for item in cart_extra_items:
                row_idx = sheet.max_row + 1
                write_source_row(row_idx, item)
                write_purchase_result_row(row_idx, item)
            
            summary_sheet_name = "采购结果汇总"
            if summary_sheet_name in workbook.sheetnames:
                del workbook[summary_sheet_name]
            
            summary_sheet = workbook.create_sheet(title=summary_sheet_name)
            status_counts = self._get_batch_purchase_status_counts(batch_id)
            
            summary_data = [
                ["采购结果汇总报告", ""],
                ["", ""],
                ["源文件", Path(source_file).name],
                ["工作表", sheet_name],
                ["导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                ["", ""],
                ["统计项目", "数值"],
                ["", ""],
                ["总行数", status_counts.get("all", 0)],
                ["成功行数", status_counts.get("success", 0)],
                ["失败行数", status_counts.get("failed", 0)],
                ["待处理行数", status_counts.get("pending", 0)],
                ["其他状态行数", status_counts.get("other", 0)],
            ]
            
            for row_idx, row_data in enumerate(summary_data, start=1):
                for col_idx, value in enumerate(row_data, start=1):
                    summary_sheet.cell(row=row_idx, column=col_idx, value=value)
            
            summary_sheet.column_dimensions["A"].width = 20
            summary_sheet.column_dimensions["B"].width = 40
            
            if output_path:
                output_file = output_path
            else:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_dir = Path(source_file).parent
                output_file = str(output_dir / f"{Path(source_file).stem}_采购结果_{timestamp}.xlsx")
            
            workbook.save(output_file)
            workbook.close()
            
            logging.info(f"智能采购结果已导出: {output_file}")
            return output_file, ""
        except Exception as e:
            logging.error(f"导出智能采购结果失败: {e}")
            return "", str(e)
