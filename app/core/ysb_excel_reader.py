import logging
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Optional
import json

import openpyxl


@dataclass
class YsbProductItem:
    raw_row_index: int
    ysb_order_no: str = ""
    ysb_store_name: str = ""
    ysb_supplier_name: str = ""
    ysb_company_name: str = ""
    purchase_time: Optional[datetime] = None
    product_name: str = ""
    manufacturer: str = ""
    spec: str = ""
    unit: str = ""
    approval_number: str = ""
    barcode: str = ""
    batch_no: str = ""
    production_date: Optional[datetime] = None
    expiry_date: Optional[datetime] = None
    unit_price: Decimal = Decimal("0")
    discount_price: Decimal = Decimal("0")
    order_quantity: Decimal = Decimal("0")
    refund_quantity: Decimal = Decimal("0")
    quantity: Decimal = Decimal("0")
    discount_amount: Decimal = Decimal("0")
    total_amount: Decimal = Decimal("0")
    actual_payment_amount: Decimal = Decimal("0")
    freight: Decimal = Decimal("0")
    discount_amount_total: Decimal = Decimal("0")
    order_type: str = ""
    raw_data: dict = field(default_factory=dict)

    def calculate_net_quantity(self):
        self.quantity = self.order_quantity - self.refund_quantity


@dataclass
class YsbSupplierSummary:
    raw_row_index: int
    ysb_supplier_name: str = ""
    ysb_company_name: str = ""
    actual_payment_amount: Decimal = Decimal("0")
    order_count: int = 1
    raw_data: dict = field(default_factory=dict)
    
    @property
    def supplier_display_name(self) -> str:
        return self.ysb_company_name or self.ysb_supplier_name


@dataclass
class YsbExcelData:
    file_path: str
    sheet_name: str = ""
    sheet_type: str = "auto"
    total_rows: int = 0
    items: list[YsbProductItem] = field(default_factory=list)
    supplier_summaries: list[YsbSupplierSummary] = field(default_factory=list)
    error_message: str = ""


class YsbExcelReader:
    TARGET_SHEET_NAMES = ["本月支付账单明细", "明细"]
    SUPPLIER_SHEET_NAMES = ["本月支付订单", "汇总"]

    COLUMN_MAPPING = {
        "订单号": "ysb_order_no",
        "订单类型": "order_type",
        "药店名称": "ysb_store_name",
        "供应商": "ysb_supplier_name",
        "企业名称": "ysb_company_name",
        "采购时间": "purchase_time",
        "商品名称": "product_name",
        "厂家": "manufacturer",
        "规格": "spec",
        "单位": "unit",
        "批准文号": "approval_number",
        "条形码": "barcode",
        "批号": "batch_no",
        "生产日期": "production_date",
        "有效期": "expiry_date",
        "单价": "unit_price",
        "折后价": "discount_price",
        "下单数量": "order_quantity",
        "退款数量": "refund_quantity",
        "折后价总额": "discount_amount",
        "商品总金额": "total_amount",
        "运费": "freight",
        "优惠金额": "discount_amount_total",
    }

    def __init__(self, file_path: str, sheet_name: str = None, sheet_type: str = "detail"):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.sheet_type = sheet_type
        self.workbook = None
        self.sheet = None
        self.header_row = []
        self.column_index_map = {}

    def read(self) -> YsbExcelData:
        result = YsbExcelData(file_path=self.file_path, sheet_type=self.sheet_type)

        try:
            self.workbook = openpyxl.load_workbook(self.file_path, read_only=False, data_only=True)
        except Exception as e:
            result.error_message = f"无法打开Excel文件: {str(e)}"
            logging.error(result.error_message)
            return result

        sheet_names = self.workbook.sheetnames
        target_sheet = self._resolve_target_sheet(sheet_names, result)
        if not target_sheet:
            self.workbook.close()
            return result

        self.sheet = self.workbook[target_sheet]
        result.sheet_name = target_sheet

        try:
            self._parse_header()
            if self.sheet_type == "supplier":
                result.supplier_summaries = self._parse_supplier_data()
                result.total_rows = len(result.supplier_summaries)
            else:
                result.items = self._parse_data()
                result.total_rows = len(result.items)
        except Exception as e:
            result.error_message = f"解析数据失败: {str(e)}"
            logging.error(result.error_message)
        finally:
            self.workbook.close()

        logging.info(f"读取药师帮Excel完成，共 {result.total_rows} 条记录")
        return result

    def _resolve_target_sheet(self, sheet_names: list[str], result: YsbExcelData) -> Optional[str]:
        if self.sheet_name:
            if self.sheet_name in sheet_names:
                return self.sheet_name
            result.error_message = f"未找到指定工作表: {self.sheet_name}\n可用工作表: {', '.join(sheet_names)}"
            logging.error(result.error_message)
            return None

        target_names = self.SUPPLIER_SHEET_NAMES if self.sheet_type == "supplier" else self.TARGET_SHEET_NAMES
        for name in target_names:
            if name in sheet_names:
                return name

        result.error_message = (
            f"未找到对账数据工作表\n可用工作表: {', '.join(sheet_names)}\n"
            f"支持的表名: {', '.join(target_names)}"
        )
        logging.error(result.error_message)
        return None

    def _parse_header(self):
        for row in self.sheet.iter_rows(min_row=1, max_row=1, values_only=True):
            self.header_row = [str(cell).strip() if cell else "" for cell in row]
            break
        
        logging.info(f"===========================================")
        logging.info(f"Excel表头 (共{len(self.header_row)}列): {self.header_row}")
        logging.info(f"工作表类型: {self.sheet_type}")
        logging.info(f"===========================================")

        for idx, header in enumerate(self.header_row):
            if not header:
                continue
            
            if header in self.COLUMN_MAPPING:
                field_name = self.COLUMN_MAPPING[header]
                self.column_index_map[field_name] = idx
                logging.info(f"  ✓ 列{idx} [{header}] -> {field_name}")
            else:
                logging.debug(f"  - 列{idx} [{header}] -> 未匹配")

        logging.info(f"列映射结果 ({len(self.column_index_map)}个字段): {self.column_index_map}")
        
        required_fields = ["discount_amount", "discount_price", "refund_quantity"]
        missing_fields = [f for f in required_fields if f not in self.column_index_map]
        if missing_fields:
            logging.warning(f"⚠️ 缺少计算实际金额所需的字段: {missing_fields}")
            logging.warning("   实际金额 = 折后价总额 - 折后价 × 退款数量")
        
        logging.info(f"Excel前5行数据预览:")
        preview_count = min(6, self.sheet.max_row)
        for row_num in range(1, preview_count + 1):
            row_data = []
            for r in self.sheet.iter_rows(min_row=row_num, max_row=row_num, values_only=True):
                row_data = [str(cell)[:20] if cell else "" for cell in r]
                break
            
            if row_num == 1:
                logging.info(f"  行{row_num} (表头): {row_data}")
            else:
                supplier_col = self.column_index_map.get("ysb_company_name", 
                            self.column_index_map.get("ysb_supplier_name", -1))
                supplier_val = row_data[supplier_col] if supplier_col >= 0 and supplier_col < len(row_data) else "N/A"
                
                discount_amount_col = self.column_index_map.get("discount_amount", -1)
                discount_price_col = self.column_index_map.get("discount_price", -1)
                refund_quantity_col = self.column_index_map.get("refund_quantity", -1)
                
                discount_amount_val = row_data[discount_amount_col] if discount_amount_col >= 0 and discount_amount_col < len(row_data) else "0"
                discount_price_val = row_data[discount_price_col] if discount_price_col >= 0 and discount_price_col < len(row_data) else "0"
                refund_quantity_val = row_data[refund_quantity_col] if refund_quantity_col >= 0 and refund_quantity_col < len(row_data) else "0"
                
                logging.info(f"  行{row_num}: 供应商=[{supplier_val}], 折后价总额=[{discount_amount_val}], 折后价=[{discount_price_val}], 退款数量=[{refund_quantity_val}]")

    def _parse_data(self) -> list[YsbProductItem]:
        items = []
        if not self.column_index_map:
            logging.warning("未匹配到任何列，尝试按默认顺序读取")
            self._auto_map_columns()

        for row_idx, row in enumerate(self.sheet.iter_rows(min_row=2, values_only=True), start=2):
            item = YsbProductItem(raw_row_index=row_idx)
            has_data = False
            
            item.raw_data = {}
            for col_idx, header in enumerate(self.header_row):
                if col_idx < len(row):
                    value = row[col_idx]
                    if value is not None:
                        item.raw_data[header] = value
                        has_data = True
            
            for field_name, col_idx in self.column_index_map.items():
                if col_idx < len(row):
                    value = row[col_idx]
                    has_data = has_data or value is not None
                    self._set_field_value(item, field_name, value)

            item.calculate_net_quantity()
            if has_data and (item.product_name or item.ysb_order_no):
                items.append(item)

        return items

    def _parse_supplier_data(self) -> list[YsbSupplierSummary]:
        items = []
        if not self.column_index_map:
            logging.warning("未匹配到任何列，尝试按默认顺序读取供应商数据")
            self._auto_map_supplier_columns()
        
        logging.info(f"===========================================")
        logging.info(f"供应商数据解析 - 列映射: {self.column_index_map}")
        logging.info(f"工作表名: {self.sheet.title}")
        logging.info(f"表头行: {self.header_row}")
        
        actual_payment_field = '实际支付金额(已减退款)'
        
        for row_idx, row in enumerate(self.sheet.iter_rows(min_row=2, values_only=True), start=2):
            item = YsbSupplierSummary(raw_row_index=row_idx)
            has_data = False
            
            item.raw_data = {}
            for col_idx, header in enumerate(self.header_row):
                if col_idx < len(row):
                    value = row[col_idx]
                    if value is not None:
                        item.raw_data[header] = value
                        has_data = True
            
            if "ysb_company_name" in self.column_index_map:
                col_idx = self.column_index_map["ysb_company_name"]
                if col_idx < len(row) and row[col_idx]:
                    item.ysb_company_name = str(row[col_idx]).strip()
                    has_data = True
            
            if "ysb_supplier_name" in self.column_index_map:
                col_idx = self.column_index_map["ysb_supplier_name"]
                if col_idx < len(row) and row[col_idx]:
                    item.ysb_supplier_name = str(row[col_idx]).strip()
                    has_data = True
            
            actual_amount = Decimal("0")
            
            if actual_payment_field in item.raw_data:
                raw_value = item.raw_data[actual_payment_field]
                actual_amount = self._parse_decimal(raw_value) or Decimal("0")
            
            item.actual_payment_amount = actual_amount
            
            supplier_display = item.ysb_company_name or item.ysb_supplier_name or "未知"
            
            if row_idx <= 5:
                logging.info(f"  ✓ 行{row_idx} [{supplier_display}]: 实际支付金额={actual_amount}")

            if item.supplier_display_name:
                items.append(item)

        total_amount = sum(item.actual_payment_amount for item in items)
        non_zero_count = sum(1 for item in items if item.actual_payment_amount > Decimal("0"))
        
        logging.info(f"✓ 供应商数据解析完成:")
        logging.info(f"   总记录数: {len(items)}")
        logging.info(f"   有金额记录: {non_zero_count}")
        logging.info(f"   零金额记录: {len(items) - non_zero_count}")
        logging.info(f"   总金额: {total_amount}")

        return items

    def _auto_map_columns(self):
        field_order = [
            "ysb_order_no", "order_type", "ysb_store_name", "ysb_supplier_name",
            "ysb_company_name", "purchase_time", "product_name", "manufacturer",
            "spec", "unit", "approval_number", "barcode", "batch_no",
            "production_date", "expiry_date", "unit_price", "discount_price",
            "order_quantity", "refund_quantity", "discount_amount", "total_amount",
            "freight", "discount_amount_total",
        ]
        for idx in range(min(len(self.header_row), len(field_order))):
            if self.header_row[idx]:
                self.column_index_map[field_order[idx]] = idx
        logging.info(f"自动列映射: {self.column_index_map}")

    def _auto_map_supplier_columns(self):
        field_order = ["ysb_supplier_name", "ysb_company_name", "actual_payment_amount"]
        for idx in range(min(len(self.header_row), len(field_order))):
            if self.header_row[idx]:
                self.column_index_map[field_order[idx]] = idx
        logging.info(f"供应商列映射: {self.column_index_map}")

    def _set_field_value(self, item: YsbProductItem, field_name: str, value):
        if value is None:
            return

        if field_name in [
            "unit_price", "discount_price", "order_quantity", "refund_quantity",
            "discount_amount", "total_amount", "freight", "discount_amount_total",
        ]:
            setattr(item, field_name, self._parse_decimal(value))
        elif field_name in ["purchase_time", "production_date", "expiry_date"]:
            setattr(item, field_name, self._parse_datetime(value))
        elif field_name == "barcode":
            setattr(item, field_name, self._parse_barcode(value))
        else:
            setattr(item, field_name, str(value).strip())

    def _parse_decimal(self, value) -> Decimal:
        if value is None:
            return Decimal("0")
        try:
            if isinstance(value, (int, float)):
                return Decimal(str(value))
            return Decimal(str(value).replace(",", "").strip())
        except (InvalidOperation, ValueError):
            return Decimal("0")

    def _parse_datetime(self, value) -> Optional[datetime]:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y.%m.%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
            try:
                return datetime.strptime(str(value).strip(), fmt)
            except ValueError:
                continue
        return None

    def _parse_barcode(self, value) -> str:
        if value is None:
            return ""
        if isinstance(value, float):
            return str(int(value))
        return str(value).strip()

    @staticmethod
    def get_sheet_names(file_path: str) -> tuple[list[str], str]:
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheets = wb.sheetnames
            wb.close()
            return sheets, ""
        except Exception as e:
            return [], str(e)


YsbItem = YsbProductItem
YsbDetailItem = YsbProductItem
