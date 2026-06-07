import logging
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.config import DATA_DIR
from app.core.reconciliation_engine import (
    SupplierReconResult,
    DetailMatchResult,
    ProductReconResult,
    ReconciliationSummary
)


class ResultExporter:
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    DIFF_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    MATCH_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    SUSPECTED_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def __init__(self):
        self.workbook = None

    def export(
        self,
        supplier_results: list[SupplierReconResult],
        detail_results: list[DetailMatchResult],
        product_results: list[ProductReconResult],
        summary: ReconciliationSummary,
        account_period: str,
        ysb_file: str = "",
        inbound_range: tuple[str, str] = ("", ""),
        output_dir: str = None,
        recon_type: str = "all"
    ) -> str:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if recon_type == "supplier":
            filename = f"药师帮供应商对账结果_{account_period}_{timestamp}.xlsx"
        elif recon_type == "supplier_product":
            filename = f"药师帮供应商商品对账结果_{account_period}_{timestamp}.xlsx"
        else:
            filename = f"药师帮入库对账结果_{account_period}_{timestamp}.xlsx"
        
        if output_dir:
            output_path = Path(output_dir) / filename
        else:
            output_path = DATA_DIR / filename

        self.workbook = openpyxl.Workbook()
        
        if recon_type == "supplier":
            self._create_supplier_summary_sheet(summary, account_period, ysb_file, inbound_range)
            self._create_supplier_sheet(supplier_results)
            self._create_supplier_diff_sheet(supplier_results)
        elif recon_type == "supplier_product":
            self._create_product_summary_sheet(summary, account_period, ysb_file, inbound_range)
            self._create_detail_sheet(detail_results)
            self._create_product_sheet(product_results)
            self._create_product_diff_sheet(detail_results, product_results)
        else:
            self._create_summary_sheet(summary, account_period, ysb_file, inbound_range)
            self._create_supplier_sheet(supplier_results)
            self._create_detail_sheet(detail_results)
            self._create_product_sheet(product_results)
            self._create_diff_sheet(supplier_results, detail_results, product_results)
        
        self.workbook.save(output_path)
        self.workbook.close()

        logging.info(f"对账结果已导出: {output_path}")
        return str(output_path)

    def _create_supplier_summary_sheet(
        self,
        summary: ReconciliationSummary,
        account_period: str,
        ysb_file: str,
        inbound_range: tuple[str, str]
    ):
        ws = self.workbook.active
        ws.title = "汇总"

        data = [
            ["供应商对账汇总报告", ""],
            ["", ""],
            ["药师帮账期", account_period],
            ["药师帮文件", ysb_file],
            ["入库查询开始日期", inbound_range[0]],
            ["入库查询结束日期", inbound_range[1]],
            ["", ""],
            ["统计项目", "数值"],
            ["", ""],
            ["药师帮供应商数", summary.ysb_supplier_count],
            ["入库供应商数", summary.inbound_supplier_count],
            ["供应商金额一致数", summary.supplier_match_count],
            ["供应商金额差异数", summary.supplier_diff_count],
            ["", ""],
            ["药师帮总金额", float(summary.ysb_total_amount)],
            ["入库总金额", float(summary.inbound_total_amount)],
            ["总金额差异", float(summary.inbound_total_amount - summary.ysb_total_amount)],
        ]

        for row_idx, row_data in enumerate(data, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.BORDER
                if row_idx == 1:
                    cell.font = Font(bold=True, size=14)
                elif row_idx == 8:
                    cell.fill = self.HEADER_FILL
                    cell.font = self.HEADER_FONT

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 40

    def _create_product_summary_sheet(
        self,
        summary: ReconciliationSummary,
        account_period: str,
        ysb_file: str,
        inbound_range: tuple[str, str]
    ):
        ws = self.workbook.active
        ws.title = "汇总"

        data = [
            ["供应商商品对账汇总报告", ""],
            ["", ""],
            ["药师帮账期", account_period],
            ["药师帮文件", ysb_file],
            ["入库查询开始日期", inbound_range[0]],
            ["入库查询结束日期", inbound_range[1]],
            ["", ""],
            ["统计项目", "数值"],
            ["", ""],
            ["药师帮商品行数", summary.ysb_row_count],
            ["入库系统商品行数", summary.inbound_row_count],
            ["", ""],
            ["成功匹配商品行数", summary.detail_matched_count],
            ["疑似匹配商品行数", summary.detail_suspected_count],
            ["未匹配商品行数", summary.detail_unmatched_count],
            ["", ""],
            ["供应商商品汇总一致数", summary.product_match_count],
            ["供应商商品汇总差异数", summary.product_diff_count],
            ["", ""],
            ["药师帮总金额", float(summary.ysb_total_amount)],
            ["入库总金额", float(summary.inbound_total_amount)],
            ["总金额差异", float(summary.inbound_total_amount - summary.ysb_total_amount)],
        ]

        for row_idx, row_data in enumerate(data, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.BORDER
                if row_idx == 1:
                    cell.font = Font(bold=True, size=14)
                elif row_idx == 8:
                    cell.fill = self.HEADER_FILL
                    cell.font = self.HEADER_FONT

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 40

    def _create_supplier_diff_sheet(self, results: list[SupplierReconResult]):
        ws = self.workbook.create_sheet(title="差异明细")
        
        headers = ["差异类型", "药师帮供应商", "入库供应商", "金额差异", "备注"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER
        
        row_idx = 2
        for result in results:
            if result.status == "差异":
                row_data = [
                    result.diff_type, result.ysb_supplier, result.inbound_supplier,
                    float(result.amount_diff), result.remark
                ]
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = self.BORDER
                    cell.fill = self.DIFF_FILL
                row_idx += 1
        
        col_widths = [20, 25, 25, 15, 30]
        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    def _create_product_diff_sheet(
        self,
        detail_results: list[DetailMatchResult],
        product_results: list[ProductReconResult]
    ):
        ws = self.workbook.create_sheet(title="差异明细")
        
        ws.cell(row=1, column=1, value="明细匹配差异").font = Font(bold=True, size=12)
        
        headers = ["匹配状态", "药师帮供应商", "商品名称", "条形码", "匹配分数", "匹配备注"]
        for col_idx, header in enumerate(headers, start=2):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER
        
        row_idx = 3
        for result in detail_results:
            if result.match_status != "自动匹配":
                row_data = [
                    result.match_status, result.ysb_supplier, result.product_name,
                    result.barcode, result.match_score, result.match_remark
                ]
                for col_idx, value in enumerate(row_data, start=2):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = self.BORDER
                    if result.match_status == "未匹配":
                        cell.fill = self.DIFF_FILL
                    else:
                        cell.fill = self.SUSPECTED_FILL
                row_idx += 1
        
        row_idx += 1
        ws.cell(row=row_idx, column=1, value="供应商商品汇总差异").font = Font(bold=True, size=12)
        row_idx += 1
        
        headers = ["差异类型", "供应商", "商品编码", "商品名称", "金额差异", "数量差异", "备注"]
        for col_idx, header in enumerate(headers, start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER
        row_idx += 1
        
        for result in product_results:
            if result.status == "差异":
                row_data = [
                    result.diff_type, result.supplier, result.product_code,
                    result.product_name, float(result.amount_diff),
                    float(result.quantity_diff), result.remark
                ]
                for col_idx, value in enumerate(row_data, start=2):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = self.BORDER
                    cell.fill = self.DIFF_FILL
                row_idx += 1
        
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 30
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 15
        ws.column_dimensions["G"].width = 30

    def _create_summary_sheet(
        self,
        summary: ReconciliationSummary,
        account_period: str,
        ysb_file: str,
        inbound_range: tuple[str, str],
    ):
        ws = self.workbook.active
        ws.title = "汇总"

        data = [
            ["对账汇总报告", ""],
            ["", ""],
            ["药师帮账期", account_period],
            ["药师帮文件", ysb_file],
            ["入库查询开始日期", inbound_range[0]],
            ["入库查询结束日期", inbound_range[1]],
            ["", ""],
            ["统计项目", "数值"],
            ["", ""],
            ["药师帮商品行数", summary.ysb_row_count],
            ["入库系统商品行数", summary.inbound_row_count],
            ["", ""],
            ["药师帮供应商数", summary.ysb_supplier_count],
            ["入库供应商数", summary.inbound_supplier_count],
            ["供应商金额一致数", summary.supplier_match_count],
            ["供应商金额差异数", summary.supplier_diff_count],
            ["", ""],
            ["成功匹配商品行数", summary.detail_matched_count],
            ["疑似匹配商品行数", summary.detail_suspected_count],
            ["未匹配商品行数", summary.detail_unmatched_count],
            ["", ""],
            ["供应商商品汇总一致数", summary.product_match_count],
            ["供应商商品汇总差异数", summary.product_diff_count],
            ["", ""],
            ["药师帮总金额", float(summary.ysb_total_amount)],
            ["入库总金额", float(summary.inbound_total_amount)],
            ["总金额差异", float(summary.inbound_total_amount - summary.ysb_total_amount)],
        ]

        for row_idx, row_data in enumerate(data, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.BORDER
                if row_idx == 1:
                    cell.font = Font(bold=True, size=14)
                elif row_idx == 8:
                    cell.fill = self.HEADER_FILL
                    cell.font = self.HEADER_FONT

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 40

    def _create_supplier_sheet(self, results: list[SupplierReconResult]):
        ws = self.workbook.create_sheet(title="供应商汇总对账")
        headers = [
            "对账状态", "差异类型", "药师帮供应商", "入库供应商",
            "药师帮金额", "入库金额", "金额差异",
            "药师帮行数", "入库行数", "匹配方式", "备注"
        ]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER
            cell.alignment = Alignment(horizontal="center")
        
        for row_idx, result in enumerate(results, start=2):
            row_data = [
                result.status, result.diff_type,
                result.ysb_supplier, result.inbound_supplier,
                float(result.ysb_amount), float(result.inbound_amount), float(result.amount_diff),
                result.ysb_count, result.inbound_count,
                result.match_method, result.remark
            ]
            
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.BORDER
                if result.status == "差异":
                    cell.fill = self.DIFF_FILL
                elif result.status == "一致":
                    cell.fill = self.MATCH_FILL
        
        col_widths = [10, 20, 25, 25, 15, 15, 15, 12, 12, 20, 30]
        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    def _create_detail_sheet(self, results: list[DetailMatchResult]):
        ws = self.workbook.create_sheet(title="药师帮明细编码回填")
        headers = [
            "原始行号", "订单号", "药师帮供应商",
            "商品名称", "规格", "厂家", "条形码",
            "药师帮数量", "药师帮金额",
            "匹配入库供应商", "匹配商品编码", "匹配商品名称", "匹配规格", "匹配厂家",
            "匹配方式", "匹配分数", "匹配状态", "匹配备注"
        ]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER
            cell.alignment = Alignment(horizontal="center")
        
        for row_idx, result in enumerate(results, start=2):
            row_data = [
                result.raw_row_index, result.ysb_order_no, result.ysb_supplier,
                result.product_name, result.spec, result.manufacturer, result.barcode,
                float(result.ysb_quantity), float(result.ysb_amount),
                result.matched_supplier_name, result.matched_product_code, result.matched_product_name,
                result.matched_spec, result.matched_manufacturer,
                result.match_method, result.match_score, result.match_status, result.match_remark
            ]
            
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.BORDER
                if result.match_status == "未匹配":
                    cell.fill = self.DIFF_FILL
                elif result.match_status == "疑似匹配":
                    cell.fill = self.SUSPECTED_FILL
                elif result.match_status == "自动匹配":
                    cell.fill = self.MATCH_FILL
        
        col_widths = [10, 20, 20, 30, 15, 20, 18, 12, 12, 20, 15, 30, 15, 20, 20, 10, 12, 30]
        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    def _create_product_sheet(self, results: list[ProductReconResult]):
        ws = self.workbook.create_sheet(title="供应商商品汇总对账")
        headers = [
            "对账状态", "差异类型", "供应商", "商品编码",
            "商品名称", "规格", "厂家",
            "药师帮金额", "入库金额", "金额差异",
            "药师帮数量", "入库数量", "数量差异",
            "药师帮供应商", "入库供应商", "备注"
        ]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER
            cell.alignment = Alignment(horizontal="center")
        
        for row_idx, result in enumerate(results, start=2):
            row_data = [
                result.status, result.diff_type,
                result.supplier, result.product_code,
                result.product_name, result.spec, result.manufacturer,
                float(result.ysb_amount), float(result.inbound_amount), float(result.amount_diff),
                float(result.ysb_quantity), float(result.inbound_quantity), float(result.quantity_diff),
                result.ysb_supplier, result.inbound_supplier, result.remark
            ]
            
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.BORDER
                if result.status == "差异":
                    cell.fill = self.DIFF_FILL
                elif result.status == "一致":
                    cell.fill = self.MATCH_FILL
        
        col_widths = [10, 20, 20, 15, 30, 15, 20, 12, 12, 12, 12, 12, 12, 20, 20, 30]
        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    def _create_diff_sheet(
        self,
        supplier_results: list[SupplierReconResult],
        detail_results: list[DetailMatchResult],
        product_results: list[ProductReconResult]
    ):
        ws = self.workbook.create_sheet(title="差异明细")
        
        ws.cell(row=1, column=1, value="供应商差异").font = Font(bold=True, size=12)
        
        headers = ["差异类型", "药师帮供应商", "入库供应商", "金额差异", "备注"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER
        
        row_idx = 3
        for result in supplier_results:
            if result.status == "差异":
                row_data = [
                    result.diff_type, result.ysb_supplier, result.inbound_supplier,
                    float(result.amount_diff), result.remark
                ]
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = self.BORDER
                    cell.fill = self.DIFF_FILL
                row_idx += 1
        
        row_idx += 1
        ws.cell(row=row_idx, column=1, value="明细匹配差异").font = Font(bold=True, size=12)
        row_idx += 1
        
        headers = ["匹配状态", "药师帮供应商", "商品名称", "条形码", "匹配分数", "匹配备注"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER
        row_idx += 1
        
        for result in detail_results:
            if result.match_status != "自动匹配":
                row_data = [
                    result.match_status, result.ysb_supplier, result.product_name,
                    result.barcode, result.match_score, result.match_remark
                ]
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = self.BORDER
                    if result.match_status == "未匹配":
                        cell.fill = self.DIFF_FILL
                    else:
                        cell.fill = self.SUSPECTED_FILL
                row_idx += 1
        
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 30
        ws.column_dimensions["F"].width = 30
