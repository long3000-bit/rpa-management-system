import logging
import uuid
import json
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple, Optional

import openpyxl
from openpyxl import load_workbook

from app.storage.database import Database


class YysStockImportService:
    
    def __init__(self, db: Database):
        self.db = db
    
    def read_excel_preview(self, file_path: str, sheet_name: str = None, max_rows: int = 100) -> Tuple[List[str], List[List[str]], str]:
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            
            sheets = wb.sheetnames
            if not sheets:
                return [], [], "Excel文件中没有工作表"
            
            if sheet_name and sheet_name in sheets:
                ws = wb[sheet_name]
            else:
                ws = wb[sheets[0]]
                sheet_name = sheets[0]
            
            headers = []
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                header = str(cell.value).strip() if cell.value else f"列{col}"
                headers.append(header)
            
            data_rows = []
            for row in range(2, min(ws.max_row + 1, max_rows + 2)):
                row_data = []
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    value = str(cell.value).strip() if cell.value else ""
                    row_data.append(value)
                data_rows.append(row_data)
            
            wb.close()
            
            return headers, data_rows, ""
            
        except Exception as e:
            logging.error(f"读取Excel预览失败: {e}")
            return [], [], str(e)
    
    def get_sheets(self, file_path: str) -> Tuple[List[str], str]:
        try:
            wb = load_workbook(file_path, read_only=True)
            sheets = wb.sheetnames
            wb.close()
            return sheets, ""
        except Exception as e:
            logging.error(f"获取工作表列表失败: {e}")
            return [], str(e)
    
    def import_excel_full(self, file_path: str, sheet_name: str, imported_by: str = "") -> Tuple[str, int, int, str]:
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            
            if sheet_name not in wb.sheetnames:
                wb.close()
                return "", 0, 0, f"工作表 '{sheet_name}' 不存在"
            
            ws = wb[sheet_name]
            
            # Use iter_rows for efficient reading in read_only mode
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [str(cell).strip() if cell else "" for cell in first_row]
            
            batch_id = f"YYS{datetime.now().strftime('%Y%m%d%H%M%S')}{uuid.uuid4().hex[:4]}"
            now = datetime.now().isoformat()
            
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('''
                INSERT INTO yys_import_batch
                (batch_id, batch_name, source_file, sheet_name, total_count, valid_count, invalid_count, status, imported_by, imported_at, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                batch_id,
                Path(file_path).name,
                file_path,
                sheet_name,
                0, 0, 0,
                'importing',
                imported_by,
                now,
                now
            ))
            
            valid_count = 0
            invalid_count = 0
            
            # Use iter_rows for efficient row iteration
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                row_dict = {}
                for col_idx, header in enumerate(headers):
                    if header and col_idx < len(row):
                        value = str(row[col_idx]).strip() if row[col_idx] else ""
                        row_dict[header] = value
                
                raw_data_json = json.dumps(row_dict, ensure_ascii=False)
                
                productno = row_dict.get('商品编码', '')
                oldproductno = row_dict.get('旧商品编码', '')
                lotno = row_dict.get('批号', '')
                
                if not productno and not oldproductno:
                    invalid_count += 1
                    continue
                
                if not lotno:
                    invalid_count += 1
                    continue
                
                yys_quantity_str = row_dict.get('商品数量', '0')
                try:
                    yys_quantity = float(yys_quantity_str) if yys_quantity_str else 0.0
                except ValueError:
                    yys_quantity = 0.0
                
                warehouse = row_dict.get('仓库名称', '')
                productname = row_dict.get('商品名称', '')
                specification = row_dict.get('商品规格', '')
                unit = row_dict.get('单位', '')
                manufacturer = row_dict.get('生产企业', '')
                supplier = row_dict.get('供应商', '')
                valid_date = row_dict.get('有效期', '')
                production_date = row_dict.get('生产日期', '')
                barcode = row_dict.get('条形码', '')
                approval_number = row_dict.get('批准文号', '')
                chinese_medicine_flag = row_dict.get('中药标志(0否，1是)', '')
                inbound_time = row_dict.get('入库时间', '')
                stock_status = row_dict.get('库存状态(1合格,0不合格)', '')
                tax_rate = row_dict.get('税率', '')
                gross_profit_rate = row_dict.get('毛利率', '')
                
                try:
                    retail_price = float(row_dict.get('零售价', '0') or '0')
                except ValueError:
                    retail_price = 0.0
                
                try:
                    batch_price = float(row_dict.get('批次单价', '0') or '0')
                except ValueError:
                    batch_price = 0.0
                
                try:
                    amount = float(row_dict.get('在库金额', '0') or '0')
                except ValueError:
                    amount = 0.0
                
                try:
                    gross_profit = float(row_dict.get('毛利', '0') or '0')
                except ValueError:
                    gross_profit = 0.0
                
                detail_id = uuid.uuid4().hex
                
                cursor.execute('''
                    INSERT INTO yys_import_detail
                    (detail_id, batch_id, row_number, productno, oldproductno, productname, lotno, yys_quantity,
                     warehouse, specification, unit, manufacturer, supplier, valid_date, production_date,
                     retail_price, batch_price, amount, tax_rate, gross_profit, gross_profit_rate,
                     stock_status, barcode, approval_number, chinese_medicine_flag, inbound_time,
                     raw_data, import_status, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    detail_id,
                    batch_id,
                    row_idx,
                    productno,
                    oldproductno,
                    productname,
                    lotno,
                    yys_quantity,
                    warehouse,
                    specification,
                    unit,
                    manufacturer,
                    supplier,
                    valid_date,
                    production_date,
                    retail_price,
                    batch_price,
                    amount,
                    tax_rate,
                    gross_profit,
                    gross_profit_rate,
                    stock_status,
                    barcode,
                    approval_number,
                    chinese_medicine_flag,
                    inbound_time,
                    raw_data_json,
                    'valid',
                    now
                ))
                
                valid_count += 1
            
            wb.close()
            
            cursor.execute('''
                UPDATE yys_import_batch
                SET total_count = ?, valid_count = ?, invalid_count = ?, status = ?
                WHERE batch_id = ?
            ''', (valid_count + invalid_count, valid_count, invalid_count, 'completed', batch_id))
            
            conn.commit()
            
            logging.info(f"云药店库存导入成功，批次号: {batch_id}, 有效记录: {valid_count}, 无效记录: {invalid_count}")
            
            return batch_id, valid_count, invalid_count, ""
            
        except Exception as e:
            logging.error(f"导入云药店库存失败: {e}")
            return "", 0, 0, str(e)
    
    def import_excel(self, file_path: str, sheet_name: str, field_mapping: Dict[str, str], imported_by: str = "") -> Tuple[str, int, int, str]:
        return self.import_excel_full(file_path, sheet_name, imported_by)
    
    def get_batches(self) -> List[Dict]:
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT batch_id, batch_name, source_file, sheet_name, total_count, valid_count, invalid_count, status, imported_by, imported_at
                FROM yys_import_batch
                ORDER BY created_at DESC
            ''')
            
            rows = cursor.fetchall()
            
            batches = []
            for row in rows:
                batches.append(dict(row))
            
            return batches
            
        except Exception as e:
            logging.error(f"获取导入批次列表失败: {e}")
            return []
    
    def get_batch_details(self, batch_id: str) -> List[Dict]:
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT detail_id, row_number, productno, oldproductno, productname, lotno, yys_quantity,
                       warehouse, specification, unit, manufacturer, supplier, valid_date,
                       production_date, retail_price, batch_price, amount, tax_rate,
                       gross_profit, gross_profit_rate, stock_status, barcode,
                       approval_number, chinese_medicine_flag, inbound_time, import_status, raw_data
                FROM yys_import_detail
                WHERE batch_id = ?
                ORDER BY row_number
            ''', (batch_id,))
            
            rows = cursor.fetchall()
            
            details = []
            for row in rows:
                detail = dict(row)
                if detail.get('raw_data'):
                    try:
                        detail['all_fields'] = json.loads(detail['raw_data'])
                    except:
                        detail['all_fields'] = {}
                details.append(detail)
            
            return details
            
        except Exception as e:
            logging.error(f"获取导入明细失败: {e}")
            return []
    
    def get_batch_details_with_filter(self, batch_id: str, productno: str = None, oldproductno: str = None, productname: str = None, lotno: str = None) -> List[Dict]:
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            sql = '''
                SELECT detail_id, row_number, productno, oldproductno, productname, lotno, yys_quantity,
                       warehouse, specification, unit, manufacturer, supplier, valid_date,
                       production_date, retail_price, batch_price, amount, tax_rate,
                       gross_profit, gross_profit_rate, stock_status, barcode,
                       approval_number, chinese_medicine_flag, inbound_time, import_status, raw_data
                FROM yys_import_detail
                WHERE batch_id = ?
            '''
            params = [batch_id]
            
            if productno:
                sql += " AND productno LIKE ?"
                params.append(f"%{productno}%")
            
            if oldproductno:
                sql += " AND oldproductno LIKE ?"
                params.append(f"%{oldproductno}%")
            
            if productname:
                sql += " AND productname LIKE ?"
                params.append(f"%{productname}%")
            
            if lotno:
                sql += " AND lotno LIKE ?"
                params.append(f"%{lotno}%")
            
            sql += " ORDER BY row_number"
            
            cursor.execute(sql, params)
            
            rows = cursor.fetchall()
            
            details = []
            for row in rows:
                detail = dict(row)
                if detail.get('raw_data'):
                    try:
                        detail['all_fields'] = json.loads(detail['raw_data'])
                    except:
                        detail['all_fields'] = {}
                details.append(detail)
            
            return details
            
        except Exception as e:
            logging.error(f"获取导入明细失败: {e}")
            return []
    
    def delete_batch(self, batch_id: str) -> Tuple[bool, str]:
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('DELETE FROM yys_import_detail WHERE batch_id = ?', (batch_id,))
            cursor.execute('DELETE FROM yys_import_batch WHERE batch_id = ?', (batch_id,))
            
            conn.commit()
            
            return True, ""
            
        except Exception as e:
            logging.error(f"删除导入批次失败: {e}")
            return False, str(e)