"""
医保价格管控 - Excel导入服务（新版）

支持导入以下数据源：
1. 医保目录-西药（直接使用Excel字段名）
2. 医保目录-中成药（直接使用Excel字段名）
3. 医保价格上限（直接使用Excel字段名）
4. 云药店商品目录（直接使用Excel字段名）
"""

import logging
import json
import uuid
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, List, Any
from dataclasses import dataclass, field

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from app.storage.database import Database


@dataclass
class MedicalImportResult:
    """导入结果"""
    batch_id: str
    batch_type: str
    file_name: str
    sheet_name: str = ""
    total_rows: int = 0
    success_rows: int = 0
    failed_rows: int = 0
    import_status: str = "pending"
    error_message: str = ""
    failures: List[Dict] = field(default_factory=list)


class MedicalPriceImportService:
    """医保价格管控导入服务 - 直接使用Excel字段名"""
    
    # 数据源类型
    BATCH_TYPES = {
        "medical_catalog_western": "医保目录-西药",
        "medical_catalog_chinese": "医保目录-中成药",
        "medical_price_limit": "三同口径文件",
        "cloud_pharmacy_catalog": "云药店商品目录",
    }
    
    # 表名映射
    TABLE_NAMES = {
        "medical_catalog_western": "medical_catalog_western",
        "medical_catalog_chinese": "medical_catalog_chinese",
        "medical_price_limit": "medical_price_limit",
        "cloud_pharmacy_catalog": "cloud_pharmacy_catalog",
    }
    
    # 字段名映射（Excel列名 -> 数据库列名）
    FIELD_MAPPING = {
        "cloud_pharmacy_catalog": {
            "商品规格": "规格",
            "生产企业": "生产厂家",
            "项目类别": "",  # 忽略此字段
            "是否医保": "",  # 忽略此字段
            "零售价": "",  # 忽略此字段
            "处方药标志": "",  # 忽略此字段
            "中药标志": "",  # 忽略此字段
            "省目录编码": "",  # 忽略此字段
            "条形码": "",  # 忽略此字段
            "限制使用标志": "",  # 忽略此字段
            "限制使用范围": "",  # 忽略此字段
            "院内收费类别": "",  # 忽略此字段
            "一次性使用标志": "",  # 忽略此字段
            "材料类别": "",  # 忽略此字段
            "特殊药品": "",  # 忽略此字段
            "特管药品": "",  # 忽略此字段
            "商品性质": "",  # 忽略此字段
            "是否冷藏": "",  # 忽略此字段
            "是否麻黄碱": "",  # 忽略此字段
            "经营分类": "",  # 忽略此字段
            "不参与促销标志": "",  # 忽略此字段
            "质量状态": "",  # 忽略此字段
            "操作码": "",  # 忽略此字段
            "云商品编码": "",  # 忽略此字段
            "会员价": "",  # 忽略此字段
            "配送价格": "",  # 忽略此字段
            "是否拆零": "",  # 忽略此字段
            "拆零价格": "",  # 忽略此字段
            "拆零数量": "",  # 忽略此字段
            "拆零单位": "",  # 忽略此字段
            "期初标志": "",  # 忽略此字段
        },
        "medical_price_limit": {
            "药品名称": "药品名称",
            "生产企业": "生产企业",
            "剂型": "剂型",
            "规格": "规格",
        },
    }
    
    def __init__(self, db: Database):
        self.db = db
    
    def generate_batch_id(self) -> str:
        """生成批次ID"""
        return f"MED_{datetime.now().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex[:8]}"
    
    def _save_batch_record(self, result: MedicalImportResult, file_path: str, imported_by: str):
        """保存批次记录"""
        conn = self.db.get_connection()
        cursor = conn.cursor()
        now = datetime.now().isoformat()
        
        cursor.execute('''
            INSERT INTO medical_import_batches (
                batch_id, batch_type, file_name, file_path, sheet_name,
                total_rows, success_rows, failed_rows, import_status,
                imported_by, imported_at, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            result.batch_id,
            result.batch_type,
            result.file_name,
            file_path,
            result.sheet_name,
            result.total_rows,
            result.success_rows,
            result.failed_rows,
            result.import_status,
            imported_by,
            now,
            now
        ))
        
        # 保存失败记录
        if result.failures:
            for failure in result.failures:
                cursor.execute('''
                    INSERT INTO medical_import_failures (
                        batch_id, row_index, raw_data, failure_reason, created_at
                    ) VALUES (?, ?, ?, ?, ?)
                ''', (
                    result.batch_id,
                    failure.get('row_index', 0),
                    failure.get('raw_data', ''),
                    failure.get('failure_reason', ''),
                    now
                ))
        
        conn.commit()
    
    def _get_table_columns(self, table_name: str) -> List[str]:
        """获取表的所有列名（除了系统字段）"""
        conn = self.db.get_connection()
        cursor = conn.cursor()
        
        # 获取表结构
        cursor.execute(f"PRAGMA table_info({table_name})")
        columns_info = cursor.fetchall()
        
        # 提取列名，排除系统字段
        system_fields = ["id", "batch_id", "sheet_type", "row_index", "原始数据", "created_at"]
        columns = []
        
        for col in columns_info:
            col_name = col['name']
            if col_name not in system_fields:
                columns.append(col_name)
        
        return columns
    
    def _find_header_row(self, worksheet: Worksheet, key_fields: List[str]) -> int:
        """自动识别表头行"""
        for row_num in range(1, min(11, worksheet.max_row + 1)):
            headers = []
            for cell in worksheet[row_num]:
                headers.append(str(cell.value or "").strip())
            
            matched_count = sum(1 for field in key_fields if field in headers)
            if matched_count >= 3:
                logging.info(f"表头在第 {row_num} 行，包含 {matched_count} 个关键字段")
                return row_num
        
        return 1  # 默认第1行
    
    def _import_data(
        self,
        file_path: str,
        table_name: str,
        batch_type: str,
        key_fields: List[str],
        sheet_name: str = None,
        imported_by: str = "admin"
    ) -> MedicalImportResult:
        """通用导入方法 - 直接使用Excel字段名"""
        batch_id = self.generate_batch_id()
        result = MedicalImportResult(
            batch_id=batch_id,
            batch_type=batch_type,
            file_name=Path(file_path).name,
            sheet_name=sheet_name or ""
        )
        
        try:
            # 不使用 read_only 模式，避免某些特殊格式文件读取问题
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            # 获取工作表
            if sheet_name:
                ws = workbook[sheet_name]
            else:
                ws = workbook.active
                result.sheet_name = ws.title
            
            # 自动识别表头位置
            header_row = self._find_header_row(ws, key_fields)
            
            # 读取表头
            headers = []
            for cell in ws[header_row]:
                headers.append(str(cell.value or "").strip())
            
            # 获取表的所有列名
            table_columns = self._get_table_columns(table_name)
            
            logging.info(f"表列名: {table_columns}")
            logging.info(f"Excel表头: {headers[:10]}")
            
            # 导入数据
            conn = self.db.get_connection()
            cursor = conn.cursor()
            now = datetime.now().isoformat()
            
            # 检查表是否有 sheet_type 字段
            has_sheet_type = table_name in ['medical_catalog_western', 'medical_catalog_chinese']
            
            # 获取字段映射
            field_mapping = self.FIELD_MAPPING.get(table_name, {})
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1), start=header_row + 1):
                try:
                    row_data = {}
                    for col_idx, cell in enumerate(row):
                        if col_idx < len(headers):
                            excel_col_name = headers[col_idx]
                            value = str(cell.value or "").strip()
                            
                            # 应用字段映射
                            if excel_col_name in field_mapping:
                                db_col_name = field_mapping[excel_col_name]
                                if db_col_name:  # 如果映射不为空，则存储
                                    row_data[db_col_name] = value
                                # 否则忽略此字段
                            else:
                                # 没有映射的字段，直接使用Excel列名
                                row_data[excel_col_name] = value
                    
                    # 转换为JSON存储原始数据
                    raw_data_json = json.dumps(row_data, ensure_ascii=False)
                    
                    # 构建插入数据
                    if has_sheet_type:
                        values = [batch_id, "main", row_idx]
                    else:
                        values = [batch_id, row_idx]
                    
                    # 对于每个表列，从Excel数据中获取值（如果不存在则为空值）
                    for col in table_columns:
                        values.append(row_data.get(col, ""))
                    
                    values.extend([raw_data_json, now])
                    
                    # 构建SQL - 将列名用引号括起来，避免特殊字符导致的语法错误
                    quoted_columns = ", ".join([f'"{col}"' for col in table_columns])
                    if has_sheet_type:
                        columns_str = f'batch_id, sheet_type, row_index, {quoted_columns}, "原始数据", created_at'
                    else:
                        columns_str = f'batch_id, row_index, {quoted_columns}, "原始数据", created_at'
                    placeholders = ", ".join(["?" for _ in values])
                    
                    cursor.execute(f'''
                        INSERT INTO {table_name} ({columns_str})
                        VALUES ({placeholders})
                    ''', values)
                    
                    result.success_rows += 1
                    
                except Exception as e:
                    result.failed_rows += 1
                    result.failures.append({
                        "row_index": row_idx,
                        "raw_data": json.dumps(row_data, ensure_ascii=False),
                        "failure_reason": str(e)
                    })
                    logging.warning(f"导入行 {row_idx} 失败: {e}")
            
            result.total_rows = result.success_rows + result.failed_rows
            result.import_status = "success" if result.failed_rows == 0 else "partial"
            
            conn.commit()
            workbook.close()
            
            self._save_batch_record(result, file_path, imported_by)
            
            logging.info(f"导入完成: {result.success_rows}/{result.total_rows} 行")
            
        except Exception as e:
            result.import_status = "failed"
            result.error_message = str(e)
            logging.error(f"导入失败: {e}")
            import traceback
            logging.error(traceback.format_exc())
            # 即使失败也保存批次记录
            try:
                workbook.close()
            except:
                pass
            self._save_batch_record(result, file_path, imported_by)
        
        return result
    
    def import_medical_catalog_western(
        self,
        file_path: str,
        sheet_name: str = None,
        imported_by: str = "admin"
    ) -> MedicalImportResult:
        """导入医保西药目录"""
        key_fields = ["药品通用名编码", "医保药品名称", "注册名称", "医保剂型", "实际规格"]
        return self._import_data(
            file_path,
            "medical_catalog_western",
            "medical_catalog_western",
            key_fields,
            sheet_name,
            imported_by
        )
    
    def import_medical_catalog_chinese(
        self,
        file_path: str,
        sheet_name: str = None,
        imported_by: str = "admin"
    ) -> MedicalImportResult:
        """导入医保中成药目录"""
        key_fields = ["药品通用名编码", "医保药品名称", "注册名称", "医保剂型", "实际规格"]
        return self._import_data(
            file_path,
            "medical_catalog_chinese",
            "medical_catalog_chinese",
            key_fields,
            sheet_name,
            imported_by
        )
    
    def import_medical_price_limit(
        self,
        file_path: str,
        sheet_name: str = None,
        imported_by: str = "admin"
    ) -> MedicalImportResult:
        """导入三同口径文件"""
        key_fields = ["医保编码", "药品名称", "生产企业", "剂型", "规格", "三同药品挂网最低单片价"]
        return self._import_data(
            file_path,
            "medical_price_limit",
            "medical_price_limit",
            key_fields,
            sheet_name,
            imported_by
        )
    
    def import_cloud_pharmacy_catalog(
        self,
        file_path: str,
        sheet_name: str = None,
        imported_by: str = "admin"
    ) -> MedicalImportResult:
        """导入云药店商品目录"""
        key_fields = ["商品编码", "商品名称", "生产厂家", "规格"]
        return self._import_data(
            file_path,
            "cloud_pharmacy_catalog",
            "cloud_pharmacy_catalog",
            key_fields,
            sheet_name,
            imported_by
        )
    
    def delete_batch(self, batch_id: str) -> bool:
        """删除批次及其数据"""
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            # 获取批次信息
            cursor.execute("SELECT * FROM medical_import_batches WHERE batch_id = ?", [batch_id])
            batch = cursor.fetchone()
            
            if not batch:
                return False
            
            batch_type = batch['batch_type']
            table_name = self.TABLE_NAMES.get(batch_type)
            
            if table_name:
                # 删除数据
                cursor.execute(f"DELETE FROM {table_name} WHERE batch_id = ?", [batch_id])
                logging.info(f"删除表 {table_name} 中的批次 {batch_id} 数据")
            
            # 删除批次记录
            cursor.execute("DELETE FROM medical_import_batches WHERE batch_id = ?", [batch_id])
            
            conn.commit()
            logging.info(f"批次 {batch_id} 已删除")
            
            return True
            
        except Exception as e:
            logging.error(f"删除批次失败: {e}")
            return False
    
    def get_import_batches(self, limit: int = 20, batch_type: str = None) -> List[Dict]:
        """获取导入批次列表"""
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            if batch_type:
                cursor.execute('''
                    SELECT * FROM medical_import_batches 
                    WHERE batch_type = ?
                    ORDER BY created_at DESC 
                    LIMIT ?
                ''', [batch_type, limit])
            else:
                cursor.execute('''
                    SELECT * FROM medical_import_batches 
                    ORDER BY created_at DESC 
                    LIMIT ?
                ''', [limit])
            
            batches = cursor.fetchall()
            return [dict(batch) for batch in batches]
            
        except Exception as e:
            logging.error(f"获取导入批次失败: {e}")
            return []
    
    def get_available_batches_for_compare(self) -> Dict[str, List[Dict]]:
        """获取可用于比对的批次列表"""
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            result = {
                'medical_catalog_western': [],
                'medical_catalog_chinese': [],
                'medical_price_limit': [],
                'cloud_pharmacy_catalog': [],
                'junyuan_sales_price': []
            }
            
            # 获取西药目录批次
            cursor.execute('''
                SELECT batch_id, file_name, total_rows, success_rows, import_status, created_at
                FROM medical_import_batches
                WHERE batch_type = 'medical_catalog_western' AND import_status = 'success'
                ORDER BY created_at DESC
                LIMIT 10
            ''')
            result['medical_catalog_western'] = [dict(row) for row in cursor.fetchall()]
            
            # 获取中成药目录批次
            cursor.execute('''
                SELECT batch_id, file_name, total_rows, success_rows, import_status, created_at
                FROM medical_import_batches
                WHERE batch_type = 'medical_catalog_chinese' AND import_status = 'success'
                ORDER BY created_at DESC
                LIMIT 10
            ''')
            result['medical_catalog_chinese'] = [dict(row) for row in cursor.fetchall()]
            
            # 获取价格上限批次
            cursor.execute('''
                SELECT batch_id, file_name, total_rows, success_rows, import_status, created_at
                FROM medical_import_batches
                WHERE batch_type = 'medical_price_limit' AND import_status = 'success'
                ORDER BY created_at DESC
                LIMIT 10
            ''')
            result['medical_price_limit'] = [dict(row) for row in cursor.fetchall()]
            
            # 获取云药店商品目录批次
            cursor.execute('''
                SELECT batch_id, file_name, total_rows, success_rows, import_status, created_at
                FROM medical_import_batches
                WHERE batch_type = 'cloud_pharmacy_catalog' AND import_status = 'success'
                ORDER BY created_at DESC
                LIMIT 10
            ''')
            result['cloud_pharmacy_catalog'] = [dict(row) for row in cursor.fetchall()]
            
            # 获取君元销售价格批次
            cursor.execute('''
                SELECT batch_id, file_name, total_rows, success_rows, import_status, created_at
                FROM medical_import_batches
                WHERE batch_type = 'junyuan_sales_price' AND import_status = 'success'
                ORDER BY created_at DESC
                LIMIT 10
            ''')
            result['junyuan_sales_price'] = [dict(row) for row in cursor.fetchall()]
            
            return result
            
        except Exception as e:
            logging.error(f"获取可用批次失败: {e}")
            return {
                'medical_catalog_western': [],
                'medical_catalog_chinese': [],
                'medical_price_limit': [],
                'cloud_pharmacy_catalog': [],
                'junyuan_sales_price': []
            }