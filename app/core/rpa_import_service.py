import logging
import json
import uuid
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Dict, Tuple
import openpyxl

from app.storage.database import Database


class RpaImportService:
    
    def __init__(self, db: Database):
        self.db = db
    
    def read_excel(self, file_path: str, sheet_name: str, max_rows: int = 0) -> Tuple[List[str], List[List[str]], str]:
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            
            if sheet_name not in wb.sheetnames:
                wb.close()
                return [], [], f"工作表 '{sheet_name}' 不存在"
            
            ws = wb[sheet_name]
            
            headers = []
            data_rows = []
            
            row_count = 0
            for row in ws.iter_rows(values_only=True):
                row_count += 1
                
                if row_count == 1:
                    headers = [str(cell) if cell else "" for cell in row]
                    continue
                
                if max_rows > 0 and row_count > max_rows + 1:
                    break
                
                row_data = [str(cell) if cell else "" for cell in row]
                if any(row_data):
                    data_rows.append(row_data)
            
            wb.close()
            
            logging.info(f"读取Excel完成: {file_path}, Sheet: {sheet_name}, 表头: {len(headers)}列, 数据: {len(data_rows)}行")
            
            return headers, data_rows, ""
            
        except Exception as e:
            logging.error(f"读取Excel失败: {e}")
            return [], [], str(e)
    
    def get_sheets(self, file_path: str) -> Tuple[List[str], str]:
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheets = wb.sheetnames
            wb.close()
            return sheets, ""
        except Exception as e:
            logging.error(f"获取Sheet列表失败: {e}")
            return [], str(e)
    
    def validate_data(self, headers: List[str], data_rows: List[List[str]], 
                      field_mapping: Dict, required_fields: List[str]) -> Tuple[List[Dict], List[Dict], str]:
        valid_rows = []
        invalid_rows = []
        
        excel_to_system = {}
        for system_field, mapping in field_mapping.items():
            excel_col = mapping.get('excel_column', '')
            if excel_col and excel_col in headers:
                excel_to_system[excel_col] = system_field
        
        missing_required = []
        for field in required_fields:
            found = False
            for excel_col, system_field in excel_to_system.items():
                if system_field == field:
                    found = True
                    break
            if not found:
                missing_required.append(field)
        
        if missing_required:
            return [], [], f"缺少必填字段映射: {', '.join(missing_required)}"
        
        for row_idx, row_data in enumerate(data_rows):
            row_dict = {}
            for col_idx, cell_value in enumerate(row_data):
                if col_idx < len(headers):
                    excel_col = headers[col_idx]
                    if excel_col in excel_to_system:
                        system_field = excel_to_system[excel_col]
                        row_dict[system_field] = cell_value
                    else:
                        row_dict[excel_col] = cell_value
            
            validation_errors = []
            for field in required_fields:
                if field not in row_dict or not row_dict[field]:
                    validation_errors.append(f"必填字段 '{field}' 为空")
            
            row_dict['_excel_row_number'] = row_idx + 2
            row_dict['_raw_data'] = json.dumps(row_data, ensure_ascii=False)
            
            if validation_errors:
                row_dict['_validation_errors'] = validation_errors
                invalid_rows.append(row_dict)
            else:
                valid_rows.append(row_dict)
        
        logging.info(f"数据校验完成: 有效 {len(valid_rows)} 行, 无效 {len(invalid_rows)} 行")
        
        return valid_rows, invalid_rows, ""
    
    def import_to_database(self, valid_rows: List[Dict], template_id: str,
                           source_file: str, sheet_name: str,
                           imported_by: str = "system") -> Tuple[str, str]:
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            import_batch_id = f"RPA{datetime.now().strftime('%Y%m%d%H%M%S')}{uuid.uuid4().hex[:4]}"
            import_name = Path(source_file).stem
            now = datetime.now().isoformat()
            
            cursor.execute('''
                INSERT INTO rpa_import_batches
                (import_batch_id, import_name, template_id, source_file, sheet_name,
                 total_count, valid_count, invalid_count, duplicate_count, status,
                 imported_by, imported_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                import_batch_id,
                import_name,
                template_id,
                source_file,
                sheet_name,
                len(valid_rows),
                len(valid_rows),
                0,
                0,
                'ready',
                imported_by,
                now
            ))
            
            business_key_field = 'business_key'
            
            for row in valid_rows:
                import_row_id = f"{import_batch_id}_{row['_excel_row_number']}"
                business_key = row.get(business_key_field, '') or f"ROW_{row['_excel_row_number']}"
                
                normalized_data = {}
                for key, value in row.items():
                    if not key.startswith('_'):
                        normalized_data[key] = value
                
                cursor.execute('''
                    INSERT INTO rpa_import_details
                    (import_row_id, import_batch_id, excel_row_number, business_key,
                     raw_data, normalized_data, data_status, business_status,
                     rpa_status, can_retry, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    import_row_id,
                    import_batch_id,
                    row['_excel_row_number'],
                    business_key,
                    row.get('_raw_data', ''),
                    json.dumps(normalized_data, ensure_ascii=False),
                    'valid',
                    'pending',
                    'pending',
                    1,
                    now,
                    now
                ))
            
            conn.commit()
            
            logging.info(f"导入完成: 批次 {import_batch_id}, 共 {len(valid_rows)} 行数据")
            
            return import_batch_id, ""
            
        except Exception as e:
            conn.rollback()
            logging.error(f"导入数据库失败: {e}")
            return "", str(e)
    
    def get_batch_data(self, import_batch_id: str, 
                       status_filter: str = "all",
                       keyword: str = "") -> Tuple[List[Dict], str]:
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            sql = '''
                SELECT import_row_id, import_batch_id, excel_row_number, business_key,
                       raw_data, normalized_data, data_status, business_status,
                       target_system_no, target_system_message, processed_at,
                       last_task_id, last_executed_by, last_executed_at, execute_count,
                       can_retry, validation_message, rpa_status, rpa_error_message,
                       rpa_system_no, rpa_screenshot_path, created_at, updated_at
                FROM rpa_import_details
                WHERE import_batch_id = ?
            '''
            
            params = [import_batch_id]
            
            if status_filter != "all":
                sql += " AND rpa_status = ?"
                params.append(status_filter)
            
            if keyword:
                sql += " AND (business_key LIKE ? OR normalized_data LIKE ?)"
                params.extend([f"%{keyword}%", f"%{keyword}%"])
            
            sql += " ORDER BY excel_row_number"
            
            cursor.execute(sql, params)
            rows = cursor.fetchall()
            
            result = []
            for row in rows:
                result.append(dict(row))
            
            return result, ""
            
        except Exception as e:
            logging.error(f"获取批次数据失败: {e}")
            return [], str(e)
    
    def export_result_excel(self, import_batch_id: str, output_path: str) -> Tuple[str, str]:
        try:
            data, error = self.get_batch_data(import_batch_id)
            if error:
                return "", error
            
            if not data:
                return "", "没有数据可导出"
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "RPA执行结果"
            
            headers = [
                "行号", "业务主键", "数据状态", "业务状态", "RPA状态",
                "系统单号", "系统返回信息", "失败原因", "执行次数",
                "执行时间", "截图路径", "是否可重试"
            ]
            
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)
            
            for row_idx, row in enumerate(data, 2):
                ws.cell(row=row_idx, column=1, value=row['excel_row_number'])
                ws.cell(row=row_idx, column=2, value=row['business_key'])
                ws.cell(row=row_idx, column=3, value=row['data_status'])
                ws.cell(row=row_idx, column=4, value=row['business_status'])
                ws.cell(row=row_idx, column=5, value=row['rpa_status'])
                ws.cell(row=row_idx, column=6, value=row['target_system_no'] or row['rpa_system_no'])
                ws.cell(row=row_idx, column=7, value=row['target_system_message'])
                ws.cell(row=row_idx, column=8, value=row['rpa_error_message'] or row['validation_message'])
                ws.cell(row=row_idx, column=9, value=row['execute_count'])
                ws.cell(row=row_idx, column=10, value=row['last_executed_at'])
                ws.cell(row=row_idx, column=11, value=row['rpa_screenshot_path'])
                ws.cell(row=row_idx, column=12, value="是" if row['can_retry'] else "否")
            
            normalized_data = json.loads(data[0]['normalized_data']) if data[0]['normalized_data'] else {}
            extra_headers = list(normalized_data.keys())
            
            if extra_headers:
                start_col = len(headers) + 1
                for col_idx, header in enumerate(extra_headers, start_col):
                    ws.cell(row=1, column=col_idx, value=header)
                
                for row_idx, row in enumerate(data, 2):
                    try:
                        row_data = json.loads(row['normalized_data']) if row['normalized_data'] else {}
                        for col_idx, header in enumerate(extra_headers, start_col):
                            ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
                    except:
                        pass
            
            wb.save(output_path)
            wb.close()
            
            logging.info(f"导出Excel完成: {output_path}")
            
            return output_path, ""
            
        except Exception as e:
            logging.error(f"导出Excel失败: {e}")
            return "", str(e)
    
    def get_batch_summary(self, import_batch_id: str) -> Tuple[Dict, str]:
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT import_batch_id, import_name, template_id, source_file, sheet_name,
                       total_count, valid_count, invalid_count, duplicate_count, status,
                       imported_by, imported_at, error_message
                FROM rpa_import_batches
                WHERE import_batch_id = ?
            ''', (import_batch_id,))
            
            row = cursor.fetchone()
            if not row:
                return {}, "批次不存在"
            
            batch_info = dict(row)
            
            cursor.execute('''
                SELECT rpa_status, COUNT(*) as count
                FROM rpa_import_details
                WHERE import_batch_id = ?
                GROUP BY rpa_status
            ''', (import_batch_id,))
            
            status_counts = {}
            for r in cursor.fetchall():
                status_counts[r['rpa_status']] = r['count']
            
            batch_info['status_counts'] = status_counts
            
            return batch_info, ""
            
        except Exception as e:
            logging.error(f"获取批次摘要失败: {e}")
            return {}, str(e)
    
    def update_row_status(self, import_row_id: str, rpa_status: str,
                          rpa_system_no: str = "", rpa_error_message: str = "",
                          rpa_screenshot_path: str = "") -> Tuple[bool, str]:
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            now = datetime.now().isoformat()
            
            cursor.execute('''
                UPDATE rpa_import_details
                SET rpa_status = ?, rpa_system_no = ?, rpa_error_message = ?,
                    rpa_screenshot_path = ?, last_executed_at = ?, 
                    execute_count = execute_count + 1, updated_at = ?
                WHERE import_row_id = ?
            ''', (
                rpa_status,
                rpa_system_no,
                rpa_error_message,
                rpa_screenshot_path,
                now,
                now,
                import_row_id
            ))
            
            conn.commit()
            
            return True, ""
            
        except Exception as e:
            conn.rollback()
            logging.error(f"更新行状态失败: {e}")
            return False, str(e)