import logging
import json
import os
from datetime import datetime, timedelta
from typing import Dict, List

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QPushButton, QDialog, QTextEdit, QLabel, QGroupBox, QComboBox,
    QDateTimeEdit, QMessageBox, QHeaderView, QSplitter, QFrame, QFileDialog
)
from PySide6.QtCore import Qt, QDateTime

from app.storage.database import Database
from app.core.permission_service import PermissionService
from app.core.permission_checker import PermissionChecker, PermissionCodes


class OperationLogPage(QWidget):
    """操作日志页面"""
    
    def __init__(self, db: Database, username: str = None):
        super().__init__()
        self.db = db
        self.username = username or 'admin'
        self.permission_service = PermissionService(db)
        self.permission_checker = PermissionChecker(db, self.username)
        
        self.current_page = 0
        self.page_size = 50
        
        self.init_ui()
        self.load_filters()
        self.load_logs()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # 筛选区域
        filter_group = QGroupBox("筛选条件")
        filter_layout = QHBoxLayout(filter_group)
        
        # 用户筛选
        self.user_combo = QComboBox()
        self.user_combo.addItem("全部用户", None)
        filter_layout.addWidget(QLabel("用户:"))
        filter_layout.addWidget(self.user_combo)
        
        # 操作类型筛选
        self.type_combo = QComboBox()
        self.type_combo.addItem("全部类型", None)
        filter_layout.addWidget(QLabel("操作类型:"))
        filter_layout.addWidget(self.type_combo)
        
        # 时间范围筛选
        filter_layout.addWidget(QLabel("开始时间:"))
        self.start_time_edit = QDateTimeEdit()
        self.start_time_edit.setDateTime(QDateTime.currentDateTime().addDays(-7))
        self.start_time_edit.setCalendarPopup(True)
        filter_layout.addWidget(self.start_time_edit)
        
        filter_layout.addWidget(QLabel("结束时间:"))
        self.end_time_edit = QDateTimeEdit()
        self.end_time_edit.setDateTime(QDateTime.currentDateTime())
        self.end_time_edit.setCalendarPopup(True)
        filter_layout.addWidget(self.end_time_edit)
        
        # 筛选按钮
        self.filter_btn = QPushButton("筛选")
        self.filter_btn.clicked.connect(self.apply_filter)
        filter_layout.addWidget(self.filter_btn)
        
        self.clear_filter_btn = QPushButton("清除筛选")
        self.clear_filter_btn.clicked.connect(self.clear_filter)
        filter_layout.addWidget(self.clear_filter_btn)
        
        filter_layout.addStretch()
        layout.addWidget(filter_group)
        
        # 操作按钮区域
        btn_group = QGroupBox("操作")
        btn_layout = QHBoxLayout(btn_group)
        
        self.refresh_btn = QPushButton("刷新")
        self.refresh_btn.clicked.connect(self.load_logs)
        btn_layout.addWidget(self.refresh_btn)
        
        self.detail_btn = QPushButton("查看详情")
        self.detail_btn.clicked.connect(self.view_detail)
        btn_layout.addWidget(self.detail_btn)
        
        self.clear_test_btn = QPushButton("清理测试日志")
        self.clear_test_btn.clicked.connect(self.clear_test_logs)
        btn_layout.addWidget(self.clear_test_btn)
        
        self.export_btn = QPushButton("导出日志")
        self.export_btn.clicked.connect(self.export_logs)
        btn_layout.addWidget(self.export_btn)
        
        btn_layout.addStretch()
        layout.addWidget(btn_group)
        
        # 分页控制
        page_layout = QHBoxLayout()
        
        self.prev_btn = QPushButton("上一页")
        self.prev_btn.clicked.connect(self.prev_page)
        page_layout.addWidget(self.prev_btn)
        
        self.page_label = QLabel("第 1 页")
        page_layout.addWidget(self.page_label)
        
        self.next_btn = QPushButton("下一页")
        self.next_btn.clicked.connect(self.next_page)
        page_layout.addWidget(self.next_btn)
        
        self.total_label = QLabel("共 0 条")
        page_layout.addWidget(self.total_label)
        
        page_layout.addStretch()
        layout.addLayout(page_layout)
        
        # 日志列表
        self.table = QTableWidget()
        self.table.setColumnCount(9)
        self.table.setHorizontalHeaderLabels([
            "ID", "用户", "操作类型", "操作描述", "目标类型", "目标ID", 
            "权限编码", "结果", "创建时间"
        ])
        
        # 设置表头自适应
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Fixed)
        header.setSectionResizeMode(1, QHeaderView.Fixed)
        header.setSectionResizeMode(2, QHeaderView.Fixed)
        header.setSectionResizeMode(3, QHeaderView.Stretch)
        header.setSectionResizeMode(4, QHeaderView.Fixed)
        header.setSectionResizeMode(5, QHeaderView.Fixed)
        header.setSectionResizeMode(6, QHeaderView.Fixed)
        header.setSectionResizeMode(7, QHeaderView.Fixed)
        header.setSectionResizeMode(8, QHeaderView.Fixed)
        
        # 设置列宽
        self.table.setColumnWidth(0, 50)
        self.table.setColumnWidth(1, 100)
        self.table.setColumnWidth(2, 120)
        self.table.setColumnWidth(4, 80)
        self.table.setColumnWidth(5, 80)
        self.table.setColumnWidth(6, 150)
        self.table.setColumnWidth(7, 60)
        self.table.setColumnWidth(8, 150)
        
        # 设置选择模式
        self.table.setSelectionMode(QTableWidget.SingleSelection)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        
        layout.addWidget(self.table)
    
    def load_filters(self):
        """加载筛选选项"""
        # 加载用户列表
        users = self.permission_service.get_all_users()
        for user in users:
            display_name = user.get('display_name', user['username'])
            self.user_combo.addItem(display_name, user['username'])
        
        # 加载操作类型列表
        operation_types = self.permission_service.get_operation_types()
        for op_type in operation_types:
            self.type_combo.addItem(op_type, op_type)
    
    def load_logs(self):
        """加载操作日志"""
        try:
            # 获取筛选条件
            username = self.user_combo.currentData()
            operation_type = self.type_combo.currentData()
            start_time = self.start_time_edit.dateTime().toString(Qt.ISODate)
            end_time = self.end_time_edit.dateTime().toString(Qt.ISODate)
            
            # 查询日志数量
            total_count = self.permission_service.get_operation_log_count(
                username=username,
                operation_type=operation_type,
                start_time=start_time,
                end_time=end_time
            )
            
            self.total_label.setText(f"共 {total_count} 条")
            
            # 计算总页数
            total_pages = (total_count + self.page_size - 1) // self.page_size
            if total_pages == 0:
                total_pages = 1
            
            # 查询日志列表
            offset = self.current_page * self.page_size
            logs = self.permission_service.get_operation_logs(
                username=username,
                operation_type=operation_type,
                start_time=start_time,
                end_time=end_time,
                limit=self.page_size,
                offset=offset
            )
            
            # 更新表格
            self.table.setRowCount(len(logs))
            
            for i, log in enumerate(logs):
                self.table.setItem(i, 0, QTableWidgetItem(str(log['id'])))
                self.table.setItem(i, 1, QTableWidgetItem(log['username'] or ""))
                self.table.setItem(i, 2, QTableWidgetItem(log['operation_type'] or ""))
                self.table.setItem(i, 3, QTableWidgetItem(log['operation_desc'] or ""))
                self.table.setItem(i, 4, QTableWidgetItem(log['target_type'] or ""))
                self.table.setItem(i, 5, QTableWidgetItem(log['target_id'] or ""))
                self.table.setItem(i, 6, QTableWidgetItem(log['permission_code'] or ""))
                
                # 结果显示
                result = log['result'] or ""
                result_item = QTableWidgetItem(result)
                if result == 'success':
                    result_item.setBackground(Qt.green)
                elif result == 'denied':
                    result_item.setBackground(Qt.red)
                self.table.setItem(i, 7, result_item)
                
                self.table.setItem(i, 8, QTableWidgetItem(log['created_at'] or ""))
            
            # 更新分页状态
            self.page_label.setText(f"第 {self.current_page + 1} 页")
            self.prev_btn.setEnabled(self.current_page > 0)
            self.next_btn.setEnabled(self.current_page < total_pages - 1)
            
        except Exception as e:
            logging.error(f"加载操作日志失败: {e}")
            QMessageBox.warning(self, "错误", f"加载日志失败: {e}")
    
    def apply_filter(self):
        """应用筛选条件"""
        self.current_page = 0
        self.load_logs()
    
    def clear_filter(self):
        """清除筛选条件"""
        self.user_combo.setCurrentIndex(0)
        self.type_combo.setCurrentIndex(0)
        self.start_time_edit.setDateTime(QDateTime.currentDateTime().addDays(-7))
        self.end_time_edit.setDateTime(QDateTime.currentDateTime())
        self.current_page = 0
        self.load_logs()
    
    def prev_page(self):
        """上一页"""
        if self.current_page > 0:
            self.current_page -= 1
            self.load_logs()
    
    def next_page(self):
        """下一页"""
        self.current_page += 1
        self.load_logs()
    
    def view_detail(self):
        """查看日志详情"""
        selected_rows = self.table.selectedItems()
        if not selected_rows:
            QMessageBox.warning(self, "提示", "请先选择一条日志")
            return
        
        row = selected_rows[0].row()
        log_id = self.table.item(row, 0).text()
        
        # 查询日志详情
        logs = self.permission_service.get_operation_logs(limit=1, offset=0)
        
        # 从当前显示的日志中获取详情
        username = self.table.item(row, 1).text()
        operation_type = self.table.item(row, 2).text()
        
        # 查询完整日志
        all_logs = self.permission_service.get_operation_logs(
            username=username,
            operation_type=operation_type,
            limit=100,
            offset=0
        )
        
        # 找到对应的日志
        log = None
        for l in all_logs:
            if str(l['id']) == log_id:
                log = l
                break
        
        if not log:
            QMessageBox.warning(self, "错误", "未找到日志详情")
            return
        
        # 显示详情对话框
        dialog = LogDetailDialog(log, self)
        dialog.exec()
    
    def clear_test_logs(self):
        """清理测试日志 - 安全版本"""
        # 权限检查
        if not self.permission_checker.check_permission(PermissionCodes.OP_LOG_DELETE, self):
            return
        
        # 显示清理选项对话框
        dialog = ClearTestLogsDialog(self)
        if dialog.exec() != QDialog.Accepted:
            return
        
        # 获取清理选项
        options = dialog.get_clear_options()
        
        # 构建清理条件
        conditions = {}
        warning_msg = "确定要清理以下日志吗？\n\n"
        
        if options['by_user']:
            conditions['username'] = options['username']
            warning_msg += f"用户: {options['username']}\n"
        
        if options['by_time']:
            conditions['start_time'] = options['start_time']
            conditions['end_time'] = options['end_time']
            warning_msg += f"时间范围: {options['start_time']} 至 {options['end_time']}\n"
        
        if options['by_type']:
            conditions['operation_type'] = options['operation_type']
            warning_msg += f"操作类型: {options['operation_type']}\n"
            # 如果选择按操作类型清理，添加警告
            if options['operation_type'] == 'permission_denied':
                warning_msg += "\n⚠️ 警告: permission_denied 类型包含真实的无权限操作记录，清理可能影响审计追溯！"
        
        if not conditions:
            QMessageBox.warning(self, "提示", "请至少选择一个清理条件")
            return
        
        # 最终确认
        reply = QMessageBox.question(
            self,
            "确认清理",
            warning_msg,
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply != QMessageBox.Yes:
            return
        
        try:
            # 删除日志
            count = self.permission_service.delete_operation_logs(
                username=conditions.get('username'),
                operation_type=conditions.get('operation_type'),
                start_time=conditions.get('start_time'),
                end_time=conditions.get('end_time')
            )
            
            # 记录清理操作日志
            self.permission_service.log_operation(
                username=self.username,
                operation_type='clear_test_logs',
                operation_desc=f'清理了 {count} 条日志',
                target_type='operation_logs',
                target_id='test_logs',
                detail={
                    'deleted_count': count,
                    'conditions': conditions
                },
                permission_code=PermissionCodes.OP_LOG_DELETE,
                result='success'
            )
            
            QMessageBox.information(self, "成功", f"已清理 {count} 条日志")
            self.load_logs()
            
        except Exception as e:
            logging.error(f"清理日志失败: {e}")
            QMessageBox.warning(self, "错误", f"清理失败: {e}")
    
    def export_logs(self):
        """导出日志"""
        # 权限检查
        if not self.permission_checker.check_permission(PermissionCodes.OP_LOG_EXPORT, self):
            return
        
        # 选择保存文件路径
        default_filename = f"操作日志_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "导出操作日志",
            default_filename,
            "Excel文件 (*.xlsx);;CSV文件 (*.csv)"
        )
        
        if not file_path:
            return
        
        try:
            # 获取当前筛选条件下的所有日志（不限制数量）
            username = self.user_combo.currentData()
            operation_type = self.type_combo.currentData()
            start_time = self.start_time_edit.dateTime().toString(Qt.ISODate)
            end_time = self.end_time_edit.dateTime().toString(Qt.ISODate)
            
            # 查询所有日志
            logs = self.permission_service.get_operation_logs(
                username=username,
                operation_type=operation_type,
                start_time=start_time,
                end_time=end_time,
                limit=10000,  # 最大导出数量
                offset=0
            )
            
            if not logs:
                QMessageBox.warning(self, "提示", "没有可导出的日志")
                return
            
            # 根据文件类型导出
            if file_path.endswith('.xlsx'):
                self._export_to_xlsx(logs, file_path)
            else:
                self._export_to_csv(logs, file_path)
            
            # 记录导出操作日志
            self.permission_service.log_operation(
                username=self.username,
                operation_type='log_export',
                operation_desc=f'导出操作日志到 {os.path.basename(file_path)}',
                target_type='operation_logs',
                target_id=file_path,
                detail={'count': len(logs), 'file_path': file_path},
                permission_code=PermissionCodes.OP_LOG_EXPORT,
                result='success'
            )
            
            QMessageBox.information(self, "成功", f"已导出 {len(logs)} 条日志到:\n{file_path}")
            
        except Exception as e:
            logging.error(f"导出日志失败: {e}")
            QMessageBox.warning(self, "错误", f"导出失败: {e}")
    
    def _export_to_xlsx(self, logs: List[Dict], file_path: str):
        """导出为Excel文件"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "操作日志"
            
            # 设置表头
            headers = ["ID", "用户", "操作类型", "操作描述", "目标类型", "目标ID", 
                       "权限编码", "结果", "IP地址", "机器名", "创建时间", "详细信息"]
            
            # 表头样式
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
            header_alignment = Alignment(horizontal="center")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # 写入数据
            for row, log in enumerate(logs, 2):
                ws.cell(row=row, column=1, value=log.get('id'))
                ws.cell(row=row, column=2, value=log.get('username', ''))
                ws.cell(row=row, column=3, value=log.get('operation_type', ''))
                ws.cell(row=row, column=4, value=log.get('operation_desc', ''))
                ws.cell(row=row, column=5, value=log.get('target_type', ''))
                ws.cell(row=row, column=6, value=log.get('target_id', ''))
                ws.cell(row=row, column=7, value=log.get('permission_code', ''))
                ws.cell(row=row, column=8, value=log.get('result', ''))
                ws.cell(row=row, column=9, value=log.get('ip_address', ''))
                ws.cell(row=row, column=10, value=log.get('machine_name', ''))
                ws.cell(row=row, column=11, value=log.get('created_at', ''))
                
                # 详细信息转为JSON字符串
                detail = log.get('detail', {})
                if detail:
                    ws.cell(row=row, column=12, value=json.dumps(detail, ensure_ascii=False))
                else:
                    ws.cell(row=row, column=12, value='')
            
            # 设置列宽
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 30
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 18
            ws.column_dimensions['H'].width = 10
            ws.column_dimensions['I'].width = 15
            ws.column_dimensions['J'].width = 15
            ws.column_dimensions['K'].width = 20
            ws.column_dimensions['L'].width = 30
            
            wb.save(file_path)
            
        except ImportError:
            # 如果没有openpyxl，导出为CSV
            logging.warning("openpyxl未安装，导出为CSV格式")
            csv_path = file_path.replace('.xlsx', '.csv')
            self._export_to_csv(logs, csv_path)
            QMessageBox.information(self, "提示", 
                f"Excel库未安装，已导出为CSV格式:\n{csv_path}")
    
    def _export_to_csv(self, logs: List[Dict], file_path: str):
        """导出为CSV文件"""
        import csv
        
        with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            
            # 写入表头
            headers = ["ID", "用户", "操作类型", "操作描述", "目标类型", "目标ID", 
                       "权限编码", "结果", "IP地址", "机器名", "创建时间", "详细信息"]
            writer.writerow(headers)
            
            # 写入数据
            for log in logs:
                detail = log.get('detail', {})
                detail_str = json.dumps(detail, ensure_ascii=False) if detail else ''
                
                writer.writerow([
                    log.get('id'),
                    log.get('username', ''),
                    log.get('operation_type', ''),
                    log.get('operation_desc', ''),
                    log.get('target_type', ''),
                    log.get('target_id', ''),
                    log.get('permission_code', ''),
                    log.get('result', ''),
                    log.get('ip_address', ''),
                    log.get('machine_name', ''),
                    log.get('created_at', ''),
                    detail_str
                ])


class ClearTestLogsDialog(QDialog):
    """清理测试日志选项对话框"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("清理测试日志")
        self.setFixedSize(450, 350)
        self._init_ui()
    
    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        
        # 提示信息
        warning_label = QLabel("⚠️ 请谨慎选择清理条件，避免误删真实审计日志")
        warning_label.setStyleSheet("color: #d32f2f; font-weight: bold;")
        layout.addWidget(warning_label)
        
        # 按用户清理
        user_group = QGroupBox("按用户清理")
        user_layout = QHBoxLayout(user_group)
        
        self.by_user_checkbox = QCheckBox("启用")
        self.by_user_checkbox.setChecked(True)
        user_layout.addWidget(self.by_user_checkbox)
        
        user_layout.addWidget(QLabel("用户名:"))
        self.username_input = QLineEdit()
        self.username_input.setPlaceholderText("输入测试用户名（如 test_*）")
        self.username_input.setText("test_")  # 默认测试用户前缀
        user_layout.addWidget(self.username_input)
        
        layout.addWidget(user_group)
        
        # 按时间范围清理
        time_group = QGroupBox("按时间范围清理")
        time_layout = QHBoxLayout(time_group)
        
        self.by_time_checkbox = QCheckBox("启用")
        time_layout.addWidget(self.by_time_checkbox)
        
        time_layout.addWidget(QLabel("从:"))
        self.start_time_edit = QDateTimeEdit()
        self.start_time_edit.setDateTime(QDateTime.currentDateTime().addHours(-1))
        self.start_time_edit.setCalendarPopup(True)
        time_layout.addWidget(self.start_time_edit)
        
        time_layout.addWidget(QLabel("到:"))
        self.end_time_edit = QDateTimeEdit()
        self.end_time_edit.setDateTime(QDateTime.currentDateTime())
        self.end_time_edit.setCalendarPopup(True)
        time_layout.addWidget(self.end_time_edit)
        
        layout.addWidget(time_group)
        
        # 按操作类型清理
        type_group = QGroupBox("按操作类型清理（谨慎使用）")
        type_layout = QHBoxLayout(type_group)
        
        self.by_type_checkbox = QCheckBox("启用")
        type_layout.addWidget(self.by_type_checkbox)
        
        type_layout.addWidget(QLabel("操作类型:"))
        self.type_combo = QComboBox()
        self.type_combo.addItems([
            'permission_denied',
            'clear_test_logs',
            'log_export',
            'user_create',
            'user_edit',
            'user_unlock',
            'config_create',
            'config_edit',
            'config_delete',
            'role_permission_update'
        ])
        type_layout.addWidget(self.type_combo)
        
        # 警告标签
        type_warning = QLabel("⚠️ permission_denied 包含真实无权限记录")
        type_warning.setStyleSheet("color: #d32f2f;")
        type_layout.addWidget(type_warning)
        
        layout.addWidget(type_group)
        
        # 按钮
        btn_layout = QHBoxLayout()
        
        confirm_btn = QPushButton("确认清理")
        confirm_btn.clicked.connect(self.accept)
        btn_layout.addWidget(confirm_btn)
        
        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)
        
        layout.addLayout(btn_layout)
    
    def get_clear_options(self) -> dict:
        """获取清理选项"""
        return {
            'by_user': self.by_user_checkbox.isChecked(),
            'username': self.username_input.text().strip(),
            'by_time': self.by_time_checkbox.isChecked(),
            'start_time': self.start_time_edit.dateTime().toString(Qt.ISODate) if self.by_time_checkbox.isChecked() else None,
            'end_time': self.end_time_edit.dateTime().toString(Qt.ISODate) if self.by_time_checkbox.isChecked() else None,
            'by_type': self.by_type_checkbox.isChecked(),
            'operation_type': self.type_combo.currentText() if self.by_type_checkbox.isChecked() else None
        }


class LogDetailDialog(QDialog):
    """日志详情对话框"""
    
    def __init__(self, log: Dict, parent=None):
        super().__init__(parent)
        self.log = log
        self.setWindowTitle("日志详情")
        self.setMinimumSize(600, 400)
        
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # 基本信息
        info_layout = QVBoxLayout()
        
        info_layout.addWidget(QLabel(f"日志ID: {self.log['id']}"))
        info_layout.addWidget(QLabel(f"用户: {self.log['username']}"))
        info_layout.addWidget(QLabel(f"操作类型: {self.log['operation_type']}"))
        info_layout.addWidget(QLabel(f"操作描述: {self.log['operation_desc']}"))
        info_layout.addWidget(QLabel(f"目标类型: {self.log['target_type']}"))
        info_layout.addWidget(QLabel(f"目标ID: {self.log['target_id']}"))
        info_layout.addWidget(QLabel(f"权限编码: {self.log['permission_code']}"))
        info_layout.addWidget(QLabel(f"结果: {self.log['result']}"))
        info_layout.addWidget(QLabel(f"IP地址: {self.log.get('ip_address', '')}"))
        info_layout.addWidget(QLabel(f"机器名: {self.log.get('machine_name', '')}"))
        info_layout.addWidget(QLabel(f"创建时间: {self.log['created_at']}"))
        
        layout.addLayout(info_layout)
        
        # 详细信息
        detail_group = QGroupBox("详细信息")
        detail_layout = QVBoxLayout(detail_group)
        
        detail_text = QTextEdit()
        detail_text.setReadOnly(True)
        
        # 格式化显示detail字段
        detail_json = json.dumps(self.log.get('detail', {}), indent=2, ensure_ascii=False)
        detail_text.setText(detail_json)
        
        detail_layout.addWidget(detail_text)
        layout.addWidget(detail_group)
        
        # 关闭按钮
        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(self.close)
        layout.addWidget(close_btn)