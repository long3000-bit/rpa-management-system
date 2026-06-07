from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, 
    QPushButton, QComboBox, QTableWidget,
    QTableWidgetItem, QMessageBox, QGroupBox,
    QLineEdit, QHeaderView
)
from PySide6.QtCore import Qt
from PySide6.QtGui import QFont
from datetime import datetime
from decimal import Decimal
from pathlib import Path
import logging
import json
import openpyxl

from app.storage.database import Database
from app.ui.widgets.table_highlight import enable_table_highlight


class YsbDataQueryPage(QWidget):
    
    def __init__(self, db: Database):
        super().__init__()
        self.db = db
        self._init_ui()
        self._load_account_periods()
    
    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        
        filter_group = QGroupBox("查询条件")
        filter_layout = QVBoxLayout(filter_group)
        
        filter_row1 = QHBoxLayout()
        filter_row1.addWidget(QLabel("核算年:"))
        
        self.year_combo = QComboBox()
        self.year_combo.setMinimumWidth(100)
        current_year = datetime.now().year
        for y in range(current_year - 2, current_year + 2):
            self.year_combo.addItem(str(y), y)
        self.year_combo.setCurrentText(str(current_year))
        self.year_combo.currentIndexChanged.connect(self._on_period_changed)
        filter_row1.addWidget(self.year_combo)
        
        filter_row1.addWidget(QLabel("核算月:"))
        
        self.month_combo = QComboBox()
        self.month_combo.setMinimumWidth(80)
        for m in range(1, 13):
            self.month_combo.addItem(f"{m}月", m)
        self.month_combo.setCurrentText(f"{datetime.now().month}月")
        self.month_combo.currentIndexChanged.connect(self._on_period_changed)
        filter_row1.addWidget(self.month_combo)
        
        self.refresh_btn = QPushButton("刷新")
        self.refresh_btn.clicked.connect(self._load_account_periods)
        filter_row1.addWidget(self.refresh_btn)
        
        filter_row1.addStretch()
        filter_layout.addLayout(filter_row1)
        
        filter_row2 = QHBoxLayout()
        filter_row2.addWidget(QLabel("数据类型:"))
        
        self.data_type_combo = QComboBox()
        self.data_type_combo.addItem("本月支付订单", "order")
        self.data_type_combo.addItem("本月支付账单明细", "bill")
        self.data_type_combo.currentIndexChanged.connect(self._query_data)
        filter_row2.addWidget(self.data_type_combo)
        
        filter_row2.addStretch()
        filter_layout.addLayout(filter_row2)
        
        filter_row3 = QHBoxLayout()
        filter_row3.addWidget(QLabel("搜索:"))
        
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("输入商品名称、企业名称或条形码搜索...")
        self.search_input.setMinimumWidth(400)
        self.search_input.textChanged.connect(self._on_search_changed)
        filter_row3.addWidget(self.search_input)
        
        self.query_btn = QPushButton("查询")
        self.query_btn.clicked.connect(self._query_data)
        filter_row3.addWidget(self.query_btn)
        
        self.export_btn = QPushButton("导出")
        self.export_btn.clicked.connect(self._export_data)
        filter_row3.addWidget(self.export_btn)
        
        filter_row3.addStretch()
        filter_layout.addLayout(filter_row3)
        
        layout.addWidget(filter_group)
        
        result_group = QGroupBox("查询结果")
        result_layout = QVBoxLayout(result_group)
        
        self.result_table = QTableWidget()
        self.result_table.setAlternatingRowColors(True)
        self.result_table.setSortingEnabled(True)
        result_layout.addWidget(self.result_table)
        
        self.result_label = QLabel("共 0 条记录")
        result_layout.addWidget(self.result_label)
        
        layout.addWidget(result_group)
    
    def _load_account_periods(self):
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT DISTINCT account_year, account_month
                FROM ysb_import_batches
                WHERE import_status = 'success'
                AND account_year IS NOT NULL AND account_month IS NOT NULL
                ORDER BY account_year DESC, account_month DESC
            ''')
            
            rows = cursor.fetchall()
            
            logging.info(f"药师帮数据查询 - 加载核算年月, 共找到 {len(rows)} 个核算年月")
            
            if rows:
                latest = rows[0]
                self.year_combo.setCurrentText(str(latest['account_year']))
                self.month_combo.setCurrentText(f"{latest['account_month']}月")
            
            self._query_data()
        except Exception as e:
            logging.error(f"加载核算年月失败: {str(e)}", exc_info=True)
            self.result_table.setRowCount(0)
            self.result_label.setText(f"查询失败: {str(e)}")
    
    def _on_period_changed(self):
        self._query_data()
    
    def _on_search_changed(self):
        pass
    
    def _query_data(self):
        try:
            data_type = self.data_type_combo.currentData()
            search_text = self.search_input.text().strip()
            account_year = self.year_combo.currentData()
            account_month = self.month_combo.currentData()
            
            logging.info(f"执行数据查询: 类型={data_type}, 年月={account_year}-{account_month}")
            
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            if data_type == "order":
                self._query_order_data(cursor, search_text, account_year, account_month)
            else:
                self._query_bill_data(cursor, search_text, account_year, account_month)
        except Exception as e:
            logging.error(f"查询数据失败: {str(e)}", exc_info=True)
            self.result_table.setRowCount(0)
            self.result_label.setText(f"查询失败: {str(e)}")
    
    def _query_order_data(self, cursor, search_text: str, account_year: int, account_month: int):
        if search_text:
            cursor.execute('''
                SELECT s.raw_data, s.actual_payment_amount AS amount_total
                FROM ysb_supplier_summary s
                JOIN ysb_import_batches b ON b.batch_id = s.import_batch_id
                WHERE b.import_status = 'success'
                AND b.account_year = ?
                AND b.account_month = ?
                AND (
                    s.ysb_company_name LIKE ? OR
                    s.ysb_supplier_name LIKE ? OR
                    s.raw_data LIKE ?
                )
                ORDER BY s.raw_row_index
            ''', (account_year, account_month, f"%{search_text}%", f"%{search_text}%", f"%{search_text}%"))
        else:
            cursor.execute('''
                SELECT s.raw_data, s.actual_payment_amount AS amount_total
                FROM ysb_supplier_summary s
                JOIN ysb_import_batches b ON b.batch_id = s.import_batch_id
                WHERE b.import_status = 'success'
                AND b.account_year = ?
                AND b.account_month = ?
                ORDER BY s.raw_row_index
            ''', (account_year, account_month))
        
        rows = cursor.fetchall()
        logging.info(f"查询本月支付订单底表数据: {len(rows)} 条")
        
        headers = self._load_excel_headers(cursor, "supplier", account_year, account_month)
        self._show_raw_data_rows(
            rows,
            empty_message="未查询到本月支付订单底表数据",
            preferred_headers=headers,
            amount_column="amount_total",
            amount_fields=["实际支付金额(已减退款)", "实际支付金额", "actual_payment_amount"]
        )
    
    def _query_bill_data(self, cursor, search_text: str, account_year: int, account_month: int):
        if search_text:
            cursor.execute('''
                SELECT d.raw_data, d.discount_amount AS amount_total
                FROM ysb_detail_data d
                JOIN ysb_import_batches b ON b.batch_id = d.import_batch_id
                WHERE b.import_status = 'success'
                AND b.account_year = ?
                AND b.account_month = ?
                AND (
                    d.product_name LIKE ? OR
                    d.ysb_company_name LIKE ? OR
                    d.barcode LIKE ? OR
                    d.manufacturer LIKE ? OR
                    d.raw_data LIKE ?
                )
                ORDER BY d.raw_row_index
            ''', (
                account_year, account_month,
                f"%{search_text}%", f"%{search_text}%", f"%{search_text}%",
                f"%{search_text}%", f"%{search_text}%"
            ))
        else:
            cursor.execute('''
                SELECT d.raw_data, d.discount_amount AS amount_total
                FROM ysb_detail_data d
                JOIN ysb_import_batches b ON b.batch_id = d.import_batch_id
                WHERE b.import_status = 'success'
                AND b.account_year = ?
                AND b.account_month = ?
                ORDER BY d.raw_row_index
            ''', (account_year, account_month))
        
        rows = cursor.fetchall()
        logging.info(f"查询本月支付账单明细底表数据: {len(rows)} 条")
        
        headers = self._load_excel_headers(cursor, "detail", account_year, account_month)
        self._show_raw_data_rows(
            rows,
            empty_message="未查询到本月支付账单明细底表数据",
            preferred_headers=headers,
            amount_column="amount_total",
            amount_fields=["折后价总额", "discount_amount", "实际支付金额(已减退款)", "商品总金额"]
        )
    
    def _load_excel_headers(self, cursor, sheet_type: str, account_year: int, account_month: int) -> list[str]:
        cursor.execute('''
            SELECT file_path, sheet_name
            FROM ysb_import_batches
            WHERE import_status = 'success'
            AND account_year = ?
            AND account_month = ?
            AND sheet_type = ?
            ORDER BY imported_at DESC
            LIMIT 1
        ''', (account_year, account_month, sheet_type))
        
        batch = cursor.fetchone()
        if not batch or not batch['file_path'] or not batch['sheet_name']:
            return []
        
        file_path = Path(batch['file_path'])
        if not file_path.exists():
            logging.warning(f"原始Excel文件不存在，无法读取表头顺序: {file_path}")
            return []
        
        workbook = None
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            if batch['sheet_name'] not in workbook.sheetnames:
                logging.warning(f"原始Excel中未找到工作表: {batch['sheet_name']}")
                return []
            
            sheet = workbook[batch['sheet_name']]
            for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
                return [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
        except Exception as e:
            logging.warning(f"读取原始Excel表头失败: {e}")
            return []
        finally:
            if workbook:
                workbook.close()
        
        return []
    
    def _show_raw_data_rows(
        self,
        rows,
        empty_message: str,
        preferred_headers: list[str] = None,
        amount_column: str = "",
        amount_fields: list[str] = None
    ):
        self.result_table.setSortingEnabled(False)
        
        if not rows:
            self.result_table.setRowCount(0)
            self.result_table.setColumnCount(0)
            self.result_label.setText("共 0 条记录，金额合计 0.00")
            logging.warning(empty_message)
            self.result_table.setSortingEnabled(True)
            return
        
        headers = list(preferred_headers or [])
        raw_data_list = []
        amount_values = []
        for row in rows:
            try:
                raw_data = json.loads(row['raw_data']) if row['raw_data'] else {}
                raw_data_list.append(raw_data)
                amount_values.append(row[amount_column] if amount_column and amount_column in row.keys() else None)
                for header in raw_data.keys():
                    if header not in headers:
                        headers.append(header)
            except Exception as e:
                logging.warning(f"解析raw_data失败: {e}")
        
        logging.info(f"解析成功 {len(raw_data_list)} 条，共 {len(headers)} 个字段")
        
        self.result_table.setColumnCount(len(headers))
        self.result_table.setHorizontalHeaderLabels(headers)
        self.result_table.setRowCount(len(raw_data_list))
        
        for row_idx, raw_data in enumerate(raw_data_list):
            for col_idx, header in enumerate(headers):
                value = raw_data.get(header, '')
                self.result_table.setItem(row_idx, col_idx, QTableWidgetItem(str(value) if value is not None else ''))
        
        self.result_table.resizeColumnsToContents()
        total_amount = self._sum_amount(raw_data_list, amount_fields or [], amount_values)
        self.result_label.setText(f"共 {len(raw_data_list)} 条记录，金额合计 {total_amount:.2f}")
        self.result_table.setSortingEnabled(True)
    
    def _sum_amount(
        self,
        rows: list[dict],
        amount_fields: list[str],
        amount_values: list = None
    ) -> Decimal:
        total = Decimal("0")
        amount_values = amount_values or []
        for index, row in enumerate(rows):
            amount_value = amount_values[index] if index < len(amount_values) else None
            if amount_value in (None, ""):
                amount_value = self._get_first_value(row, amount_fields)
            total += self._to_decimal(amount_value)
        return total
    
    def _get_first_value(self, row: dict, fields: list[str]):
        if not row:
            return None
        
        normalized = {str(key).strip().lower(): value for key, value in row.items()}
        for field in fields:
            value = row.get(field)
            if value not in (None, ""):
                return value
            
            value = normalized.get(str(field).strip().lower())
            if value not in (None, ""):
                return value
        
        return None
    
    def _to_decimal(self, value) -> Decimal:
        if value in (None, ""):
            return Decimal("0")
        
        text = str(value).replace(",", "").replace("￥", "").replace("¥", "").strip()
        try:
            return Decimal(text)
        except Exception:
            return Decimal("0")
    
    def _export_data(self):
        QMessageBox.information(self, "提示", "导出功能开发中...")
