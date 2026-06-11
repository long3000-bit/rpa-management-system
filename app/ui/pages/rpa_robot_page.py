from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, 
    QPushButton, QComboBox, QTableWidget, QTableWidgetItem,
    QMessageBox, QGroupBox, QLineEdit, QTextEdit, 
    QSpinBox, QCheckBox, QFileDialog, QSplitter, QTabWidget,
    QProgressBar, QHeaderView
)
from PySide6.QtCore import Qt, QThread, Signal
from pathlib import Path
import logging
import json
import uuid
from datetime import datetime
from typing import Dict, List, Tuple

from app.storage.database import Database
from app.core.permission_service import PermissionService
from app.core.rpa_import_service import RpaImportService
from app.core.rpa_template_service import RpaTemplateService
from app.core.rpa_exe_config_service import RpaExeConfigService
from app.core.rpa_executor import RpaExecutor
from app.core.rpa_auto_login import RpaAutoLogin
from app.core.data_permission_service import DataPermissionService
from app.core.permission_checker import PermissionChecker, PermissionCodes


class RpaRobotPage(QWidget):
    
    def __init__(self, db: Database, username: str = None, role_code: str = None):
        super().__init__()
        self.db = db
        self.username = username or 'admin'
        self.role_code = role_code or 'store_manager'
        self.permission_service = PermissionService(db)
        self.permission_checker = PermissionChecker(db, self.username)
        self.import_service = RpaImportService(db)
        self.template_service = RpaTemplateService(db)
        self.exe_config_service = RpaExeConfigService(db)
        self.executor = RpaExecutor(db)
        self.auto_login = RpaAutoLogin(db)
        self.data_permission_service = DataPermissionService(db)
        
        self.excel_file_path = ""
        self.sheet_names = []
        self.current_sheet = ""
        self.preview_data = []
        self.headers = []
        self.import_batch_id = ""
        self.field_mapping = {}
        self.current_template_id = ""
        self.valid_rows = []
        self.invalid_rows = []
        self.is_executing = False
        self.current_task_id = ""
        
        self._init_ui()
        self._load_templates()
        self._load_exe_configs()
        self._load_batches()
    
    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        
        splitter = QSplitter(Qt.Vertical)
        
        top_widget = QWidget()
        top_layout = QVBoxLayout(top_widget)
        top_layout.setSpacing(5)
        
        import_group = QGroupBox("数据导入区")
        import_layout = QVBoxLayout(import_group)
        
        file_row = QHBoxLayout()
        file_row.addWidget(QLabel("Excel文件:"))
        self.file_path_input = QLineEdit()
        self.file_path_input.setReadOnly(True)
        self.file_path_input.setMinimumWidth(300)
        file_row.addWidget(self.file_path_input)
        
        self.select_file_btn = QPushButton("选择文件")
        self.select_file_btn.clicked.connect(self._select_excel_file)
        file_row.addWidget(self.select_file_btn)
        
        file_row.addWidget(QLabel("工作表:"))
        self.sheet_combo = QComboBox()
        self.sheet_combo.setMinimumWidth(150)
        self.sheet_combo.currentTextChanged.connect(self._on_sheet_changed)
        file_row.addWidget(self.sheet_combo)
        
        file_row.addWidget(QLabel("导入模板:"))
        self.template_combo = QComboBox()
        self.template_combo.setMinimumWidth(200)
        self.template_combo.currentIndexChanged.connect(self._on_template_changed)
        file_row.addWidget(self.template_combo)
        
        self.new_template_btn = QPushButton("新建模板")
        file_row.addWidget(self.new_template_btn)
        
        file_row.addStretch()
        import_layout.addLayout(file_row)
        
        preview_row = QHBoxLayout()
        self.preview_btn = QPushButton("预览数据")
        self.preview_btn.clicked.connect(self._preview_data)
        preview_row.addWidget(self.preview_btn)
        
        self.validate_btn = QPushButton("校验数据")
        self.validate_btn.clicked.connect(self._validate_data)
        preview_row.addWidget(self.validate_btn)
        
        self.import_btn = QPushButton("导入系统")
        self.import_btn.clicked.connect(self._import_data)
        self.import_btn.setEnabled(False)
        preview_row.addWidget(self.import_btn)
        
        self.import_status_label = QLabel("")
        preview_row.addWidget(self.import_status_label)
        
        preview_row.addStretch()
        import_layout.addLayout(preview_row)
        
        top_layout.addWidget(import_group)
        
        data_group = QGroupBox("待处理数据区")
        data_layout = QVBoxLayout(data_group)
        
        filter_row = QHBoxLayout()
        filter_row.addWidget(QLabel("导入批次:"))
        self.batch_combo = QComboBox()
        self.batch_combo.setMinimumWidth(200)
        self.batch_combo.currentIndexChanged.connect(self._on_batch_changed)
        filter_row.addWidget(self.batch_combo)
        
        self.refresh_batch_btn = QPushButton("刷新批次")
        self.refresh_batch_btn.clicked.connect(self._load_batches)
        filter_row.addWidget(self.refresh_batch_btn)
        
        filter_row.addWidget(QLabel("状态筛选:"))
        self.status_filter_combo = QComboBox()
        self.status_filter_combo.addItem("全部", "all")
        self.status_filter_combo.addItem("待处理", "pending")
        self.status_filter_combo.addItem("成功", "success")
        self.status_filter_combo.addItem("失败", "failed")
        self.status_filter_combo.addItem("跳过", "skipped")
        self.status_filter_combo.addItem("重复", "duplicate")
        self.status_filter_combo.currentIndexChanged.connect(self._filter_data)
        filter_row.addWidget(self.status_filter_combo)
        
        filter_row.addWidget(QLabel("关键字:"))
        self.keyword_search = QLineEdit()
        self.keyword_search.setPlaceholderText("业务主键/关键字")
        self.keyword_search.setMaximumWidth(150)
        self.keyword_search.textChanged.connect(self._filter_data)
        filter_row.addWidget(self.keyword_search)
        
        filter_row.addStretch()
        data_layout.addLayout(filter_row)
        
        self.data_table = QTableWidget()
        self.data_table.setAlternatingRowColors(True)
        self.data_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.data_table.setMinimumHeight(150)
        data_layout.addWidget(self.data_table)
        
        top_layout.addWidget(data_group)
        
        splitter.addWidget(top_widget)
        
        bottom_widget = QWidget()
        bottom_layout = QVBoxLayout(bottom_widget)
        bottom_layout.setSpacing(5)
        
        config_group = QGroupBox("任务配置区")
        config_layout = QVBoxLayout(config_group)
        
        config_row1 = QHBoxLayout()
        config_row1.addWidget(QLabel("执行模式:"))
        self.exec_mode_combo = QComboBox()
        self.exec_mode_combo.addItem("全部执行", "all")
        self.exec_mode_combo.addItem("仅执行待处理", "pending_only")
        self.exec_mode_combo.addItem("仅重试失败", "retry_failed")
        self.exec_mode_combo.addItem("从指定记录开始", "from_record")
        self.exec_mode_combo.addItem("只预演不提交", "preview_only")
        config_row1.addWidget(self.exec_mode_combo)
        
        config_row1.addWidget(QLabel("起始行:"))
        self.start_row_spin = QSpinBox()
        self.start_row_spin.setMinimum(1)
        self.start_row_spin.setMaximum(999999)
        self.start_row_spin.setValue(1)
        config_row1.addWidget(self.start_row_spin)
        
        config_row1.addWidget(QLabel("最大执行行数:"))
        self.max_rows_spin = QSpinBox()
        self.max_rows_spin.setMinimum(0)
        self.max_rows_spin.setMaximum(999999)
        self.max_rows_spin.setValue(0)
        self.max_rows_spin.setSpecialValueText("不限制")
        config_row1.addWidget(self.max_rows_spin)
        
        config_row1.addWidget(QLabel("目标EXE:"))
        self.exe_config_combo = QComboBox()
        self.exe_config_combo.setMinimumWidth(150)
        config_row1.addWidget(self.exe_config_combo)
        
        config_row1.addStretch()
        config_layout.addLayout(config_row1)
        
        config_row2 = QHBoxLayout()
        self.skip_success_check = QCheckBox("跳过已成功行")
        self.skip_success_check.setChecked(True)
        config_row2.addWidget(self.skip_success_check)
        
        self.screenshot_on_fail_check = QCheckBox("失败时截图")
        self.screenshot_on_fail_check.setChecked(True)
        config_row2.addWidget(self.screenshot_on_fail_check)
        
        self.check_duplicate_check = QCheckBox("执行前重复检查")
        self.check_duplicate_check.setChecked(True)
        config_row2.addWidget(self.check_duplicate_check)
        
        self.auto_export_check = QCheckBox("完成后自动导出Excel")
        self.auto_export_check.setChecked(False)
        config_row2.addWidget(self.auto_export_check)
        
        config_row2.addStretch()
        config_layout.addLayout(config_row2)
        
        bottom_layout.addWidget(config_group)
        
        control_group = QGroupBox("执行控制区")
        control_layout = QVBoxLayout(control_group)
        
        control_row = QHBoxLayout()
        self.connect_exe_btn = QPushButton("连接EXE")
        self.connect_exe_btn.clicked.connect(self._connect_exe)
        control_row.addWidget(self.connect_exe_btn)
        
        self.test_login_btn = QPushButton("测试登录")
        self.test_login_btn.clicked.connect(self._test_login)
        control_row.addWidget(self.test_login_btn)
        
        self.test_current_btn = QPushButton("测试当前记录")
        self.test_current_btn.clicked.connect(self._test_current_record)
        control_row.addWidget(self.test_current_btn)
        
        self.start_btn = QPushButton("开始执行")
        self.start_btn.clicked.connect(self._start_execution)
        self.start_btn.setStyleSheet("background-color: #4CAF50; color: white;")
        control_row.addWidget(self.start_btn)
        
        self.pause_btn = QPushButton("暂停")
        self.pause_btn.clicked.connect(self._pause_execution)
        self.pause_btn.setEnabled(False)
        control_row.addWidget(self.pause_btn)
        
        self.continue_btn = QPushButton("继续")
        self.continue_btn.clicked.connect(self._continue_execution)
        self.continue_btn.setEnabled(False)
        control_row.addWidget(self.continue_btn)
        
        self.stop_btn = QPushButton("停止")
        self.stop_btn.clicked.connect(self._stop_execution)
        self.stop_btn.setEnabled(False)
        self.stop_btn.setStyleSheet("background-color: #f44336; color: white;")
        control_row.addWidget(self.stop_btn)
        
        self.retry_failed_btn = QPushButton("重试失败")
        self.retry_failed_btn.clicked.connect(self._retry_failed)
        control_row.addWidget(self.retry_failed_btn)
        
        self.export_result_btn = QPushButton("导出结果Excel")
        self.export_result_btn.clicked.connect(self._export_result)
        control_row.addWidget(self.export_result_btn)
        
        control_row.addStretch()
        control_layout.addLayout(control_row)
        
        progress_row = QHBoxLayout()
        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        progress_row.addWidget(self.progress_bar)
        
        self.progress_label = QLabel("总行数: 0 | 待处理: 0 | 成功: 0 | 失败: 0 | 跳过: 0")
        progress_row.addWidget(self.progress_label)
        
        progress_row.addStretch()
        control_layout.addLayout(progress_row)
        
        bottom_layout.addWidget(control_group)
        
        log_group = QGroupBox("实时日志区")
        log_layout = QVBoxLayout(log_group)
        
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setMinimumHeight(100)
        log_layout.addWidget(self.log_text)
        
        bottom_layout.addWidget(log_group)
        
        splitter.addWidget(bottom_widget)
        
        splitter.setSizes([400, 300])
        
        layout.addWidget(splitter)
    
    def _select_excel_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "选择Excel文件",
            "",
            "Excel文件 (*.xlsx *.xls)"
        )
        
        if file_path:
            self.excel_file_path = file_path
            self.file_path_input.setText(file_path)
            self._load_sheets()
    
    def _load_sheets(self):
        if not self.excel_file_path:
            return
        
        try:
            import openpyxl
            wb = openpyxl.load_workbook(self.excel_file_path, read_only=True)
            self.sheet_names = wb.sheetnames
            wb.close()
            
            self.sheet_combo.clear()
            self.sheet_combo.addItems(self.sheet_names)
            
            if self.sheet_names:
                self.current_sheet = self.sheet_names[0]
                self._preview_data()
                
        except Exception as e:
            QMessageBox.warning(self, "错误", f"读取Excel文件失败:\n{str(e)}")
            logging.error(f"读取Excel文件失败: {e}")
    
    def _on_sheet_changed(self, sheet_name):
        self.current_sheet = sheet_name
        if sheet_name:
            self._preview_data()
    
    def _preview_data(self):
        if not self.excel_file_path or not self.current_sheet:
            return
        
        try:
            import openpyxl
            wb = openpyxl.load_workbook(self.excel_file_path, read_only=True)
            ws = wb[self.current_sheet]
            
            self.headers = []
            self.preview_data = []
            
            for row_idx, row in enumerate(ws.iter_rows(max_row=100, values_only=True)):
                if row_idx == 0:
                    self.headers = [str(cell) if cell else "" for cell in row]
                else:
                    row_data = [str(cell) if cell else "" for cell in row]
                    if any(row_data):
                        self.preview_data.append(row_data)
            
            wb.close()
            
            self._show_preview_table()
            
            self.import_btn.setEnabled(True)
            self.import_status_label.setText(f"预览: {len(self.preview_data)} 行数据")
            
        except Exception as e:
            QMessageBox.warning(self, "错误", f"预览数据失败:\n{str(e)}")
            logging.error(f"预览数据失败: {e}")
    
    def _show_preview_table(self):
        self.data_table.clear()
        self.data_table.setColumnCount(len(self.headers))
        self.data_table.setHorizontalHeaderLabels(self.headers)
        self.data_table.setRowCount(len(self.preview_data))
        
        for row_idx, row_data in enumerate(self.preview_data):
            for col_idx, cell_value in enumerate(row_data):
                self.data_table.setItem(row_idx, col_idx, QTableWidgetItem(cell_value))
        
        self.data_table.resizeColumnsToContents()
    
    def _validate_data(self):
        if not self.preview_data:
            QMessageBox.warning(self, "提示", "请先预览数据")
            return
        
        if not self.current_template_id:
            QMessageBox.warning(self, "提示", "请先选择导入模板")
            return
        
        template, error = self.template_service.get_template(self.current_template_id)
        if error:
            QMessageBox.warning(self, "错误", f"获取模板失败: {error}")
            return
        
        import_field_mapping = template.get('import_field_mapping', {})
        if not import_field_mapping:
            QMessageBox.warning(self, "提示", "模板未配置字段映射")
            return
        
        required_fields = []
        for field_name, mapping in import_field_mapping.items():
            if mapping.get('required', False):
                required_fields.append(field_name)
        
        self.valid_rows, self.invalid_rows, error = self.import_service.validate_data(
            self.headers, self.preview_data, import_field_mapping, required_fields
        )
        
        if error:
            QMessageBox.warning(self, "校验结果", f"数据校验失败:\n{error}")
            return
        
        self._log(f"数据校验完成: 有效 {len(self.valid_rows)} 行, 无效 {len(self.invalid_rows)} 行")
        
        if self.invalid_rows:
            error_count = len(self.invalid_rows)
            QMessageBox.warning(
                self, "校验结果",
                f"数据校验完成:\n有效数据: {len(self.valid_rows)} 行\n无效数据: {error_count} 行\n\n是否继续导入有效数据?",
                QMessageBox.Yes | QMessageBox.No
            )
            if QMessageBox.Yes:
                self.import_btn.setEnabled(True)
        else:
            QMessageBox.information(self, "校验结果", f"数据校验通过，共 {len(self.valid_rows)} 行有效数据")
            self.import_btn.setEnabled(True)
        
        self.import_status_label.setText(f"校验: 有效 {len(self.valid_rows)} 行, 无效 {len(self.invalid_rows)} 行")
    
    def _import_data(self):
        if not self.valid_rows:
            QMessageBox.warning(self, "提示", "没有有效数据可导入")
            return
        
        if not self.current_template_id:
            QMessageBox.warning(self, "提示", "请先选择导入模板")
            return
        
        reply = QMessageBox.question(
            self, "确认导入",
            f"确认导入 {len(self.valid_rows)} 行数据到系统数据库?",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply != QMessageBox.Yes:
            return
        
        self._log("开始导入数据...")
        
        import_batch_id, error = self.import_service.import_to_database(
            self.valid_rows,
            self.current_template_id,
            self.excel_file_path,
            self.current_sheet,
            self.username
        )
        
        if error:
            QMessageBox.warning(self, "导入失败", f"导入数据失败:\n{error}")
            self._log(f"导入失败: {error}")
            return
        
        self.import_batch_id = import_batch_id
        self._log(f"导入成功, 批次ID: {import_batch_id}")
        
        QMessageBox.information(self, "导入成功", f"数据导入成功!\n批次ID: {import_batch_id}\n共导入 {len(self.valid_rows)} 行数据")
        
        self.import_status_label.setText(f"已导入: 批次 {import_batch_id}")
        
        self._load_batches()
        
        index = self.batch_combo.findData(import_batch_id)
        if index >= 0:
            self.batch_combo.setCurrentIndex(index)
        
        self.import_btn.setEnabled(False)
    
    def _load_templates(self):
        self.template_combo.clear()
        self.template_combo.addItem("请选择模板", None)
        
        templates, error = self.template_service.get_all_templates()
        if error:
            self._log(f"加载模板失败: {error}")
            return
        
        for template in templates:
            self.template_combo.addItem(template['template_name'], template['template_id'])
        
        self._log(f"已加载 {len(templates)} 个模板")
    
    def _on_template_changed(self, index):
        template_id = self.template_combo.currentData()
        if template_id:
            self.current_template_id = template_id
            template, error = self.template_service.get_template(template_id)
            if error:
                self._log(f"获取模板失败: {error}")
                return
            
            self.field_mapping = template.get('import_field_mapping', {})
            self._log(f"已选择模板: {template['template_name']}")
            
            if self.preview_data:
                self.import_btn.setEnabled(False)
                self.import_status_label.setText("请先校验数据")
        else:
            self.current_template_id = ""
            self.field_mapping = {}
    
    def _load_exe_configs(self):
        self.exe_config_combo.clear()
        self.exe_config_combo.addItem("请选择EXE", None)
        
        configs, error = self.exe_config_service.get_all_configs()
        if error:
            self._log(f"加载EXE配置失败: {error}")
            return
        
        for config in configs:
            self.exe_config_combo.addItem(config['config_name'], config['config_id'])
        
        self._log(f"已加载 {len(configs)} 个EXE配置")
    
    def _load_batches(self):
        self.batch_combo.clear()
        self.batch_combo.addItem("请选择批次", None)
        
        # 使用数据权限服务获取过滤后的批次
        batches = self.data_permission_service.get_filtered_batches(
            'rpa_import_batches', self.role_code, self.username,
            order_by="imported_at DESC"
        )
        
        # 限制显示数量
        batches = batches[:20]
        
        for batch in batches:
            display_text = f"{batch['import_name']} ({batch['total_count']}条) - {batch['imported_at'][:10] if batch['imported_at'] else ''}"
            self.batch_combo.addItem(display_text, batch['import_batch_id'])
        
        self._log(f"已加载 {len(batches)} 个导入批次")
    
    def _on_batch_changed(self, index):
        batch_id = self.batch_combo.currentData()
        if batch_id:
            self.import_batch_id = batch_id
            self._load_batch_data(batch_id)
    
    def _load_batch_data(self, batch_id):
        status_filter = self.status_filter_combo.currentData()
        keyword = self.keyword_search.text()
        
        rows, error = self.import_service.get_batch_data(batch_id, status_filter, keyword)
        if error:
            self._log(f"加载批次数据失败: {error}")
            return
        
        headers = ["行号", "业务主键", "数据状态", "RPA状态", "系统单号", "失败原因", "执行时间"]
        self.data_table.setColumnCount(len(headers))
        self.data_table.setHorizontalHeaderLabels(headers)
        self.data_table.setRowCount(len(rows))
        
        for row_idx, row in enumerate(rows):
            self.data_table.setItem(row_idx, 0, QTableWidgetItem(str(row['excel_row_number'])))
            self.data_table.setItem(row_idx, 1, QTableWidgetItem(str(row['business_key'] or '')))
            self.data_table.setItem(row_idx, 2, QTableWidgetItem(str(row['data_status'] or '')))
            self.data_table.setItem(row_idx, 3, QTableWidgetItem(str(row['rpa_status'] or '')))
            self.data_table.setItem(row_idx, 4, QTableWidgetItem(str(row['target_system_no'] or row['rpa_system_no'] or '')))
            self.data_table.setItem(row_idx, 5, QTableWidgetItem(str(row['rpa_error_message'] or row['validation_message'] or '')))
            self.data_table.setItem(row_idx, 6, QTableWidgetItem(str(row['last_executed_at'] or '')))
        
        self.data_table.resizeColumnsToContents()
        
        self._update_progress_stats(rows)
        
        self._log(f"已加载批次 {batch_id} 的 {len(rows)} 条数据")
    
    def _filter_data(self):
        if self.import_batch_id:
            self._load_batch_data(self.import_batch_id)
    
    def _update_progress_stats(self, rows):
        total = len(rows)
        pending = sum(1 for r in rows if r['rpa_status'] == 'pending' or not r['rpa_status'])
        success = sum(1 for r in rows if r['rpa_status'] == 'success')
        failed = sum(1 for r in rows if r['rpa_status'] == 'failed')
        skipped = sum(1 for r in rows if r['rpa_status'] == 'skipped')
        
        self.progress_label.setText(f"总行数: {total} | 待处理: {pending} | 成功: {success} | 失败: {failed} | 跳过: {skipped}")
        
        if total > 0:
            progress = int((success + skipped) / total * 100)
            self.progress_bar.setValue(progress)
    
    def _connect_exe(self):
        if not self.permission_checker.check_permission(PermissionCodes.OP_RPA_EXECUTE, self):
            return

        config_id = self.exe_config_combo.currentData()
        if not config_id:
            QMessageBox.warning(self, "提示", "请先选择目标EXE配置")
            return
        
        success, message = self.exe_config_service.test_connection(config_id)
        if success:
            QMessageBox.information(self, "连接成功", message)
            self._log(f"EXE连接成功: {message}")
        else:
            QMessageBox.warning(self, "连接失败", message)
            self._log(f"EXE连接失败: {message}")
    
    def _test_login(self):
        if not self.permission_checker.check_permission(PermissionCodes.OP_RPA_EXECUTE, self):
            return

        config_id = self.exe_config_combo.currentData()
        if not config_id:
            QMessageBox.warning(self, "提示", "请先选择目标EXE配置")
            return
        
        success, message = self.exe_config_service.test_launch(config_id)
        if success:
            QMessageBox.information(self, "启动成功", message)
            self._log(f"EXE启动成功: {message}")
        else:
            QMessageBox.warning(self, "启动失败", message)
            self._log(f"EXE启动失败: {message}")
    
    def _test_current_record(self):
        if not self.permission_checker.check_permission(PermissionCodes.OP_RPA_EXECUTE, self):
            return

        current_row = self.data_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "提示", "请先选择要测试的记录")
            return
        
        if not self.current_template_id:
            QMessageBox.warning(self, "提示", "请先选择业务模板")
            return
        
        exe_config_id = self.exe_config_combo.currentData()
        if not exe_config_id:
            QMessageBox.warning(self, "提示", "请先选择目标EXE配置")
            return
        
        template, error = self.template_service.get_template(self.current_template_id)
        if error:
            QMessageBox.warning(self, "错误", f"获取模板失败: {error}")
            return
        
        exe_config, error = self.exe_config_service.get_config(exe_config_id)
        if error:
            QMessageBox.warning(self, "错误", f"获取EXE配置失败: {error}")
            return
        
        self._log(f"开始测试当前记录...")
        
        success, message = self.auto_login.auto_login(exe_config)
        if not success:
            QMessageBox.warning(self, "登录失败", f"自动登录失败:\n{message}")
            self._log(f"登录失败: {message}")
            return
        
        self._log(f"登录成功: {message}")
        
        row_data = self._get_row_data(current_row)
        
        workflow_steps = []
        if template.get('workflow_steps'):
            for step_dict in template['workflow_steps']:
                from app.core.rpa_action_model import ActionStep
                workflow_steps.append(ActionStep.from_dict(step_dict))
        
        if not workflow_steps:
            QMessageBox.warning(self, "提示", "模板中没有定义工作流步骤")
            return
        
        screenshot_dir = Path("output/rpa") / self.current_task_id / "screenshots"
        screenshot_dir.mkdir(parents=True, exist_ok=True)
        
        success, result = self.executor.execute_workflow(
            workflow_steps, row_data, self.current_task_id, str(screenshot_dir)
        )
        
        if success:
            QMessageBox.information(self, "测试成功", f"测试执行成功")
            self._log(f"测试成功")
        else:
            QMessageBox.warning(self, "测试失败", f"测试执行失败:\n{result.get('error_message', '')}")
            self._log(f"测试失败: {result.get('error_message', '')}")
    
    def _start_execution(self):
        if not self.permission_checker.check_permission(PermissionCodes.OP_RPA_EXECUTE, self):
            return

        if self.is_executing:
            QMessageBox.warning(self, "提示", "任务正在执行中")
            return
        
        if not self.import_batch_id:
            QMessageBox.warning(self, "提示", "请先选择导入批次")
            return
        
        if not self.current_template_id:
            QMessageBox.warning(self, "提示", "请先选择业务模板")
            return
        
        exe_config_id = self.exe_config_combo.currentData()
        if not exe_config_id:
            QMessageBox.warning(self, "提示", "请先选择目标EXE配置")
            return
        
        template, error = self.template_service.get_template(self.current_template_id)
        if error:
            QMessageBox.warning(self, "错误", f"获取模板失败: {error}")
            return
        
        exe_config, error = self.exe_config_service.get_config(exe_config_id)
        if error:
            QMessageBox.warning(self, "错误", f"获取EXE配置失败: {error}")
            return
        
        rows, error = self._get_execution_rows()
        if error:
            QMessageBox.warning(self, "错误", f"获取待执行数据失败: {error}")
            return
        
        if len(rows) == 0:
            QMessageBox.information(self, "提示", "没有需要执行的数据")
            return
        
        self.current_task_id = uuid.uuid4().hex[:8]
        self.is_executing = True
        
        task_dir = Path("output/rpa") / self.current_task_id
        task_dir.mkdir(parents=True, exist_ok=True)
        screenshot_dir = task_dir / "screenshots"
        screenshot_dir.mkdir(parents=True, exist_ok=True)
        
        self._create_task_record(len(rows))
        
        self._log(f"任务创建成功，任务ID: {self.current_task_id}")
        self._log(f"待执行记录数: {len(rows)}")
        
        success, message = self.auto_login.auto_login(exe_config)
        if not success:
            self._log(f"登录失败: {message}")
            self._update_task_status("失败", f"登录失败: {message}")
            self.is_executing = False
            QMessageBox.warning(self, "登录失败", f"自动登录失败:\n{message}")
            return
        
        self._log(f"登录成功: {message}")
        
        workflow_steps = []
        if template.get('workflow_steps'):
            for step_dict in template['workflow_steps']:
                from app.core.rpa_action_model import ActionStep
                workflow_steps.append(ActionStep.from_dict(step_dict))
        
        if not workflow_steps:
            self._log("模板中没有定义工作流步骤")
            self._update_task_status("失败", "模板中没有定义工作流步骤")
            self.is_executing = False
            QMessageBox.warning(self, "提示", "模板中没有定义工作流步骤")
            return
        
        success_count = 0
        failed_count = 0
        
        for idx, row_data in enumerate(rows):
            if not self.is_executing:
                self._log("任务已停止")
                break
            
            row_id = row_data.get('import_row_id', '')
            business_key = row_data.get('business_key', '')
            
            self._log(f"第{idx+1}行开始执行，业务主键: {business_key}")
            
            self._create_row_record(row_id, idx+1, business_key)
            
            success, result = self.executor.execute_workflow(
                workflow_steps, row_data, self.current_task_id, str(screenshot_dir)
            )
            
            if success:
                success_count += 1
                system_no = result.get('saved_values', {}).get('system_no', '')
                self.executor.save_task_result(row_id, True, system_no, "执行成功")
                self._log(f"第{idx+1}行执行成功")
            else:
                failed_count += 1
                error_msg = result.get('error_message', '')
                screenshot_path = result.get('screenshot_path', '')
                self.executor.save_task_result(row_id, False, "", "", error_msg, screenshot_path)
                self._log(f"第{idx+1}行执行失败: {error_msg}")
            
            self._update_progress()
        
        self.is_executing = False
        
        self._update_task_status("完成", f"成功:{success_count}, 失败:{failed_count}")
        self._log(f"任务执行完成，成功:{success_count}, 失败:{failed_count}")
        
        QMessageBox.information(self, "执行完成", 
            f"任务执行完成\n成功: {success_count}\n失败: {failed_count}")
    
    def _pause_execution(self):
        self.is_executing = False
        self._log("任务已暂停")
    
    def _continue_execution(self):
        if not self.current_task_id:
            QMessageBox.warning(self, "提示", "没有暂停的任务")
            return
        
        self._log("任务继续执行")
        self._start_execution()
    
    def _stop_execution(self):
        self.is_executing = False
        self._log("任务已停止")
    
    def _retry_failed(self):
        if not self.permission_checker.check_permission(PermissionCodes.OP_RPA_EXECUTE, self):
            return

        self.exec_mode_combo.setCurrentIndex(2)
        self._start_execution()
    
    def _export_result(self):
        if not self.permission_checker.check_permission(PermissionCodes.OP_RPA_EXECUTE, self):
            return

        if not self.import_batch_id:
            QMessageBox.warning(self, "提示", "请先选择导入批次")
            return
        
        default_name = f"RPA结果_{self.import_batch_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "导出结果Excel",
            default_name,
            "Excel文件 (*.xlsx)"
        )
        
        if not file_path:
            return
        
        self._log(f"开始导出结果到: {file_path}")
        
        output_path, error = self.import_service.export_result_excel(self.import_batch_id, file_path)
        
        if error:
            QMessageBox.warning(self, "导出失败", f"导出结果失败:\n{error}")
            self._log(f"导出失败: {error}")
            return
        
        QMessageBox.information(self, "导出成功", f"结果已导出到:\n{output_path}")
        self._log(f"导出成功: {output_path}")
    
    def _get_row_data(self, row_idx: int) -> Dict:
        row_data = {}
        for col_idx in range(self.data_table.columnCount()):
            header = self.data_table.horizontalHeaderItem(col_idx).text()
            item = self.data_table.item(row_idx, col_idx)
            value = item.text() if item else ""
            row_data[header] = value
        return row_data
    
    def _get_execution_rows(self) -> Tuple[List[Dict], str]:
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            exec_mode = self.exec_mode_combo.currentData()
            
            if exec_mode == "all":
                cursor.execute('''
                    SELECT import_row_id, business_key, normalized_data
                    FROM rpa_import_details
                    WHERE import_batch_id = ?
                    ORDER BY excel_row_number
                ''', (self.import_batch_id,))
            elif exec_mode == "pending_only":
                cursor.execute('''
                    SELECT import_row_id, business_key, normalized_data
                    FROM rpa_import_details
                    WHERE import_batch_id = ? AND rpa_status = '待处理'
                    ORDER BY excel_row_number
                ''', (self.import_batch_id,))
            elif exec_mode == "retry_failed":
                cursor.execute('''
                    SELECT import_row_id, business_key, normalized_data
                    FROM rpa_import_details
                    WHERE import_batch_id = ? AND rpa_status = '失败'
                    ORDER BY excel_row_number
                ''', (self.import_batch_id,))
            else:
                start_row = self.start_row_spin.value()
                max_rows = self.max_rows_spin.value()
                
                if max_rows > 0:
                    cursor.execute('''
                        SELECT import_row_id, business_key, normalized_data
                        FROM rpa_import_details
                        WHERE import_batch_id = ?
                        ORDER BY excel_row_number
                        LIMIT ? OFFSET ?
                    ''', (self.import_batch_id, max_rows, start_row - 1))
                else:
                    cursor.execute('''
                        SELECT import_row_id, business_key, normalized_data
                        FROM rpa_import_details
                        WHERE import_batch_id = ?
                        ORDER BY excel_row_number
                        OFFSET ?
                    ''', (self.import_batch_id, start_row - 1))
            
            rows = cursor.fetchall()
            
            result_rows = []
            for row in rows:
                row_dict = dict(row)
                if row_dict.get('normalized_data'):
                    try:
                        import json
                        normalized_data = json.loads(row_dict['normalized_data'])
                        row_dict.update(normalized_data)
                    except:
                        pass
                result_rows.append(row_dict)
            
            return result_rows, ""
            
        except Exception as e:
            logging.error(f"获取待执行数据失败: {e}")
            return [], str(e)
    
    def _create_task_record(self, total_count: int):
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            now = datetime.now().isoformat()
            
            cursor.execute('''
                INSERT INTO rpa_tasks
                (task_id, task_type, template_id, import_batch_id,
                 exe_config_id, status, total_count, started_at, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                self.current_task_id,
                'batch',
                self.current_template_id,
                self.import_batch_id,
                self.exe_config_combo.currentData(),
                '执行中',
                total_count,
                now,
                now
            ))
            
            conn.commit()
            
        except Exception as e:
            logging.error(f"创建任务记录失败: {e}")
    
    def _create_row_record(self, import_row_id: str, row_number: int, business_key: str):
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            now = datetime.now().isoformat()
            
            row_id = uuid.uuid4().hex[:8]
            
            cursor.execute('''
                INSERT INTO rpa_task_rows
                (row_id, task_id, import_row_id, excel_row_number,
                 business_key, status, started_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (
                row_id,
                self.current_task_id,
                import_row_id,
                row_number,
                business_key,
                '处理中',
                now
            ))
            
            conn.commit()
            
        except Exception as e:
            logging.error(f"创建行记录失败: {e}")
    
    def _update_task_status(self, status: str, message: str):
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            now = datetime.now().isoformat()
            
            cursor.execute('''
                UPDATE rpa_tasks
                SET status = ?, error_message = ?, finished_at = ?
                WHERE task_id = ?
            ''', (status, message, now, self.current_task_id))
            
            conn.commit()
            
        except Exception as e:
            logging.error(f"更新任务状态失败: {e}")
    
    def _update_progress(self):
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT 
                    COUNT(*) as total,
                    SUM(CASE WHEN status = '成功' THEN 1 ELSE 0 END) as success,
                    SUM(CASE WHEN status = '失败' THEN 1 ELSE 0 END) as failed,
                    SUM(CASE WHEN status = '跳过' THEN 1 ELSE 0 END) as skipped
                FROM rpa_task_rows
                WHERE task_id = ?
            ''', (self.current_task_id,))
            
            row = cursor.fetchone()
            if row:
                total = row['total']
                success = row['success']
                failed = row['failed']
                skipped = row['skipped']
                
                self.total_label.setText(f"总行数: {total}")
                self.success_label.setText(f"成功: {success}")
                self.failed_label.setText(f"失败: {failed}")
                self.skipped_label.setText(f"跳过: {skipped}")
                
                if total > 0:
                    progress = int((success + skipped) / total * 100)
                    self.progress_bar.setValue(progress)
                
        except Exception as e:
            logging.error(f"更新进度失败: {e}")
    
    def _log(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.append(f"{timestamp} {message}")
        logging.info(f"RPA: {message}")