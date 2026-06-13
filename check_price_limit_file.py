"""检查医保价格上限Excel文件"""

import openpyxl
import glob

# 查找价格上限Excel文件
excel_files = glob.glob("D:/project/RPA/*.xlsx")
price_limit_files = [f for f in excel_files if "价格" in f or "上限" in f or "三同" in f]

print(f"找到的价格上限相关文件: {price_limit_files}")

if price_limit_files:
    file_path = price_limit_files[0]
    print(f"\n检查文件: {file_path}")
    
    workbook = openpyxl.load_workbook(file_path, read_only=True)
    print(f"工作表列表: {workbook.sheetnames}")
    
    for sheet_name in workbook.sheetnames[:2]:
        ws = workbook[sheet_name]
        print(f"\n工作表: {sheet_name}")
        print(f"最大行数: {ws.max_row}")
        
        # 打印前5行
        for row_num in range(1, min(6, ws.max_row + 1)):
            row_data = []
            for cell in ws[row_num]:
                value = str(cell.value or "").strip()
                row_data.append(value)
            
            print(f"第 {row_num} 行: {row_data[:10]}")
            
            # 检查是否包含关键字段
            key_fields = ["医保编码", "药品名称", "企业名称", "医保价格上限", "价格生效日期"]
            matched_count = sum(1 for field in key_fields if field in row_data)
            if matched_count > 0:
                print(f"  -> 包含 {matched_count} 个关键字段")
    
    workbook.close()
else:
    print("未找到价格上限Excel文件")