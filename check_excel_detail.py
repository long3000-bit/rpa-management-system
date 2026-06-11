import openpyxl
import pandas as pd

file_path = 'D:/project/RPA/商品信息维护20260608162511.xlsx'

# 使用openpyxl检查
wb = openpyxl.load_workbook(file_path, read_only=True)
print('=== openpyxl 检查 ===')
print('工作表列表:', wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n工作表: {sheet_name}')
    print(f'最大行数: {ws.max_row}')
    print(f'最大列数: {ws.max_column}')
    
    print('\n前10行内容:')
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10), 1):
        values = []
        for cell in row:
            val = cell.value
            if val is not None:
                values.append(f'{cell.column_letter}:{val}')
        print(f'Row {i}: {values}')

wb.close()

# 使用pandas检查
print('\n=== pandas 检查 ===')
try:
    df = pd.read_excel(file_path, sheet_name=None)
    for sheet_name, sheet_df in df.items():
        print(f'\n工作表: {sheet_name}')
        print(f'行数: {len(sheet_df)}')
        print(f'列数: {len(sheet_df.columns)}')
        print(f'列名: {list(sheet_df.columns)}')
        print(f'\n前5行数据:')
        print(sheet_df.head())
except Exception as e:
    print(f'pandas读取失败: {e}')